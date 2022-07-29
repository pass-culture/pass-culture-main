import csv
from pprint import PrettyPrinter
from typing import Iterable

import openpyxl
import sqlalchemy.exc

from pcapi.core.booking_providers.models import BookingProvider
from pcapi.core.booking_providers.models import BookingProviderName
from pcapi.core.booking_providers.models import VenueBookingProvider
from pcapi.core.offerers.models import Venue
from pcapi.core.providers.models import CDSCinemaDetails
from pcapi.core.providers.models import CinemaProviderPivot
from pcapi.core.providers.models import Provider
from pcapi.models import db


class CDSCinema:
    def __init__(self, account_id: str, cinema_id: str, cds_token: str, venue_siret: str) -> None:
        self.account_id = account_id
        self.cinema_id = cinema_id
        self.cds_token = cds_token
        self.venue_siret = venue_siret


def add_all_cinema_pivot_from_file(
    file_path: str,
    file_ext: str = "excel",
    starting_excel_row_number_to_import: int = 2,
    ending_excel_row_number_to_import: int = 2,
    dry_run: bool = True,
) -> None:
    cds_provider = Provider.query.filter(Provider.localClass == "CDSStocks").one_or_none()
    booking_provider = BookingProvider.query.filter(
        BookingProvider.name == BookingProviderName.CINE_DIGITAL_SERVICE
    ).one_or_none()
    if cds_provider and booking_provider:

        if file_ext == "excel":
            print("read excel file")
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            cinemas_to_add = get_all_cinemas_to_add_from_excel(
                sheet, starting_excel_row_number_to_import, ending_excel_row_number_to_import
            )
            print(f"{len(cinemas_to_add)} cinemas to add")

        else:
            print("read csv file")
            with open(file_path, mode="r", newline="", encoding="utf-8") as file:
                csv_reader = csv.DictReader(file)
                cinemas_to_add = get_all_cinemas_to_add_from_csv(csv_reader)
                print(f"{len(cinemas_to_add)} cinemas to add")

        for cinema in cinemas_to_add:
            add_pivot_and_booking_provider_for_cinema(cinema, cds_provider, booking_provider)

        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()
    else:
        print("CDS Provider not found")


def add_pivot_and_booking_provider_for_cinema(
    cinema: CDSCinema, provider: Provider, booking_provider: BookingProvider
) -> None:
    print(f"Adding Pivot for {cinema.account_id} - {cinema.cinema_id} - {cinema.venue_siret} -  ...")
    venue = Venue.query.filter(Venue.siret == cinema.venue_siret).one_or_none()
    if venue:
        try:
            with db.session.begin_nested():
                pivot = CinemaProviderPivot(venue=venue, provider=provider, idAtProvider=cinema.cinema_id)
                cds_details = CDSCinemaDetails(
                    cinemaProviderPivot=pivot, cinemaApiToken=cinema.cds_token, accountId=cinema.account_id
                )
                venue_booking_provider = VenueBookingProvider(
                    isActive=True,
                    venue=venue,
                    bookingProvider=booking_provider,
                    idAtProvider=cinema.cinema_id,
                    token=cinema.cds_token,
                )
                db.session.add(pivot)
                db.session.add(cds_details)
                db.session.add(venue_booking_provider)
                db.session.flush()
        except sqlalchemy.exc.IntegrityError as e:
            print("Ignoring cinema: %s because %s" % (str(cinema.cinema_id), e.orig.diag.message_detail))
        print(f"{cinema.cinema_id} added successfully")

    else:
        print(f"No venue found for cinema : {cinema.cinema_id}")


def get_all_cinemas_to_add_from_csv(rows: Iterable[dict]) -> list[CDSCinema]:
    cinemas_to_add = []
    for row in rows:
        cinema = get_cds_cinema_from_csv_row(row)
        if isValidSiret(cinema.venue_siret) and isValidToken(cinema.cds_token):
            cinemas_to_add.append(cinema)
        else:
            print(f"Ignoring {cinema.account_id} - {cinema.cinema_id} : incorrect siret or token")

    return cinemas_to_add


def get_cds_cinema_from_csv_row(row: dict) -> CDSCinema:
    return CDSCinema(
        account_id=row["compte"],
        cinema_id=row["cinemaid"],
        cds_token=row["vadtoken"],
        venue_siret=row["SIRET (manuel)"],
    )


def get_all_cinemas_to_add_from_excel(
    sheet: openpyxl.Workbook.active, starting_row_number_to_import: int, ending_row_number_to_import: int
) -> list[CDSCinema]:

    # Create a dictionary of column names
    column_names = {}
    current_column_index = 0
    for column in sheet.iter_cols(1, sheet.max_column):
        column_names[column[0].value] = current_column_index
        current_column_index += 1

    # access by column name
    # (My data has a column named 'Dogs')
    cinemas_in_error = []
    cinemas_to_add = []
    for row_cells in sheet.iter_rows(min_row=starting_row_number_to_import, max_row=ending_row_number_to_import):
        cinema = CDSCinema(
            account_id=row_cells[column_names["compte"]].value,
            cinema_id=row_cells[column_names["cinemaid"]].value,
            cds_token=row_cells[column_names["vadtoken"]].value,
            venue_siret=str(row_cells[column_names["SIRET (manuel)"]].value),
        )

        valid_siret = isValidSiret(cinema.venue_siret)
        valid_token = isValidToken(cinema.cds_token)

        if not valid_siret and not valid_token:
            error = f"Incorrect siret and token for account {cinema.account_id} - cinema {cinema.cinema_id}"
            cinemas_in_error.append(error)
        elif not valid_siret:
            error = f"Incorrect siret for account {cinema.account_id} - cinema {cinema.cinema_id} "
            cinemas_in_error.append(error)
        elif not valid_token:
            error = f"Incorrect token for account {cinema.account_id} - cinema {cinema.cinema_id}"
            cinemas_in_error.append(error)

        if valid_siret and valid_token:
            cinemas_to_add.append(cinema)

    print(f"{len(cinemas_in_error)} cinemas with incorrect siret or/and cds_token")
    pp = PrettyPrinter(width=200)
    pp.pprint(cinemas_in_error)

    return cinemas_to_add


def get_cds_cinema_from_excel_row(row: dict) -> CDSCinema:
    return CDSCinema(
        account_id=row["compte"],
        cinema_id=row["cinemaid"],
        cds_token=row["vadtoken"],
        venue_siret=row["SIRET (manuel)"],
    )


def isValidSiret(siret: str) -> bool:
    return len(siret) == 14 and siret.isdigit()


def isValidToken(token: str) -> bool:
    return token not in ["", "null"]
