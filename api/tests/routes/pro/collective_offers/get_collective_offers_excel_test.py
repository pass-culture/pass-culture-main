import datetime
import io
from operator import attrgetter

import openpyxl
import pytest
import pytz

from pcapi.core import testing
from pcapi.core.educational import factories
from pcapi.core.educational import models
from pcapi.core.educational.utils import format_collective_offer_displayed_status
from pcapi.core.offerers import factories as offerers_factories
from pcapi.core.testing import assert_num_queries
from pcapi.core.users.factories import ProFactory


pytestmark = pytest.mark.usefixtures("db_session")

EXPECTED_HEADER = [
    "Nom de l'offre",
    "Numéro de l'offre",
    "Statut de l'offre",
    "Type de localisation de l'offre",
    "Localisation de l'offre",
    "Structure",
    "Etablissement",
    "Code postal de l'établissement",
    "UAI de l'établissement",
    "Email de l'enseignant",
    "Prénom de l'enseignant",
    "Nom de l'enseignant",
    "Date de début de l'évènement",
    "Date de fin de l'évènement",
    "Prix",
    "Nombre de participants",
    "Date de préréservation de l'offre",
    "Date de réservation de l'offre",
    "Date de remboursement",
]


def _reader_from_response(response):
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    sheet = wb.active

    for i, header in enumerate(EXPECTED_HEADER):
        assert sheet.cell(row=1, column=i + 1).value == header

    return sheet


def _format_date(date_time, tz):
    return date_time.replace(tzinfo=datetime.timezone.utc).astimezone(tz).isoformat()


class Returns200Test:
    num_queries = testing.AUTHENTICATION_QUERIES
    num_queries += 1  # select collective_offer ids
    num_queries += 1  # select collective_offer
    num_queries_with_bookings = num_queries + 1  # selectinload collective_booking

    def test_one_offer(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer)
        offer = factories.ReimbursedCollectiveOfferFactory(
            venue=venue, locationType=models.CollectiveLocationType.ADDRESS, offererAddress=venue.offererAddress
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries_with_bookings):
            response = client.get("/collective/offers/excel")
            assert response.status_code == 200

        venue = offer.venue
        stock = offer.collectiveStock
        [booking] = stock.collectiveBookings
        tz = pytz.timezone(offer.offererAddress.address.timezone)

        reader = _reader_from_response(response)
        row_values = [reader.cell(row=2, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert row_values == [
            offer.name,
            offer.id,
            format_collective_offer_displayed_status(offer.displayedStatus).lower(),
            "À une adresse précise",
            offer.offererAddress.address.fullAddress,
            venue.common_name,
            offer.institution.full_name,
            offer.institution.postalCode,
            offer.institution.institutionId,
            booking.educationalRedactor.email,
            booking.educationalRedactor.firstName,
            booking.educationalRedactor.lastName,
            _format_date(stock.startDatetime, tz),
            _format_date(stock.endDatetime, tz),
            stock.price,
            stock.numberOfTickets,
            _format_date(booking.dateCreated, tz),
            _format_date(booking.confirmationDate, tz),
            _format_date(booking.reimbursementDate, tz),
        ]

    def test_one_offer_draft(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer)
        offer = factories.DraftCollectiveOfferFactory(
            name="Special name &@🤡%",
            venue=venue,
            locationType=models.CollectiveLocationType.TO_BE_DEFINED,
            locationComment="Chez toi",
        )
        offer.institution = None

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get("/collective/offers/excel")
            assert response.status_code == 200

        reader = _reader_from_response(response)
        row_values = [reader.cell(row=2, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert row_values == [
            "Special name &@🤡%",
            offer.id,
            format_collective_offer_displayed_status(offer.displayedStatus).lower(),
            "À déterminer avec l'enseignant",
            "Chez toi",
            venue.common_name,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]

    def test_multiple_offers(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        offers = factories.BookedCollectiveOfferFactory.create_batch(
            3, venue__managingOfferer=user_offerer.offerer, locationType=models.CollectiveLocationType.SCHOOL
        )
        offers = sorted(offers, key=attrgetter("dateCreated"), reverse=True)

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries_with_bookings):
            response = client.get("/collective/offers/excel")
            assert response.status_code == 200

        reader = _reader_from_response(response)

        for i, offer in enumerate(offers):
            venue = offer.venue
            stock = offer.collectiveStock
            [booking] = stock.collectiveBookings
            tz = pytz.timezone(venue.offererAddress.address.timezone)

            row_values = [
                reader.cell(row=i + 2, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))
            ]
            assert row_values == [
                offer.name,
                offer.id,
                "réservée",
                "En établissement scolaire",
                None,
                venue.common_name,
                offer.institution.full_name,
                offer.institution.postalCode,
                offer.institution.institutionId,
                booking.educationalRedactor.email,
                booking.educationalRedactor.firstName,
                booking.educationalRedactor.lastName,
                _format_date(stock.startDatetime, tz),
                _format_date(stock.endDatetime, tz),
                stock.price,
                stock.numberOfTickets,
                _format_date(booking.dateCreated, tz),
                _format_date(booking.confirmationDate, tz),
                None,
            ]

    def test_user_no_result(self, client):
        factories.PublishedCollectiveOfferFactory()
        user = ProFactory()

        client = client.with_session_auth(user.email)
        # no request to get collective offers as there are no ids
        with assert_num_queries(self.num_queries - 1):
            response = client.get("/collective/offers/excel")
            assert response.status_code == 200

        reader = _reader_from_response(response)
        row_values = [reader.cell(row=2, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert all(value is None for value in row_values)

    def test_filter_status(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer)
        offer = factories.PublishedCollectiveOfferFactory(venue=venue)
        _offer_not_in_result = factories.PrebookedCollectiveOfferFactory(venue=venue)

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries_with_bookings):
            response = client.get("/collective/offers/excel?status=PUBLISHED")
            assert response.status_code == 200

        reader = _reader_from_response(response)
        assert reader.cell(row=2, column=1).value == offer.name
        assert reader.cell(row=2, column=2).value == offer.id
        row_values = [reader.cell(row=3, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert all(value is None for value in row_values)

    def test_filter_date_range(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer)
        offer = factories.PublishedCollectiveOfferFactory(venue=venue)
        offer.collectiveStock.startDatetime = datetime.datetime.fromisoformat("2020-01-05")
        offer_not_in_result = factories.PublishedCollectiveOfferFactory(venue=venue)
        offer_not_in_result.collectiveStock.startDatetime = datetime.datetime.fromisoformat("2020-02-05")

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries_with_bookings):
            response = client.get("/collective/offers/excel?periodBeginningDate=2020-01-01&periodEndingDate=2020-02-01")
            assert response.status_code == 200

        reader = _reader_from_response(response)
        assert reader.cell(row=2, column=1).value == offer.name
        assert reader.cell(row=2, column=2).value == offer.id
        row_values = [reader.cell(row=3, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert all(value is None for value in row_values)

    def test_filter_offerer_id(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer)
        offer = factories.PublishedCollectiveOfferFactory(venue=venue)
        user_offerer_not_in_result = offerers_factories.UserOffererFactory(user=user_offerer.user)
        _offer_not_in_result = factories.PublishedCollectiveOfferFactory(
            venue__managingOfferer=user_offerer_not_in_result.offerer
        )

        client = client.with_session_auth(user_offerer.user.email)
        offerer_id = user_offerer.offerer.id
        with assert_num_queries(self.num_queries_with_bookings):
            response = client.get(f"/collective/offers/excel?offererId={offerer_id}")
            assert response.status_code == 200

        reader = _reader_from_response(response)
        assert reader.cell(row=2, column=1).value == offer.name
        assert reader.cell(row=2, column=2).value == offer.id
        row_values = [reader.cell(row=3, column=col_number + 1).value for col_number in range(len(EXPECTED_HEADER))]
        assert all(value is None for value in row_values)
