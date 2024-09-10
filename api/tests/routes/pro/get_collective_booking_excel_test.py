from datetime import datetime
from datetime import timedelta
from io import BytesIO

import openpyxl
import pytest

from pcapi.core.educational import factories as educational_factories
import pcapi.core.offerers.factories as offerers_factories
from pcapi.core.testing import assert_num_queries
from pcapi.core.testing import override_features
from pcapi.routes.serialization.collective_bookings_serialize import COLLECTIVE_BOOKING_EXPORT_HEADER


pytestmark = pytest.mark.usefixtures("db_session")


def reader_from_response(response):
    wb = openpyxl.load_workbook(BytesIO(response.data))
    return wb.active


class Returns200Test:
    num_queries = 1  # Fetch the session
    num_queries += 1  # Fetch the user
    num_queries += 1  # Fetch the collective_booking
    num_queries += 1  # Check if WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE is active

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_complete_booking_single(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Bob",
            educationalRedactor__lastName="Kelso",
            educationalInstitution=educational_factories.EducationalInstitutionFactory(),
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Celles
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Bob"
        assert sheet.cell(row=2, column=5).value == "Kelso"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=12).value == booking.educationalInstitution.institutionId
        assert (
            sheet.cell(row=2, column=13).value
            == f"{booking.educationalInstitution.institutionType} {booking.educationalInstitution.name}"
        )

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_created_booking_single(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=None,
            dateUsed=None,
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Perry",
            educationalRedactor__lastName="Cox",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Perry"
        assert sheet.cell(row=2, column=5).value == "Cox"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "None"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "None"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_one_invisible_rights_booking(self, client):
        invisible_user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CollectiveBookingFactory(
            collectiveStock__collectiveOffer__venue__managingOfferer=invisible_user_offerer.offerer,
            offerer=invisible_user_offerer.offerer,
        )
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="John Michael",
            educationalRedactor__lastName="Dorian",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "John Michael"
        assert sheet.cell(row=2, column=5).value == "Dorian"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"

        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=3, column=i).value is None

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_one_invisible_date_range_booking(self, client):
        invisible_user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2015, 8, 11, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=invisible_user_offerer.offerer,
            offerer=invisible_user_offerer.offerer,
        )
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Eliot",
            educationalRedactor__lastName="Reid",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2015-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Eliot"
        assert sheet.cell(row=2, column=5).value == "Reid"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_complete_booking_multiple(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        bookings = [
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Carla",
                educationalRedactor__lastName="Espinosa",
            ),
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Janitor",
                educationalRedactor__lastName="The",
            ),
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Ted",
                educationalRedactor__lastName="Buckland",
            ),
        ]

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        for i in range(2, 5):
            assert sheet.cell(row=i, column=1).value == bookings[i - 2].venue.name
            assert sheet.cell(row=i, column=2).value == bookings[i - 2].collectiveStock.collectiveOffer.name
            assert sheet.cell(row=i, column=3).value == "2020-08-13 14:00:00+02:00"
            assert sheet.cell(row=i, column=4).value == bookings[i - 2].educationalRedactor.firstName
            assert sheet.cell(row=i, column=5).value == bookings[i - 2].educationalRedactor.lastName
            assert sheet.cell(row=i, column=6).value == bookings[i - 2].educationalRedactor.email
            assert sheet.cell(row=i, column=7).value == "2020-08-11 14:00:00+02:00"
            assert sheet.cell(row=i, column=8).value == "2020-08-15 14:00:00+02:00"
            assert sheet.cell(row=i, column=9).value == bookings[i - 2].collectiveStock.price
            assert sheet.cell(row=2, column=10).value == "réservé"
            assert sheet.cell(row=i, column=11).value == "2021-08-11 14:00:00+02:00"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_booking_status_when_used(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.UsedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "validé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_booking_status_when_reimbursed(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.ReimbursedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "remboursé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_booking_status_when_cancelled(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CancelledCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "annulé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_booking_status_when_pending(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.PendingCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "préréservé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=False)
    def test_booking_status_when_confirmed(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.ConfirmedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "confirmé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_complete_booking_single_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Bob",
            educationalRedactor__lastName="Kelso",
            educationalInstitution=educational_factories.EducationalInstitutionFactory(),
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Celles
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Bob"
        assert sheet.cell(row=2, column=5).value == "Kelso"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=12).value == booking.educationalInstitution.institutionId
        assert (
            sheet.cell(row=2, column=13).value
            == f"{booking.educationalInstitution.institutionType} {booking.educationalInstitution.name}"
        )

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_created_booking_single_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=None,
            dateUsed=None,
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Perry",
            educationalRedactor__lastName="Cox",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Perry"
        assert sheet.cell(row=2, column=5).value == "Cox"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "None"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "None"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_one_invisible_rights_booking_with_offerer_address_as_data_source(self, client):
        invisible_user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CollectiveBookingFactory(
            collectiveStock__collectiveOffer__venue__managingOfferer=invisible_user_offerer.offerer,
            offerer=invisible_user_offerer.offerer,
        )
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="John Michael",
            educationalRedactor__lastName="Dorian",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "John Michael"
        assert sheet.cell(row=2, column=5).value == "Dorian"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"

        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=3, column=i).value is None

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_one_invisible_date_range_booking_with_offerer_address_as_data_source(self, client):
        invisible_user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2015, 8, 11, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=invisible_user_offerer.offerer,
            offerer=invisible_user_offerer.offerer,
        )
        user_offerer = offerers_factories.UserOffererFactory()
        booking = educational_factories.CollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
            educationalRedactor__firstName="Eliot",
            educationalRedactor__lastName="Reid",
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2015-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        # Headers
        for i in range(1, len(COLLECTIVE_BOOKING_EXPORT_HEADER)):
            assert sheet.cell(row=1, column=i).value == COLLECTIVE_BOOKING_EXPORT_HEADER[i - 1]
        # Cells
        assert sheet.cell(row=2, column=1).value == booking.venue.name
        assert sheet.cell(row=2, column=2).value == booking.collectiveStock.collectiveOffer.name
        assert sheet.cell(row=2, column=3).value == "2020-08-13 14:00:00+02:00"
        assert sheet.cell(row=2, column=4).value == "Eliot"
        assert sheet.cell(row=2, column=5).value == "Reid"
        assert sheet.cell(row=2, column=6).value == booking.educationalRedactor.email
        assert sheet.cell(row=2, column=7).value == "2020-08-11 14:00:00+02:00"
        assert sheet.cell(row=2, column=8).value == "2020-08-15 14:00:00+02:00"
        assert sheet.cell(row=2, column=9).value == booking.collectiveStock.price
        assert sheet.cell(row=2, column=10).value == "réservé"
        assert sheet.cell(row=2, column=11).value == "2021-08-11 14:00:00+02:00"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_complete_booking_multiple_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        bookings = [
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Carla",
                educationalRedactor__lastName="Espinosa",
            ),
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Janitor",
                educationalRedactor__lastName="The",
            ),
            educational_factories.CollectiveBookingFactory(
                dateCreated=datetime(2020, 8, 11, 12, 0, 0),
                collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
                cancellationLimitDate=datetime.utcnow() + timedelta(days=1),
                reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
                dateUsed=datetime(2020, 8, 15, 12, 0, 0),
                collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
                offerer=user_offerer.offerer,
                educationalRedactor__firstName="Ted",
                educationalRedactor__lastName="Buckland",
            ),
        ]

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        for i in range(2, 5):
            assert sheet.cell(row=i, column=1).value == bookings[i - 2].venue.name
            assert sheet.cell(row=i, column=2).value == bookings[i - 2].collectiveStock.collectiveOffer.name
            assert sheet.cell(row=i, column=3).value == "2020-08-13 14:00:00+02:00"
            assert sheet.cell(row=i, column=4).value == bookings[i - 2].educationalRedactor.firstName
            assert sheet.cell(row=i, column=5).value == bookings[i - 2].educationalRedactor.lastName
            assert sheet.cell(row=i, column=6).value == bookings[i - 2].educationalRedactor.email
            assert sheet.cell(row=i, column=7).value == "2020-08-11 14:00:00+02:00"
            assert sheet.cell(row=i, column=8).value == "2020-08-15 14:00:00+02:00"
            assert sheet.cell(row=i, column=9).value == bookings[i - 2].collectiveStock.price
            assert sheet.cell(row=2, column=10).value == "réservé"
            assert sheet.cell(row=i, column=11).value == "2021-08-11 14:00:00+02:00"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_booking_status_when_used_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.UsedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "validé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_booking_status_when_reimbursed_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.ReimbursedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            reimbursementDate=datetime(2021, 8, 11, 12, 0, 0),
            dateUsed=datetime(2020, 8, 15, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "remboursé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_booking_status_when_cancelled_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.CancelledCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "annulé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_booking_status_when_pending_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.PendingCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "préréservé"

    @override_features(WIP_USE_OFFERER_ADDRESS_AS_DATA_SOURCE=True)
    def test_booking_status_when_confirmed_with_offerer_address_as_data_source(self, client):
        user_offerer = offerers_factories.UserOffererFactory()
        educational_factories.ConfirmedCollectiveBookingFactory(
            dateCreated=datetime(2020, 8, 11, 12, 0, 0),
            collectiveStock__beginningDatetime=datetime(2020, 8, 13, 12, 0, 0),
            collectiveStock__collectiveOffer__venue__managingOfferer=user_offerer.offerer,
            offerer=user_offerer.offerer,
        )

        client = client.with_session_auth(user_offerer.user.email)
        with assert_num_queries(self.num_queries):
            response = client.get(
                "/collective/bookings/excel?bookingPeriodBeginningDate=2000-01-01&bookingPeriodEndingDate=2030-01-01"
            )
            assert response.status_code == 200

        sheet = reader_from_response(response)
        assert sheet.cell(row=2, column=10).value == "confirmé"
