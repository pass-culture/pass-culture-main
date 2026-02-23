import dataclasses
import datetime
import re
from io import BytesIO

import factory
import openpyxl
import pytest
from flask import url_for

from pcapi.core.bookings import factories as bookings_factories
from pcapi.core.bookings import models as bookings_models
from pcapi.core.categories import pro_categories
from pcapi.core.categories import subcategories
from pcapi.core.external_bookings import factories as external_bookings_factories
from pcapi.core.finance import factories as finance_factories
from pcapi.core.finance import models as finance_models
from pcapi.core.geography import factories as geography_factories
from pcapi.core.mails import testing as mails_testing
from pcapi.core.mails.transactional.sendinblue_template_ids import TransactionalEmail
from pcapi.core.offerers import factories as offerers_factories
from pcapi.core.offers import factories as offers_factories
from pcapi.core.offers import models as offers_models
from pcapi.core.permissions import models as perm_models
from pcapi.core.providers import factories as providers_factories
from pcapi.core.providers.repository import get_provider_by_local_class
from pcapi.core.testing import assert_num_queries
from pcapi.core.users import factories as users_factories
from pcapi.models import db
from pcapi.routes.backoffice.bookings import forms
from pcapi.utils import date as date_utils

from tests.connectors.cgr import soap_definitions
from tests.local_providers.cinema_providers.cgr import fixtures as cgr_fixtures

from .helpers import flash
from .helpers import html_parser
from .helpers.get import GetEndpointHelper
from .helpers.post import PostEndpointHelper


pytestmark = [
    pytest.mark.usefixtures("db_session"),
    pytest.mark.backoffice,
]


@pytest.fixture(scope="function", name="bookings")
def bookings_fixture() -> tuple:
    user1 = users_factories.BeneficiaryFactory(firstName="Napoléon", lastName="Bonaparte", email="napo@leon.com")
    user2 = users_factories.BeneficiaryFactory(age=17, firstName="Joséphine", lastName="de Beauharnais")
    user3 = users_factories.BeneficiaryFactory(age=17, firstName="Marie-Louise", lastName="d'Autriche")
    user4 = users_factories.BeneficiaryFactory(age=17, firstName="Marie-Louise", lastName="d'Autriche")
    used = bookings_factories.UsedBookingFactory(
        id=1000001,  # Avoid flaky test because of same ids in bookings and offers
        user=user1,
        quantity=2,
        token="WTRL00",
        stock__price="15.2",
        stock__quantity="212",
        stock__offer__isDuo=True,
        stock__offer__name="Guide du Routard Sainte-Hélène",
        stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        stock__offer__withdrawalType=offers_models.WithdrawalTypeEnum.IN_APP,
        dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=4),
        validationAuthorType=bookings_models.BookingValidationAuthorType.BACKOFFICE,
    )
    offerers_factories.UserOffererFactory(offerer=used.offerer)
    cancelled = bookings_factories.CancelledBookingFactory(
        id=1000002,
        user=user2,
        quantity=1,
        amount=12.5,
        token="CNCL02",
        stock__offer__subcategoryId=subcategories.FESTIVAL_SPECTACLE.id,
        stock__beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=11),
        dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=3),
    )
    confirmed = bookings_factories.BookingFactory(
        id=1000003,
        user=user1,
        quantity=1,
        token="ELBEIT",
        stock__price=13.95,
        stock__quantity="2",
        stock__offer__name="Guide Ile d'Elbe 1814 Petit Futé",
        stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=2),
        cancellation_limit_date=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
    )
    reimbursed = bookings_factories.ReimbursedBookingFactory(
        id=1000004,
        user=user3,
        token="REIMB3",
        stock__offer__subcategoryId=subcategories.SPECTACLE_REPRESENTATION.id,
        stock__beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=12),
        dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
    )
    booked = bookings_factories.BookingFactory(
        id=1000005,
        user=user4,
        quantity=1,
        token="ADFTH9",
        stock__price=12.34,
        stock__quantity="1",
        stock__offer__name="Guide Ile d'Elbe 1814 Petit Futé",
        stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        dateCreated=date_utils.get_naive_utc_now(),
    )

    return used, cancelled, confirmed, reimbursed, booked


class ListIndividualBookingsTest(GetEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.list_individual_bookings"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # Use assert_num_queries() instead of assert_no_duplicated_queries() which does not detect one extra query caused
    # by a field added in the jinja template.
    # - fetch session + user (1 query)
    expected_num_queries_when_no_query = 1
    # - fetch individual bookings with extra data (1 query)
    expected_num_queries = expected_num_queries_when_no_query + 1

    def test_list_bookings_without_filter(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries_when_no_query):
            response = authenticated_client.get(url_for(self.endpoint))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_bookings_by_token(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="WTRL00"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID résa"] == str(bookings[0].id)
        assert rows[0]["Contremarque"] == "WTRL00"
        assert rows[0]["Bénéficiaire"].startswith("Napoléon Bonaparte (")
        assert rows[0]["Nom de l'offre"] == "Guide du Routard Sainte-Hélène"
        assert rows[0]["ID offre"].isdigit()
        assert rows[0]["Résa duo"] == "Oui"
        assert rows[0]["Stock"] == "212"
        assert rows[0]["Montant"] == "30,40 €"
        assert rows[0]["Statut"] == "Validée"
        assert rows[0]["Crédit actif"] == "Oui"
        assert rows[0]["Fraude"] == ""
        assert rows[0]["Auteur de la validation"] == "Backoffice"
        assert rows[0]["Date de réservation"].startswith(
            (datetime.date.today() - datetime.timedelta(days=4)).strftime("%d/%m/%Y")
        )
        assert rows[0]["Entité juridique"] == bookings[0].offerer.name
        assert rows[0]["Partenaire culturel"] == bookings[0].venue.name

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]
        assert f"Catégorie : {pro_categories.LIVRE.pro_label}" in extra_data
        assert f"Sous-catégorie : {subcategories.LIVRE_PAPIER.pro_label}" in extra_data
        assert f"Date de validation : {datetime.date.today().strftime('%d/%m/%Y')} à " in extra_data
        assert "Date limite de réservation" not in extra_data
        assert "Date d'annulation" not in extra_data
        assert "Modalités de retrait : Dans l'app" in extra_data

        assert html_parser.extract_pagination_info(response.data) == (1, 1, 1)

    def test_list_bookings_with_expired_deposit(self, authenticated_client):
        user = users_factories.BeneficiaryFactory()
        booking = bookings_factories.UsedBookingFactory(
            user=user,
            token="WTRL00",
        )

        users_factories.DepositGrantFactory(
            bookings=[booking],
            dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=5),
            expirationDate=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="WTRL00"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Crédit actif"] == "Non"

    def test_list_fraudulent_booking(self, authenticated_client):
        fraudulent_booking_tag = bookings_factories.FraudulentBookingTagFactory(booking__token="WTRL00")

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="WTRL00"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Contremarque"] == "WTRL00"
        assert rows[0]["Fraude"] == "Frauduleuse"

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]
        assert f"Date et auteur du marquage frauduleux : {datetime.date.today().strftime('%d/%m/%Y')} à " in extra_data
        assert f"par {fraudulent_booking_tag.author.full_name}" in extra_data

    def test_list_bookings_by_list_of_tokens(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="WTRL00 ELBEIT\tREIMB3\n"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 3
        assert set(row["Contremarque"] for row in rows) == {"WTRL00", "ELBEIT", "REIMB3"}

    @pytest.mark.parametrize(
        "incident_status, display_alert",
        [
            (finance_models.IncidentStatus.VALIDATED, True),
            (finance_models.IncidentStatus.CREATED, False),
            (finance_models.IncidentStatus.CANCELLED, False),
            (finance_models.IncidentStatus.INVOICED, True),
        ],
    )
    def test_display_incident_alert(self, authenticated_client, incident_status, display_alert):
        booking_finance_incident = finance_factories.IndividualBookingFinanceIncidentFactory(
            incident=finance_factories.FinanceIncidentFactory(status=incident_status),
            booking=bookings_factories.ReimbursedBookingFactory(),
        )
        booking_id = booking_finance_incident.booking.id

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=booking_id))
            assert response.status_code == 200

        if display_alert:
            assert "bi-exclamation-triangle-fill" in str(response.data)
        else:
            assert "bi-exclamation-triangle-fill" not in str(response.data)

    @pytest.mark.parametrize(
        "query_args",
        [
            {},
            {"event_from_date": datetime.date(1970, 1, 1)},
            {"event_to_date": datetime.date(2037, 12, 31)},
            {
                "event_from_date": datetime.date(1970, 1, 1),
                "event_to_date": datetime.date(2037, 12, 31),
            },
        ],
    )
    def test_display_download_link(self, authenticated_client, bookings, query_args):
        venue_id = [bookings[0].venueId]
        kwargs = {**query_args, "venue": venue_id, "deposit": "all"}
        response = authenticated_client.get(url_for(self.endpoint, **kwargs))
        assert (b"pc-clipboard" in response.data) == (not query_args)

    def test_list_bookings_by_offer_name(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="Routard Sainte-Hélène"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID résa"] == str(bookings[0].id)
        assert rows[0]["Nom de l'offre"] == "Guide du Routard Sainte-Hélène"
        assert html_parser.extract_pagination_info(response.data) == (1, 1, 1)

    def test_list_bookings_by_booking_id(self, authenticated_client, bookings):
        searched_booking_id = bookings[0].id
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=searched_booking_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID résa"] == str(bookings[0].id)

    def test_list_bookings_by_token_not_found(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="IENA06"))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_bookings_by_offer_id(self, authenticated_client, bookings):
        searched_id = str(bookings[2].stock.offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=searched_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # Warning: test may return more than 1 row when a user id is the same as expected offer id
        assert len(rows) >= 1
        for row_index, row in enumerate(rows):
            if row["ID offre"] == searched_id:
                break
        else:
            assert False, f"Expected offer {searched_id} not found in results"
        assert row["ID résa"] == str(bookings[2].id)
        assert row["Contremarque"] == "ELBEIT"
        assert row["Bénéficiaire"].startswith("Napoléon Bonaparte (")
        assert row["Nom de l'offre"] == "Guide Ile d'Elbe 1814 Petit Futé"
        assert row["Résa duo"] == "Non"
        assert row["Stock"] == "2"
        assert row["Montant"] == "13,95 €"
        assert row["Statut"] == "Confirmée"
        assert row["Date de réservation"].startswith(
            (datetime.date.today() - datetime.timedelta(days=2)).strftime("%d/%m/%Y à")
        )
        assert row["Entité juridique"] == bookings[2].offerer.name
        assert row["Partenaire culturel"] == bookings[2].venue.name

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[row_index]
        assert f"Catégorie : {pro_categories.LIVRE.pro_label}" in extra_data
        assert f"Sous-catégorie : {subcategories.LIVRE_PAPIER.pro_label}" in extra_data
        assert "Date" not in extra_data

        assert html_parser.extract_pagination_info(response.data) == (1, 1, len(rows))

    def test_list_bookings_by_user_id(self, authenticated_client, bookings):
        search_query = str(bookings[1].user.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=search_query))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # Warning: test may return more than 1 row when an offer id is the same as expected user id
        assert len(rows) >= 1
        assert bookings[1].token in set(row["Contremarque"] for row in rows)

    @pytest.mark.parametrize("search_query", ["napo@leon.com", "Napo@Leon.com"])
    def test_list_bookings_by_user(self, authenticated_client, bookings, search_query):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=search_query))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == {"WTRL00", "ELBEIT"}

    def test_list_bookings_by_category(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, category=pro_categories.SPECTACLE.id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == {"CNCL02", "REIMB3"}

    def test_list_bookings_by_cashflow_batch(self, authenticated_client):
        cashflows = finance_factories.CashflowFactory.create_batch(
            3,
            bankAccount=finance_factories.BankAccountFactory(),
        )

        finance_factories.PricingFactory(booking=bookings_factories.UsedBookingFactory())
        finance_factories.PricingFactory(
            booking=bookings_factories.ReimbursedBookingFactory(), cashflows=[cashflows[1]]
        )

        reimbursed_pricing1 = finance_factories.PricingFactory(
            booking=bookings_factories.ReimbursedBookingFactory(), cashflows=[cashflows[0]]
        )
        reimbursed_pricing3 = finance_factories.PricingFactory(
            booking=bookings_factories.ReimbursedBookingFactory(), cashflows=[cashflows[2]]
        )

        searched_cashflow_batches = [cashflows[0].batch.id, cashflows[2].batch.id]
        # one more query because of cashflow_batches validation
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, cashflow_batches=searched_cashflow_batches))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {
            reimbursed_pricing1.bookingId,
            reimbursed_pricing3.bookingId,
        }

    @pytest.mark.parametrize(
        "status, expected_tokens",
        [
            ([forms.BookingStatus.CONFIRMED.name], {"ELBEIT"}),
            ([forms.BookingStatus.USED.name], {"WTRL00"}),
            ([forms.BookingStatus.CANCELLED.name], {"CNCL02"}),
            ([forms.BookingStatus.REIMBURSED.name], {"REIMB3"}),
            ([forms.BookingStatus.BOOKED.name], {"ADFTH9"}),
            (
                [forms.BookingStatus.CONFIRMED.name, forms.BookingStatus.USED.name],
                {"ELBEIT", "WTRL00"},
            ),
        ],
    )
    def test_list_bookings_by_status(self, authenticated_client, bookings, status, expected_tokens):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, status=status))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == expected_tokens

    def test_list_bookings_by_cancellation_reason(self, authenticated_client):
        bookings_factories.CancelledBookingFactory.create_batch(
            size=3,
            token=factory.Iterator(["CNCL01", "CNCL02", "CNCL03"]),
            cancellationReason=factory.Iterator(
                [
                    bookings_models.BookingCancellationReasons.OFFERER.value,
                    bookings_models.BookingCancellationReasons.BENEFICIARY.value,
                    bookings_models.BookingCancellationReasons.FRAUD.value,
                ]
            ),
        )

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, cancellation_reason=["OFFERER", "FRAUD"]))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)

        assert {row["Contremarque"] for row in rows} == {"CNCL01", "CNCL03"}

    def test_list_bookings_by_date(self, authenticated_client, bookings):
        date_from = datetime.date.today() - datetime.timedelta(days=3)
        date_to = datetime.date.today() - datetime.timedelta(days=2)
        separator = " - "
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    from_to_date=f"{date_from.strftime('%d/%m/%Y')}{separator}{date_to.strftime('%d/%m/%Y')}",
                )
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == {"CNCL02", "ELBEIT"}

    @pytest.mark.parametrize(
        "from_date, to_date, expected_tokens",
        [
            ((datetime.date.today() + datetime.timedelta(days=12)).isoformat(), None, {"REIMB3"}),
            (None, (datetime.date.today() + datetime.timedelta(days=12)).isoformat(), {"CNCL02", "REIMB3"}),
            (
                (datetime.date.today() + datetime.timedelta(days=12)).isoformat(),
                (datetime.date.today() + datetime.timedelta(days=12)).isoformat(),
                {"REIMB3"},
            ),
        ],
    )
    def test_list_bookings_by_event_date(self, authenticated_client, bookings, from_date, to_date, expected_tokens):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    event_from_date=from_date,
                    event_to_date=to_date,
                )
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == expected_tokens

    def test_list_bookings_by_offerer(self, authenticated_client, bookings):
        offerer_id = bookings[1].offererId
        # one more query because of offerer validation
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, offerer=offerer_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Contremarque"] == bookings[1].token

    def test_list_bookings_by_venue(self, authenticated_client, bookings):
        venue_ids = [bookings[0].venueId, bookings[2].venueId]
        # one more query because of venue validation
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_ids))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == {bookings[0].token, bookings[2].token}

    @pytest.mark.parametrize(
        "deposit_filter_value, result_token, result_active",
        [
            ("expired", "EXPIRD", "Non"),
            ("active", "ACTIVE", "Oui"),
        ],
    )
    def test_list_bookings_by_deposit_expiration_status(
        self, authenticated_client, deposit_filter_value, result_token, result_active
    ):
        offer = offers_factories.OfferFactory()
        offer_id = offer.id

        expired_deposit_booking = bookings_factories.UsedBookingFactory(token="EXPIRD", stock__offer=offer)
        expired_deposit_booking.deposit.expirationDate = date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
        db.session.flush()

        bookings_factories.UsedBookingFactory(token="ACTIVE", stock__offer=offer)

        with assert_num_queries(self.expected_num_queries):
            # Also search on offer ID to have something realistic (cartesian product found and fixed in such a case)
            response = authenticated_client.get(url_for(self.endpoint, q=str(offer_id), deposit=deposit_filter_value))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Contremarque"] == result_token
        assert rows[0]["Crédit actif"] == result_active

    def test_list_bookings_with_only_all_deposit_filter_considered_as_empty_form(self, authenticated_client):
        old_user = users_factories.BeneficiaryFactory()
        expired_deposit_booking = bookings_factories.UsedBookingFactory(
            user=old_user,
            token="EXPIRD",
        )
        users_factories.DepositGrantFactory(
            bookings=[expired_deposit_booking],
            dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=5),
            expirationDate=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )

        new_user = users_factories.BeneficiaryFactory()
        bookings_factories.UsedBookingFactory(
            user=new_user,
            token="ACTIVE",
        )

        with assert_num_queries(self.expected_num_queries_when_no_query):
            response = authenticated_client.get(url_for(self.endpoint, deposit="all"))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    def test_additional_data_when_reimbursed(self, authenticated_client, bookings):
        reimbursed = bookings[3]
        pricing_venue = offerers_factories.VenueFactory()
        bank_account = finance_factories.BankAccountFactory()
        offerers_factories.VenueBankAccountLinkFactory(venue=pricing_venue, bankAccount=bank_account)
        pricing = finance_factories.PricingFactory(
            booking=reimbursed,
            status=finance_models.PricingStatus.INVOICED,
            venue=pricing_venue,
        )
        cashflow = finance_factories.CashflowFactory(bankAccount=bank_account, pricings=[pricing])

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(self.endpoint, status=bookings_models.BookingStatus.REIMBURSED.name)
            )
            assert response.status_code == 200

        reimbursement_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]
        assert "Total payé par l'utilisateur : 10,10 €" in reimbursement_data
        assert f"Date de remboursement : {reimbursed.reimbursementDate.strftime('%d/%m/%Y à ')}" in reimbursement_data
        assert "Montant remboursé : 10,10 €" in reimbursement_data
        assert f"N° de virement : {cashflow.batch.label}" in reimbursement_data
        assert "Taux de remboursement : 100,0 %" in reimbursement_data
        assert "CFP" not in reimbursement_data

    def test_sort_bookings_by_date(self, authenticated_client, bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, status=[s.name for s in forms.BookingStatus]))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # booked J not used, booked J-2 not used, event J+12, event J+11, used J
        assert [row["Contremarque"] for row in rows] == ["ADFTH9", "ELBEIT", "REIMB3", "CNCL02", "WTRL00"]

    @pytest.mark.parametrize(
        "cancellation_reason, expected_text",
        [
            (bookings_models.BookingCancellationReasons.OFFERER, "Annulée par l'acteur culturel"),
            (
                bookings_models.BookingCancellationReasons.OFFERER_CONNECT_AS,
                "Annulée par Hercule Poirot via Connect As",
            ),
            (bookings_models.BookingCancellationReasons.BENEFICIARY, "Annulée par le bénéficiaire"),
            (bookings_models.BookingCancellationReasons.EXPIRED, "Expirée"),
            (bookings_models.BookingCancellationReasons.FRAUD, "Fraude avérée"),
            (bookings_models.BookingCancellationReasons.FRAUD_SUSPICION, "Suspicion de fraude"),
            (bookings_models.BookingCancellationReasons.FRAUD_INAPPROPRIATE, "Offre non conforme"),
            (
                bookings_models.BookingCancellationReasons.REFUSED_BY_INSTITUTE,
                "Refusée par l'institution",
            ),
            (bookings_models.BookingCancellationReasons.FINANCE_INCIDENT, "Incident finance"),
            (
                bookings_models.BookingCancellationReasons.BACKOFFICE_EVENT_CANCELLED,
                "Annulée depuis le backoffice par Hercule Poirot pour annulation d’évènement",
            ),
            (
                bookings_models.BookingCancellationReasons.BACKOFFICE_BENEFICIARY_REQUEST,
                "Annulée depuis le backoffice par Hercule Poirot sur demande du bénéficiaire",
            ),
            (
                bookings_models.BookingCancellationReasons.BACKOFFICE_OVERBOOKING,
                "Annulée depuis le backoffice par Hercule Poirot pour surbooking",
            ),
            (
                bookings_models.BookingCancellationReasons.BACKOFFICE_OFFER_MODIFIED,
                "Annulée depuis le backoffice par Hercule Poirot pour modification des informations de l'offre",
            ),
            (
                bookings_models.BookingCancellationReasons.BACKOFFICE_OFFER_WITH_WRONG_INFORMATION,
                "Annulée depuis le backoffice par Hercule Poirot pour erreur d'information dans l'offre",
            ),
        ],
    )
    def test_list_cancelled_booking_information(
        self, authenticated_client, legit_user, cancellation_reason, expected_text
    ):
        bookings_factories.CancelledBookingFactory(
            cancellationReason=cancellation_reason,
            cancellationUserId=legit_user.id,
        )

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, cancellation_reason=cancellation_reason.name))
            assert response.status_code == 200

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]

        assert "Catégorie" in extra_data
        assert "Sous-catégorie" in extra_data
        assert "Date d'annulation" in extra_data
        assert "Raison de l'annulation" in extra_data
        assert expected_text in extra_data

    @pytest.mark.parametrize("field", ["offerer", "venue", "cashflow_batches"])
    def test_list_bookings_by_invalid_autocomplete_id(self, authenticated_client, field):
        response = authenticated_client.get(url_for(self.endpoint, **{field: "1)"}))
        assert response.status_code == 400
        assert html_parser.extract_warnings(response.data) == [f"ID invalide pour {field} : 1)"]

    def test_list_caledonian_booking(self, authenticated_client):
        nc_booking = bookings_factories.ReimbursedBookingFactory(
            user=users_factories.CaledonianBeneficiaryFactory(),
            token="NC988F",
            amount=150.0,
            dateCreated=datetime.datetime.combine(datetime.date.today(), datetime.time.min),
            stock=offers_factories.EventStockFactory(
                beginningDatetime=datetime.datetime(2035, 12, 31, 20, 30),
                offer__offererAddress=offerers_factories.OffererAddressFactory(
                    address=geography_factories.CaledonianAddressFactory()
                ),
            ),
        )
        bank_account = finance_factories.BankAccountFactory()
        offerers_factories.VenueBankAccountLinkFactory(venue=nc_booking.venue, bankAccount=bank_account)
        pricing = finance_factories.PricingFactory(
            booking=nc_booking,
            status=finance_models.PricingStatus.INVOICED,
            venue=nc_booking.venue,
        )
        finance_factories.CashflowFactory(bankAccount=bank_account, pricings=[pricing])

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="NC988F"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID résa"] == str(nc_booking.id)
        assert rows[0]["Montant"] == "150,00 € (17900 CFP)"
        assert rows[0]["Date de réservation"] == f"{(datetime.date.today()).strftime('%d/%m/%Y')} à 11h00"
        assert rows[0]["Date de l'évènement"] == "01/01/2036 à 07h30"

        reimbursement_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]
        assert "Total payé par l'utilisateur : 150,00 € (17900 CFP)" in reimbursement_data
        assert "Montant remboursé : 150,00 € (17900 CFP)" in reimbursement_data

    @pytest.mark.parametrize(
        "has_incident, expected_results",
        [
            ("true", ["XXXXXX"]),
            ("false", ["YYYYYY", "ZZZZZZ"]),
        ],
    )
    def test_list_bookings_with_incidents(self, authenticated_client, has_incident, expected_results):
        finance_factories.IndividualBookingFinanceIncidentFactory(
            incident=finance_factories.FinanceIncidentFactory(status=finance_models.IncidentStatus.VALIDATED),
            booking=bookings_factories.ReimbursedBookingFactory(token="XXXXXX"),
        )

        finance_factories.IndividualBookingFinanceIncidentFactory(
            incident=finance_factories.FinanceIncidentFactory(status=finance_models.IncidentStatus.CREATED),
            booking=bookings_factories.ReimbursedBookingFactory(token="YYYYYY"),
        )

        bookings_factories.ReimbursedBookingFactory(token="ZZZZZZ")

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, has_incident=has_incident))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == set(expected_results)

    @pytest.mark.parametrize(
        "is_fraudulent, expected_results",
        [
            ("true", ["XXXXXX"]),
            ("false", ["YYYYYY"]),
        ],
    )
    def test_list_fraudulent_bookings(self, authenticated_client, is_fraudulent, expected_results):
        bookings_factories.FraudulentBookingTagFactory(booking__token="XXXXXX")
        bookings_factories.BookingFactory(token="YYYYYY")

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, is_fraudulent=is_fraudulent))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == set(expected_results)

    @pytest.mark.parametrize("quantity", [1, 2])
    def test_display_duo_bookings(self, authenticated_client, bookings, quantity):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, is_duo=[quantity]))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert {int(r["ID résa"]) for r in rows} == {b.id for b in bookings if b.quantity == quantity}

    @pytest.mark.parametrize(
        "is_free, expected_results",
        [
            ("true", ["XXXXXX"]),
            ("false", ["YYYYYY"]),
        ],
    )
    def test_list_free_bookings(self, authenticated_client, is_free, expected_results):
        bookings_factories.BookingFactory(token="XXXXXX", amount=0)
        bookings_factories.BookingFactory(token="YYYYYY", amount=1)

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, is_free=is_free))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Contremarque"] for row in rows) == set(expected_results)


class MarkBookingAsUsedTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.mark_booking_as_used"
    endpoint_kwargs = {"booking_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_uncancel_and_mark_as_used(self, authenticated_client, bookings):
        cancelled = bookings[1]

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=cancelled.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{cancelled.id}")
        assert cells[2] == str(cancelled.id)

        db.session.refresh(cancelled)
        assert cancelled.status is bookings_models.BookingStatus.USED
        assert cancelled.validationAuthorType == bookings_models.BookingValidationAuthorType.BACKOFFICE

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert "1 réservation a été validée" in alerts["success"]

    def test_uncancel_booking_is_already_used(self, authenticated_client, bookings):
        booking = bookings_factories.UsedBookingFactory()

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
        assert cells[2] == str(booking.id)

        booking = db.session.query(bookings_models.Booking).filter_by(id=booking.id).one()
        assert booking.status == bookings_models.BookingStatus.USED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (f"Impossible de valider ces réservations : - 1 réservation déjà validée ({booking.token})") in alerts[
            "warning"
        ]

    def test_uncancel_booking_is_already_refunded(self, authenticated_client, bookings):
        booking = bookings_factories.ReimbursedBookingFactory()

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
        assert cells[2] == str(booking.id)

        booking = db.session.query(bookings_models.Booking).filter_by(id=booking.id).one()
        assert booking.status == bookings_models.BookingStatus.REIMBURSED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            f"Impossible de valider ces réservations : - 1 réservation déjà remboursée ({booking.token})"
        ) in alerts["warning"]

    def test_uncancel_booking_expired_deposit(self, authenticated_client, bookings):
        booking = bookings_factories.BookingFactory(
            status=bookings_models.BookingStatus.CANCELLED,
            deposit=users_factories.DepositGrantFactory(
                expirationDate=date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
            ),
        )

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
        assert cells[2] == str(booking.id)

        booking = db.session.query(bookings_models.Booking).filter_by(id=booking.id).one()
        assert booking.status == bookings_models.BookingStatus.CANCELLED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            "Impossible de valider ces réservations : "
            f"- 1 réservation dont le crédit associé est expiré ({booking.token})"
        ) in alerts["warning"]

    def test_uncancel_booking_insufficient_funds(self, authenticated_client, bookings):
        beneficiary = users_factories.BeneficiaryFactory()
        booking_id = bookings_factories.CancelledBookingFactory(user=beneficiary, stock__price="250").id
        bookings_factories.ReimbursedBookingFactory(user=beneficiary, stock__price="100")

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking_id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking_id}")
        assert cells[2] == str(booking_id)

        booking = db.session.query(bookings_models.Booking).filter_by(id=booking_id).one()
        assert booking.status == bookings_models.BookingStatus.CANCELLED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            "Impossible de valider ces réservations : "
            f"- 1 réservation dont le crédit associé est insuffisant ({booking.token})"
        ) in alerts["warning"]

    def test_uncancel_booking_no_stock(self, authenticated_client, bookings):
        beneficiary = users_factories.BeneficiaryFactory()
        booking_id = bookings_factories.CancelledBookingFactory(user=beneficiary, quantity=2, stock__quantity=1).id

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking_id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking_id}")
        assert cells[2] == str(booking_id)

        booking = db.session.query(bookings_models.Booking).filter_by(id=booking_id).one()
        assert booking.status == bookings_models.BookingStatus.CANCELLED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            "Impossible de valider ces réservations : "
            f"- 1 réservation dont l'offre n'a plus assez de stock disponible ({booking.token})"
        ) in alerts["warning"]

    def test_uncancel_and_mark_as_used_unlocks_achievement(self, authenticated_client, bookings):
        festival_booking = bookings[1]

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=festival_booking.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{festival_booking.id}")
        assert cells[2] == str(festival_booking.id)
        assert festival_booking.user.achievements


class CancelBookingTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.mark_booking_as_cancelled"
    endpoint_kwargs = {"booking_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_cancel_booking(self, authenticated_client, bookings):
        confirmed = bookings[2]

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=confirmed.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{confirmed.id}")
        assert cells[2] == str(confirmed.id)

        db.session.refresh(confirmed)
        assert confirmed.status is bookings_models.BookingStatus.CANCELLED
        assert confirmed.cancellationReason == bookings_models.BookingCancellationReasons.BACKOFFICE

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert "1 réservation a été annulée" in alerts["success"]

    def test_cancel_booking_decreases_stock_quantity(self, authenticated_client, bookings):
        stock = offers_factories.StockFactory(quantity=3)

        booking = bookings_factories.BookingFactory(stock=stock)
        booking_to_cancel = bookings_factories.BookingFactory(stock=stock)
        offers_factories.ActivationCodeFactory(booking=booking, stock=stock)
        offers_factories.ActivationCodeFactory(booking=booking_to_cancel, stock=stock)

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking_to_cancel.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking_to_cancel.id}")
        assert cells[2] == str(booking_to_cancel.id)

        db.session.refresh(booking_to_cancel)
        assert booking_to_cancel.status == bookings_models.BookingStatus.CANCELLED
        assert booking_to_cancel.cancellationReason == bookings_models.BookingCancellationReasons.BACKOFFICE
        assert stock.quantity == 2
        assert stock.dnBookedQuantity == 1

    def test_cant_cancel_booking_with_processed_pricing(self, authenticated_client, bookings):
        pricing = finance_factories.PricingFactory(status=finance_models.PricingStatus.PROCESSED)

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=pricing.bookingId,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{pricing.bookingId}")
        assert cells[2] == str(pricing.bookingId)

        db.session.refresh(pricing)
        assert pricing.booking.status == bookings_models.BookingStatus.USED

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            f"Impossible d'annuler ces réservations : - 1 réservation déjà remboursée ({pricing.booking.token})"
            in alerts["warning"]
        )

    def test_cant_cancel_reimbursed_booking(self, authenticated_client, bookings):
        reimbursed = bookings[3]
        old_status = reimbursed.status

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=reimbursed.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{reimbursed.id}")
        assert cells[2] == str(reimbursed.id)

        db.session.refresh(reimbursed)
        assert reimbursed.status == old_status

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            f"Impossible d'annuler ces réservations : - 1 réservation déjà remboursée ({reimbursed.token})"
            in alerts["warning"]
        )

    def test_cant_cancel_cancelled_booking(self, authenticated_client, bookings):
        cancelled = bookings[1]
        old_status = cancelled.status

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=cancelled.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{cancelled.id}")
        assert cells[2] == str(cancelled.id)

        db.session.refresh(cancelled)
        assert cancelled.status == old_status

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            f"Impossible d'annuler ces réservations : - 1 réservation déjà annulée ({cancelled.token})"
            in alerts["warning"]
        )

    def test_cant_cancel_booking_without_reason(self, authenticated_client, bookings):
        confirmed = bookings[2]

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=confirmed.id,
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        assert len(response.data) == 0

        db.session.refresh(confirmed)
        assert confirmed.status == bookings_models.BookingStatus.CONFIRMED
        assert confirmed.cancellationReason is None

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert "Les données envoyées comportent des erreurs. Raison : Information obligatoire ;" in alerts["warning"]

    @pytest.mark.settings(EMS_SUPPORT_EMAIL_ADDRESS="ems.support@example.com")
    @pytest.mark.features(ENABLE_EMS_INTEGRATION=True)
    def test_ems_cancel_external_booking_from_backoffice(self, authenticated_client, requests_mock):
        beneficiary = users_factories.BeneficiaryFactory()
        ems_provider = get_provider_by_local_class("EMSStocks")
        venue_provider = providers_factories.VenueProviderFactory(provider=ems_provider)
        providers_factories.CinemaProviderPivotFactory(venue=venue_provider.venue)
        offer = offers_factories.EventOfferFactory(
            name="Film",
            venue=venue_provider.venue,
            subcategoryId=subcategories.SEANCE_CINE.id,
            lastProviderId=ems_provider.id,
        )
        stock = offers_factories.EventStockFactory(
            offer=offer,
            idAtProviders="1111%2222%EMS#3333",
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=2),
        )
        booking = bookings_factories.BookingFactory(stock=stock, user=beneficiary)
        external_bookings_factories.ExternalBookingFactory(
            booking=booking,
            barcode="123456789",
            additional_information={
                "num_cine": "9997",
                "num_caisse": "255",
                "num_trans": 1257,
                "num_ope": 147149,
            },
        )
        requests_mock.post(
            url=re.compile(r"https://fake_url.com/ANNULATION/*"),
            json={"code_erreur": 5, "message_erreur": "Erreur lors de l'appel au cinéma", "statut": 0},
        )

        assert booking.status is bookings_models.BookingStatus.CONFIRMED
        assert len(booking.externalBookings) == 1
        assert booking.externalBookings[0].barcode
        assert booking.externalBookings[0].seat
        assert booking.externalBookings[0].additional_information == {
            "num_cine": "9997",
            "num_caisse": "255",
            "num_trans": 1257,
            "num_ope": 147149,
        }

        old_quantity = stock.dnBookedQuantity

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
        assert cells[2] == str(booking.id)

        db.session.refresh(booking)
        assert booking.status == bookings_models.BookingStatus.CANCELLED
        assert stock.dnBookedQuantity == old_quantity - 1
        assert booking.cancellationReason == bookings_models.BookingCancellationReasons.BACKOFFICE

        assert len(mails_testing.outbox) == 1
        assert mails_testing.outbox[0]["To"] == "ems.support@example.com"
        assert mails_testing.outbox[0]["template"] == dataclasses.asdict(
            TransactionalEmail.EXTERNAL_BOOKING_SUPPORT_CANCELLATION.value
        )
        assert mails_testing.outbox[0]["params"] == {
            "EXTERNAL_BOOKING_INFORMATION": "barcode: 123456789, additional_information: {'num_ope': 147149, 'num_cine': '9997', 'num_trans': 1257, 'num_caisse': '255'}"
        }

    def test_cancel_cgr_external_booking_unilaterally(self, authenticated_client, requests_mock):
        requests_mock.get(
            "https://cgr-cinema-0.example.com/web_service?wsdl", text=soap_definitions.WEB_SERVICE_DEFINITION
        )
        requests_mock.post(
            "https://cgr-cinema-0.example.com/web_service",
            text=cgr_fixtures.cgr_annulation_response_template(
                message_error="L'annulation n'a pas pu être prise en compte : Code barre non reconnu / annulation impossible",
            ),
        )
        cgr_provider = get_provider_by_local_class("CGRStocks")
        venue_provider = providers_factories.VenueProviderFactory(provider=cgr_provider)
        cinema_provider_pivot = providers_factories.CGRCinemaProviderPivotFactory(
            venue=venue_provider.venue, idAtProvider=venue_provider.venueIdAtOfferProvider
        )
        providers_factories.CGRCinemaDetailsFactory(
            cinemaProviderPivot=cinema_provider_pivot, cinemaUrl="https://cgr-cinema-0.example.com/web_service"
        )

        offer = offers_factories.OfferFactory(
            lastProvider=cgr_provider,
            idAtProvider="123%12354114%CGR",
            subcategoryId=subcategories.SEANCE_CINE.id,
            venue=venue_provider.venue,
        )
        stock = offers_factories.StockFactory(
            lastProvider=cgr_provider,
            idAtProviders="123%12354114%CGR#111",
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=2),
            offer=offer,
        )
        booking_to_cancel = bookings_factories.BookingFactory(stock=stock)
        external_bookings_factories.ExternalBookingFactory(booking=booking_to_cancel)

        response = self.post_to_endpoint(
            authenticated_client,
            booking_id=booking_to_cancel.id,
            form={"reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking_to_cancel.id}")
        assert cells[2] == str(booking_to_cancel.id)

        db.session.refresh(booking_to_cancel)
        assert booking_to_cancel.status == bookings_models.BookingStatus.CANCELLED
        assert booking_to_cancel.cancellationReason == bookings_models.BookingCancellationReasons.BACKOFFICE

        alerts = flash.get_htmx_flash_messages(authenticated_client)

        assert "1 réservation a été annulée" in alerts["success"]


class GetBatchMarkAsUsedIndividualBookingsFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_batch_validate_individual_bookings_form"
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_get_batch_mark_as_used_booking_form(self, legit_user, authenticated_client):
        bookings_factories.BookingFactory()
        with assert_num_queries(1):  # session
            response = authenticated_client.get(url_for(self.endpoint))
            # Rendering is not checked, but at least the fetched frame does not crash
            assert response.status_code == 200


class BatchMarkBookingAsUsedTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.batch_validate_individual_bookings"
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_batch_mark_as_used_bookings(self, legit_user, authenticated_client):
        bookings = bookings_factories.BookingFactory.create_batch(3)
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            db.session.refresh(booking)
            assert booking.status is bookings_models.BookingStatus.USED
            assert booking.validationAuthorType == bookings_models.BookingValidationAuthorType.BACKOFFICE

            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

    def test_batch_mark_as_used_cancelled_bookings(self, legit_user, authenticated_client):
        bookings = bookings_factories.BookingFactory.create_batch(3, status=bookings_models.BookingStatus.CANCELLED)
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            db.session.refresh(booking)
            assert booking.status is bookings_models.BookingStatus.USED
            assert booking.validationAuthorType == bookings_models.BookingValidationAuthorType.BACKOFFICE

            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

    def test_batch_mark_as_used_with_already_used_bookings(self, legit_user, authenticated_client):
        bookings = bookings_factories.BookingFactory.create_batch(3, status=bookings_models.BookingStatus.CANCELLED)
        bookings[0].status = bookings_models.BookingStatus.USED
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

        assert bookings[0].status is bookings_models.BookingStatus.USED
        for booking in bookings[1:]:
            db.session.refresh(booking)
            assert booking.status is bookings_models.BookingStatus.USED
            assert booking.validationAuthorType == bookings_models.BookingValidationAuthorType.BACKOFFICE

    def test_batch_mark_as_used_bookings_with_expired_deposit(self, legit_user, authenticated_client):
        expired_booking = bookings_factories.BookingFactory(
            status=bookings_models.BookingStatus.CANCELLED,
            deposit=users_factories.DepositGrantFactory(
                expirationDate=date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
            ),
        )
        other_booking = bookings_factories.BookingFactory(status=bookings_models.BookingStatus.CANCELLED)

        bookings = (
            expired_booking,
            other_booking,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

        db.session.refresh(expired_booking)
        assert expired_booking.status is bookings_models.BookingStatus.CANCELLED
        db.session.refresh(other_booking)
        assert other_booking.status is bookings_models.BookingStatus.USED

    def test_batch_mark_as_used_bookings_with_insufficient_funds(self, legit_user, authenticated_client):
        poor_user = users_factories.BeneficiaryFactory(deposit__amount=1)
        insufficient_funds_booking = bookings_factories.BookingFactory(
            status=bookings_models.BookingStatus.CANCELLED, user=poor_user
        )
        other_booking = bookings_factories.BookingFactory()

        bookings = (
            insufficient_funds_booking,
            other_booking,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200

        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

    def test_batch_mark_as_used_bookings_insufficient_stock(self, legit_user, authenticated_client):
        insufficient_stock_booking = bookings_factories.BookingFactory(
            status=bookings_models.BookingStatus.CANCELLED,
            deposit=users_factories.DepositGrantFactory(
                expirationDate=date_utils.get_naive_utc_now() + datetime.timedelta(days=1)
            ),
            stock__quantity=0,
        )
        other_booking = bookings_factories.BookingFactory()

        bookings = (
            insufficient_stock_booking,
            other_booking,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200

        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

    def test_batch_mark_as_used_when_offerer_is_closed(self, legit_user, authenticated_client):
        booking = bookings_factories.BookingFactory(
            stock__offer__venue__managingOfferer=offerers_factories.ClosedOffererFactory()
        )

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": booking.id},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
        assert cells[2] == str(booking.id)

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert (
            f"Une erreur s'est produite pour la réservation ({booking.token}) : "
            "Une contremarque ne peut plus être validée sur une structure fermée"
        ) in alerts["warning"]
        assert booking.status == bookings_models.BookingStatus.CONFIRMED

    def test_batch_mark_as_used_bookings_with_multiple_errors(self, legit_user, authenticated_client):
        cancelled_booking = bookings_factories.CancelledBookingFactory()
        insufficient_stock_booking = bookings_factories.BookingFactory(
            status=bookings_models.BookingStatus.CANCELLED,
            deposit=users_factories.DepositGrantFactory(
                expirationDate=date_utils.get_naive_utc_now() + datetime.timedelta(days=1)
            ),
            stock__quantity=0,
        )
        already_reimbursed_booking = bookings_factories.ReimbursedBookingFactory()
        another_reimbursed_booking = bookings_factories.ReimbursedBookingFactory()
        bookings = (
            cancelled_booking,
            insufficient_stock_booking,
            already_reimbursed_booking,
            another_reimbursed_booking,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        warning = (
            alerts["warning"]
            .pop()
            .split("Certaines réservations n'ont pas pu être validées et ont été ignorées : ")[-1]
        )
        warnings = warning.split("- ")
        assert "1 réservation a été validée" in alerts["success"]
        assert (
            f"1 réservation dont l'offre n'a plus assez de stock disponible ({insufficient_stock_booking.token})"
            in warnings
        )
        assert (
            f"2 réservations déjà remboursées ({already_reimbursed_booking.token}, {another_reimbursed_booking.token})"
            in warnings
            or f"2 réservations déjà remboursées ({another_reimbursed_booking.token}, {already_reimbursed_booking.token})"
            in warnings
        )

    def test_batch_mark_as_used_unlocks_achievement(self, authenticated_client, bookings):
        festival_booking = bookings[1]

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": str(festival_booking.id)},
            headers={"hx-request": "true"},
        )

        cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{festival_booking.id}")
        assert cells[2] == str(festival_booking.id)

        assert festival_booking.user.achievements


class GetBatchCancelIndividualBookingsFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_batch_cancel_individual_bookings_form"
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_get_batch_cancel_booking_form(self, legit_user, authenticated_client):
        bookings_factories.BookingFactory()
        with assert_num_queries(1):  # session
            url = url_for(self.endpoint)
            response = authenticated_client.get(url)
            # Rendering is not checked, but at least the fetched frame does not crash
            assert response.status_code == 200


class BatchCancelIndividualBookingsTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.batch_cancel_individual_bookings"
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_batch_cancel_bookings(self, legit_user, authenticated_client):
        bookings = bookings_factories.BookingFactory.create_batch(3)
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={
                "object_ids": parameter_ids,
                "reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value,
            },
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            db.session.refresh(booking)
            assert booking.status is bookings_models.BookingStatus.CANCELLED
            assert booking.cancellationReason == bookings_models.BookingCancellationReasons.BACKOFFICE

        alerts = flash.get_htmx_flash_messages(authenticated_client)
        assert "3 réservations ont été annulées" in alerts["success"]

    def test_batch_cancel_booking_with_multiple_errors(self, legit_user, authenticated_client):
        booking_to_cancel = bookings_factories.UsedBookingFactory()
        already_cancelled_booking = bookings_factories.CancelledBookingFactory()
        already_reimbursed_booking = bookings_factories.ReimbursedBookingFactory()

        bookings = (
            booking_to_cancel,
            already_cancelled_booking,
            already_reimbursed_booking,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={
                "object_ids": parameter_ids,
                "reason": bookings_models.BookingCancellationReasons.BACKOFFICE.value,
            },
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)


class GetBatchTagFraudulentBookingsFormTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_batch_tag_fraudulent_bookings_form"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_get_batch_tag_fraudulent_bookings_form(self, legit_user, authenticated_client):
        booking = bookings_factories.BookingFactory()

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": f"{booking.id}"},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        response_text = html_parser.content_as_text(response.data)
        assert "Voulez-vous vraiment marquer ces réservations comme frauduleuses ?" in response_text

    def test_get_batch_tag_fraudulent_bookings_form_with_no_valid_booking(self, legit_user, authenticated_client):
        booking = bookings_factories.FraudulentBookingTagFactory().booking

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": f"{booking.id}"},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        response_text = html_parser.content_as_text(response.data)
        assert "Toutes les réservations sélectionnées sont déjà frauduleuses." in response_text


class BatchTagFraudulentBookingsTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.batch_tag_fraudulent_bookings"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_batch_tag_fraudulent_bookings_without_email(self, legit_user, authenticated_client):
        bookings = bookings_factories.BookingFactory.create_batch(3)
        bookings.append(
            bookings_factories.FraudulentBookingTagFactory(
                dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=4)
            ).booking
        )
        bookings.append(
            bookings_factories.FraudulentBookingTagFactory(
                dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=4)
            ).booking
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={
                "object_ids": parameter_ids,
                "send_mails": False,
            },
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

        all_tags = db.session.query(bookings_models.FraudulentBookingTag).all()
        assert len(all_tags) == 5
        just_created_tags = (
            db.session.query(bookings_models.FraudulentBookingTag)
            .filter(
                bookings_models.FraudulentBookingTag.dateCreated
                > date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
            )
            .all()
        )
        assert len(just_created_tags) == 3
        assert len(mails_testing.outbox) == 0

    def test_batch_tag_fraudulent_bookings_with_email(self, legit_user, authenticated_client):
        booking_with_email = bookings_factories.BookingFactory(stock__offer__bookingEmail="email1@example.com")
        booking_with_venue_email = bookings_factories.BookingFactory(
            stock__offer__venue__bookingEmail="email1@example.com"
        )
        booking_with_other_email = bookings_factories.BookingFactory(stock__offer__bookingEmail="email2@example.com")
        already_fraudulent_booking_with_other_email = bookings_factories.FraudulentBookingTagFactory(
            booking__stock__offer__bookingEmail="email2@example.com",
            dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=4),
        ).booking
        already_fraudulent_booking_with_different_email = bookings_factories.FraudulentBookingTagFactory(
            booking__stock__offer__bookingEmail="email3@example.com",
            dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=4),
        ).booking

        bookings = (
            booking_with_email,
            booking_with_venue_email,
            booking_with_other_email,
            already_fraudulent_booking_with_other_email,
            already_fraudulent_booking_with_different_email,
        )
        parameter_ids = ",".join(str(booking.id) for booking in bookings)

        response = self.post_to_endpoint(
            authenticated_client,
            form={
                "object_ids": parameter_ids,
                "send_mails": True,
            },
        )

        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

        all_tags = db.session.query(bookings_models.FraudulentBookingTag).all()
        assert len(all_tags) == 5
        just_created_tags = (
            db.session.query(bookings_models.FraudulentBookingTag)
            .filter(
                bookings_models.FraudulentBookingTag.dateCreated
                > date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
            )
            .all()
        )
        assert len(just_created_tags) == 3

        assert len(mails_testing.outbox) == 2
        assert {e["To"] for e in mails_testing.outbox} == {"email1@example.com", "email2@example.com"}
        mail1 = [e for e in mails_testing.outbox if e["To"] == "email1@example.com"][0]
        assert mail1["template"] == dataclasses.asdict(TransactionalEmail.FRAUDULENT_BOOKING_SUSPICION.value)
        assert set(mail1["params"]["TOKEN_LIST"].split(", ")) == {
            booking_with_email.token,
            booking_with_venue_email.token,
        }
        mail2 = [e for e in mails_testing.outbox if e["To"] == "email2@example.com"][0]
        assert mail2["template"] == dataclasses.asdict(TransactionalEmail.FRAUDULENT_BOOKING_SUSPICION.value)
        assert set(mail2["params"]["TOKEN_LIST"].split(", ")) == {booking_with_other_email.token}


class GetBatchRemoveFraudulentBookingTagFormTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_batch_remove_fraudulent_booking_tag_form"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_get_batch_remove_fraudulent_booking_tag_form(self, legit_user, authenticated_client):
        booking = bookings_factories.FraudulentBookingTagFactory().booking

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": f"{booking.id}"},
        )

        assert response.status_code == 200
        response_text = html_parser.content_as_text(response.data)
        assert "Voulez-vous vraiment ne plus marquer ces réservations comme frauduleuses ?" in response_text

    def test_get_batch_remove_fraudulent_booking_tag_form_with_no_valid_booking(self, legit_user, authenticated_client):
        booking = bookings_factories.BookingFactory()

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": f"{booking.id}"},
        )

        assert response.status_code == 200
        response_text = html_parser.content_as_text(response.data)
        assert "Aucune réservation sélectionnée n'a de marquage frauduleux." in response_text


class BatchRemoveFraudulentBookingTagTest(PostEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.batch_remove_fraudulent_booking_tag"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_batch_remove_fraudulent_booking_tag(self, legit_user, authenticated_client):
        bookings_factories.BookingFactory()
        bookings = {
            bookings_factories.FraudulentBookingTagFactory().booking,
            bookings_factories.FraudulentBookingTagFactory().booking,
        }
        unrelated_tag = bookings_factories.FraudulentBookingTagFactory()
        parameter_ids = ",".join(str(booking.id) for booking in bookings)
        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            headers={"hx-request": "true"},
        )

        assert response.status_code == 200
        for booking in bookings:
            cells = html_parser.extract_plain_row(response.data, id=f"booking-row-{booking.id}")
            assert cells[2] == str(booking.id)

        all_tags = db.session.query(bookings_models.FraudulentBookingTag).all()
        assert len(all_tags) == 1
        assert all_tags[0].bookingId == unrelated_tag.bookingId


class GetIndividualBookingCSVDownloadTest(GetEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_individual_booking_csv_download"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # session + bookings
    expected_num_queries = 2

    def test_csv_length(self, authenticated_client, bookings):
        venue_id = bookings[0].venueId

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_id))
            assert response.status_code == 200

        assert len(response.data.split(b"\n")) == 4


class GetIndividualBookingXLSXDownloadTest(GetEndpointHelper):
    endpoint = "backoffice_web.individual_bookings.get_individual_booking_xlsx_download"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # session + bookings
    expected_num_queries = 2

    def reader_from_response(self, response):
        wb = openpyxl.load_workbook(BytesIO(response.data))
        return wb.active

    def test_csv_length(self, authenticated_client, bookings):
        venue_id = bookings[0].venueId

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_id))
            assert response.status_code == 200

        sheet = self.reader_from_response(response)
        assert sheet.cell(row=1, column=1).value == "Structure"
        assert sheet.cell(row=2, column=1).value == bookings[0].venue.name
        assert sheet.cell(row=3, column=1).value == bookings[0].venue.name
        assert sheet.cell(row=4, column=1).value == None
