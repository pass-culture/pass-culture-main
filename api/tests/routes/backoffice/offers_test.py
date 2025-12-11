import dataclasses
import datetime
import decimal
from io import BytesIO
from itertools import count
from operator import itemgetter
from unittest import mock
from unittest.mock import patch

import openpyxl
import pytest
from flask import g
from flask import url_for

from pcapi.core.bookings import factories as bookings_factories
from pcapi.core.bookings.models import BookingStatus
from pcapi.core.categories import pro_categories
from pcapi.core.categories import subcategories
from pcapi.core.criteria import factories as criteria_factories
from pcapi.core.finance import conf as finance_conf
from pcapi.core.finance import factories as finance_factories
from pcapi.core.finance import models as finance_models
from pcapi.core.geography import factories as geography_factories
from pcapi.core.highlights import factories as highlight_factories
from pcapi.core.mails import testing as mails_testing
from pcapi.core.mails.transactional.sendinblue_template_ids import TransactionalEmail
from pcapi.core.offerers import constants as offerers_constants
from pcapi.core.offerers import factories as offerers_factories
from pcapi.core.offerers import models as offerers_models
from pcapi.core.offerers import schemas as offerers_schemas
from pcapi.core.offers import factories as offers_factories
from pcapi.core.offers import models as offers_models
from pcapi.core.permissions import factories as perm_factories
from pcapi.core.permissions import models as perm_models
from pcapi.core.providers import factories as providers_factories
from pcapi.core.search.models import IndexationReason
from pcapi.core.testing import assert_num_queries
from pcapi.core.users import factories as users_factories
from pcapi.models import db
from pcapi.models.offer_mixin import OfferValidationType
from pcapi.routes.backoffice.filters import format_date
from pcapi.utils import date as date_utils
from pcapi.utils.regions import get_department_code_from_city_code

from .helpers import button as button_helpers
from .helpers import html_parser
from .helpers.get import GetEndpointHelper
from .helpers.post import PostEndpointHelper


pytestmark = [
    pytest.mark.usefixtures("db_session"),
    pytest.mark.backoffice,
]


@pytest.fixture(scope="function", name="criteria")
def criteria_fixture() -> list:
    return criteria_factories.CriterionFactory.create_batch(4)


@pytest.fixture(scope="function", name="offers")
def offers_fixture(criteria) -> tuple:
    offer_with_unlimited_stock = offers_factories.OfferFactory(
        criteria=[criteria[0]],
        offererAddress=offerers_factories.OffererAddressFactory(
            address__postalCode="47000", address__departmentCode="47"
        ),
        subcategoryId=subcategories.SUPPORT_PHYSIQUE_MUSIQUE_VINYLE.id,
        author=users_factories.ProFactory(),
        extraData={"musicType": 501, "musicSubType": 510, "gtl_id": "02050000"},
        lastProvider=providers_factories.ProviderFactory(name="Music Provider"),
        product=offers_factories.ProductFactory(
            subcategoryId=subcategories.SUPPORT_PHYSIQUE_MUSIQUE_VINYLE.id,
            extraData={"musicType": 501, "musicSubType": 510, "gtl_id": "02050000"},
        ),
    )
    offer_with_limited_stock = offers_factories.EventOfferFactory(
        name="A Very Specific Name",
        lastValidationDate=datetime.date(2022, 2, 22),
        offererAddress=offerers_factories.OffererAddressFactory(
            address__postalCode="97400", address__departmentCode="974"
        ),
        subcategoryId=subcategories.FESTIVAL_LIVRE.id,
        validation=offers_models.OfferValidationStatus.APPROVED,
        extraData={"visa": "2023123456", "showType": 100, "showSubType": 104},
    )
    offer_with_two_criteria = offers_factories.OfferFactory(
        name="A Very Specific Name That Is Longer",
        criteria=[criteria[0], criteria[1]],
        dateCreated=datetime.datetime.now(datetime.UTC) - datetime.timedelta(days=2),
        validation=offers_models.OfferValidationStatus.REJECTED,
        offererAddress=offerers_factories.OffererAddressFactory(
            address__postalCode="74000", address__departmentCode="74"
        ),
        subcategoryId=subcategories.LIVRE_PAPIER.id,
        ean="9781234567890",
    )
    offer_with_a_lot_of_types = offers_factories.OfferFactory(
        criteria=[criteria[3]],
        offererAddress=offerers_factories.OffererAddressFactory(
            address__postalCode="10000", address__departmentCode="10"
        ),
        subcategoryId=subcategories.JEU_EN_LIGNE.id,
        author=users_factories.ProFactory(),
        extraData={"musicType": 870, "musicSubType": 871, "gtl_id": "14000000", "showType": 1510, "showSubType": 1511},
    )
    offers_factories.StockFactory(quantity=None, offer=offer_with_unlimited_stock, price=10.1)
    offers_factories.StockFactory(offer=offer_with_unlimited_stock, price=15)
    offers_factories.EventStockFactory(
        quantity=10,
        dnBookedQuantity=0,
        offer=offer_with_limited_stock,
        beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=1),
    )
    offers_factories.EventStockFactory(
        quantity=10,
        dnBookedQuantity=5,
        offer=offer_with_limited_stock,
        beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=3),
    )
    return offer_with_unlimited_stock, offer_with_limited_stock, offer_with_two_criteria, offer_with_a_lot_of_types


class ListOffersTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.list_offers"
    needed_permission = perm_models.Permissions.READ_OFFERS

    def _get_query_args_by_id(self, id_: int) -> dict[str, str]:
        return {
            "search-0-search_field": "ID",
            "search-0-operator": "IN",
            "search-0-string": str(id_),
        }

    # Use assert_num_queries() instead of assert_no_duplicated_queries() which does not detect one extra query caused
    # by a field added in the jinja template.
    # - fetch session + user (1 query)
    # - fetch offers with joinedload including extra data (1 query)
    expected_num_queries_with_no_result = 2
    # - fetch venue addresses and offer addresses (2 x selectinload: 2 queries)
    expected_num_queries_with_results = expected_num_queries_with_no_result + 2
    # - fetch providers (selectinload: 1 query)
    expected_num_queries_with_provider = expected_num_queries_with_results + 1

    def test_list_offers_without_filter(self, authenticated_client, offers):
        # no filter => no query to fetch offers
        with assert_num_queries(self.expected_num_queries_with_no_result - 1):
            response = authenticated_client.get(url_for(self.endpoint))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    @pytest.mark.parametrize(
        "admin_user,stock_data_expected,tag_data_expected,fraud_rules_expected,fraud_data_expected",
        [
            ("read_only_bo_user", True, False, False, False),
            ("pro_fraud_admin", False, False, True, True),
            ("support_pro_n2_admin", True, True, False, False),
            ("homologation_admin", True, False, True, False),
        ],
    )
    def test_list_offers_by_id(
        self,
        client,
        read_only_bo_user,
        pro_fraud_admin,
        support_pro_n2_admin,
        homologation_admin,
        offers,
        admin_user,
        stock_data_expected,
        tag_data_expected,
        fraud_rules_expected,
        fraud_data_expected,
    ):
        user = locals()[admin_user]
        client = client.with_bo_session_auth(user)
        query_args = self._get_query_args_by_id(offers[0].id)
        with assert_num_queries(self.expected_num_queries_with_provider):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offers[0].id)
        assert rows[0]["Nom de l'offre"] == offers[0].name
        assert rows[0]["EAN / Allociné ID"] == offers[0].product.ean
        assert rows[0]["Catégorie"] == offers[0].category.pro_label
        assert rows[0]["Sous-catégorie"] == offers[0].subcategory.pro_label
        assert rows[0]["État"] == "• Validée"
        assert rows[0]["Date de création"] == (datetime.date.today()).strftime("%d/%m/%Y")
        assert rows[0]["Dernière validation"] == ""
        assert rows[0]["Dép."] == offers[0].offererAddress.address.departmentCode
        assert rows[0]["Entité juridique"] == offers[0].venue.managingOfferer.name
        assert rows[0]["Partenaire culturel"] == offers[0].venue.name
        assert rows[0]["Date(s) de l'évènement"] == ""
        assert rows[0]["Date(s) limite(s) de réservation"] == ""
        assert rows[0]["Partenaire technique"] == "Music Provider"
        assert rows[0]["Tarif"] == "10,10 € - 15,00 €"
        assert rows[0]["Créateur de l'offre"] == offers[0].author.full_name

        if stock_data_expected:
            assert rows[0]["Stock réservé"] == "0"
            assert rows[0]["Stock restant"] == "Illimité"
        else:
            assert "Stock réservé" not in rows[0]
            assert "Stock restant" not in rows[0]

        if tag_data_expected:
            assert rows[0]["Tag"] == offers[0].criteria[0].name
            assert rows[0]["Pond."] == ""
        else:
            assert "Tag" not in rows[0]
            assert "Pond." not in rows[0]

        if fraud_rules_expected:
            assert rows[0]["Règles de conformité"] == ""
        else:
            assert "Règles de conformité" not in rows[0]

        if fraud_data_expected:
            assert rows[0]["Score data"] == ""
        else:
            assert "Score data" not in rows[0]

    def test_list_offers_by_ids_list(self, authenticated_client, offers):
        query_args = self._get_query_args_by_id(f"{offers[0].id}, {offers[2].id}\n")
        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 2
        assert set(int(row["ID"]) for row in rows) == {offers[0].id, offers[2].id}

    def test_list_offers_by_name(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "NAME",
            "search-0-operator": "CONTAINS",
            "search-0-string": offers[1].name,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 2
        rows = sorted(rows, key=itemgetter("ID"))  # ensures deterministic order
        assert rows[0]["ID"] == str(offers[1].id)
        assert rows[0]["Nom de l'offre"] == offers[1].name
        assert rows[0]["Catégorie"] == offers[1].category.pro_label
        assert rows[0]["Sous-catégorie"] == offers[1].subcategory.pro_label
        assert rows[0]["État"] == "• Validée"
        assert rows[0]["Date de création"] == (datetime.date.today()).strftime("%d/%m/%Y")
        assert rows[0]["Dernière validation"] == "22/02/2022"
        assert rows[0]["Dép."] == offers[1].offererAddress.address.departmentCode
        assert rows[0]["Entité juridique"] == offers[1].venue.managingOfferer.name
        assert rows[0]["Partenaire culturel"] == offers[1].venue.name
        assert rows[1]["ID"] == str(offers[2].id)
        assert rows[1]["Nom de l'offre"] == offers[2].name

    def test_list_offers_by_product(self, authenticated_client):
        product = offers_factories.ProductFactory()
        offer1 = offers_factories.OfferFactory(product=product)
        offer2 = offers_factories.OfferFactory(product=product)
        offers_factories.OfferFactory(product=offers_factories.ProductFactory())
        offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "PRODUCT",
            "search-0-operator": "EQUALS",
            "search-0-integer": offer1.id,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 2
        assert int(rows[0]["ID"]) in {offer1.id, offer2.id}
        assert int(rows[1]["ID"]) in {offer1.id, offer2.id}

    @pytest.mark.parametrize("ean", ["9781234567890", " 978-1234567890", "978 1234567890\t"])
    def test_list_offers_by_ean(self, authenticated_client, offers, ean):
        query_args = {
            "search-0-search_field": "EAN",
            "search-0-operator": "EQUALS",
            "search-0-string": ean,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert int(rows[0]["ID"]) == offers[2].id

    def test_list_offers_by_allocine_id(self, authenticated_client):
        allocine_id = 123456789
        offer = offers_factories.OfferFactory(
            product=offers_factories.ProductFactory(
                extraData={"allocineId": allocine_id},
            ),
        )
        offers_factories.OfferFactory(
            product=offers_factories.ProductFactory(
                extraData={"allocineId": 111111111111},
            ),
        )

        query_args = {
            "search-0-search_field": "ALLOCINE_ID",
            "search-0-operator": "EQUALS",
            "search-0-integer": int(allocine_id),
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert int(rows[0]["ID"]) == offer.id

    def test_list_offers_by_creation_date(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "CREATION_DATE",
            "search-0-operator": "DATE_TO",
            "search-0-date": (datetime.date.today() - datetime.timedelta(days=1)).isoformat(),
            "search-2-search_field": "CREATION_DATE",
            "search-2-operator": "DATE_FROM",
            "search-2-date": (datetime.date.today() - datetime.timedelta(days=3)).isoformat(),
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[2].id}

    def test_list_offers_by_status_and_event_date(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "STATUS",
            "search-0-operator": "IN",
            "search-0-status": offers_models.OfferStatus.ACTIVE.value,
            "search-2-search_field": "EVENT_DATE",
            "search-2-operator": "DATE_TO",
            "search-2-date": (datetime.date.today() + datetime.timedelta(days=2)).isoformat(),
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    def test_list_offers_without_sort_should_not_have_created_date_sort_link(self, authenticated_client, offers):
        query_args = self._get_query_args_by_id(offers[0].id)
        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        assert "sort=dateCreated&amp;order=desc" not in str(response.data)

    def test_list_offers_with_sort_should_have_created_date_sort_link(self, authenticated_client, offers):
        query_args = self._get_query_args_by_id(offers[0].id) | {"sort": "dateCreated", "order": "asc", "q": "e"}

        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        assert "sort=dateCreated&amp;order=desc" in str(response.data)

    def test_list_offers_with_and_sort_should_have_created_date_sort_link(self, authenticated_client, offers):
        query_args = {
            "sort": "dateCreated",
            "order": "asc",
            "search-0-search_field": "NAME",
            "search-0-operator": "NAME_EQUALS",
            "search-0-string": "A Very Specific Name",
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        expected_url = (
            "/pro/offer?sort=dateCreated&amp;order=desc&amp;search-0-search_field=NAME&amp;"
            "search-0-operator=NAME_EQUALS&amp;search-0-boolean=true&amp;search-0-string=A+Very+Specific+Name"
        )
        assert expected_url in str(response.data)

    @pytest.mark.parametrize(
        "operator,valid_date,not_valid_date",
        [
            (
                "DATE_FROM",
                date_utils.get_naive_utc_now() + datetime.timedelta(days=1),
                date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
            ),
            (
                "DATE_TO",
                date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
                date_utils.get_naive_utc_now() + datetime.timedelta(days=1),
            ),
        ],
    )
    def test_list_offers_with_booking_limit_date_filter(
        self, authenticated_client, operator, valid_date, not_valid_date
    ):
        should_be_displayed_offer = offers_factories.OfferFactory(
            stocks=[offers_factories.StockFactory(bookingLimitDatetime=valid_date)]
        )
        offers_factories.OfferFactory(stocks=[offers_factories.StockFactory()])
        offers_factories.OfferFactory(stocks=[offers_factories.StockFactory(bookingLimitDatetime=not_valid_date)])

        query_args = {
            "order": "asc",
            "search-0-search_field": "BOOKING_LIMIT_DATE",
            "search-0-operator": operator,
            "search-0-date": datetime.date.today(),
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(should_be_displayed_offer.id)

    def test_list_offers_with_offerer_tag_filter_in(self, authenticated_client):
        tag = offerers_factories.OffererTagFactory()
        offer_to_display = offers_factories.OfferFactory()
        offer_to_display.venue.managingOfferer.tags = [tag]
        hidden_offer = offers_factories.OfferFactory()
        hidden_offer.venue.managingOfferer.tags = [offerers_factories.OffererTagFactory()]

        query_args = {
            "search-0-search_field": "OFFERER_TAG",
            "search-0-operator": "IN",
            "search-0-offerer_tags": tag.id,
        }

        # +1 because of reloading selected tag in the form when tag arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_with_offerer_tag_filter_not_in(self, authenticated_client):
        tag = offerers_factories.OffererTagFactory()
        offers_factories.OfferFactory().venue.managingOfferer.tags = [tag]
        offer_to_display = offers_factories.OfferFactory()

        query_args = {
            "search-0-search_field": "OFFERER_TAG",
            "search-0-operator": "NOT_IN",
            "search-0-offerer_tags": tag.id,
        }

        # +1 because of reloading selected tag in the form when tag arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_with_offerer_tag_filter_not_exist(self, authenticated_client):
        tag = offerers_factories.OffererTagFactory(name="plouf")
        offers_factories.OfferFactory().venue.managingOfferer.tags = [tag]
        offers_factories.OfferFactory().venue.managingOfferer.tags = [offerers_factories.OffererTagFactory()]
        offer_to_display = offers_factories.OfferFactory()

        query_args = {
            "search-0-search_field": "OFFERER_TAG",
            "search-0-operator": "NOT_EXIST",
            "search-0-offerer_tags": tag.id,
        }

        # +1 because of reloading selected tag in the form when tag arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_with_venue_criterion_filter_in(self, authenticated_client):
        should_be_displayed_offer = offers_factories.OfferFactory()
        tag = criteria_factories.CriterionFactory()
        should_be_displayed_offer.venue.criteria = [tag]
        hidden_offer = offers_factories.OfferFactory()
        hidden_offer.venue.criteria = [criteria_factories.CriterionFactory()]

        query_args = {
            "search-0-search_field": "VENUE_TAG",
            "search-0-operator": "IN",
            "search-0-criteria": tag.id,
        }

        # +1 because of reloading selected criterion in the form when criterion arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(should_be_displayed_offer.id)

    def test_list_offers_with_venue_criterion_filter_not_in(self, authenticated_client):
        tag = criteria_factories.CriterionFactory()
        offers_factories.OfferFactory().venue.criteria = [tag]
        offer_to_display = offers_factories.OfferFactory()
        offer_to_display.venue.criteria = [criteria_factories.CriterionFactory()]

        query_args = {
            "search-0-search_field": "VENUE_TAG",
            "search-0-operator": "NOT_IN",
            "search-0-criteria": tag.id,
        }

        # +1 because of reloading selected criterion in the form when criterion arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_with_venue_criterion_filter_no_exist(self, authenticated_client):
        tag = criteria_factories.CriterionFactory()
        offers_factories.OfferFactory().venue.criteria = [tag]
        offers_factories.OfferFactory().venue.criteria = [criteria_factories.CriterionFactory()]
        offer_to_display = offers_factories.OfferFactory()

        query_args = {
            "search-0-search_field": "VENUE_TAG",
            "search-0-operator": "NOT_EXIST",
            "search-0-criteria": tag.id,
        }

        # +1 because of reloading selected criterion in the form when criterion arg is set
        with assert_num_queries(self.expected_num_queries_with_results + 1):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_by_event_date(self, authenticated_client, offers):
        offers_factories.EventStockFactory(beginningDatetime=datetime.date.today() + datetime.timedelta(days=1))
        stock = offers_factories.EventStockFactory(beginningDatetime=datetime.date.today())
        offers_factories.EventStockFactory(beginningDatetime=datetime.date.today() - datetime.timedelta(days=1))

        query_args = {
            "search-0-search_field": "EVENT_DATE",
            "search-0-operator": "DATE_EQUALS",
            "search-0-date": datetime.date.today().isoformat(),
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {stock.offer.id}

    def test_list_offers_by_event_date_gte_only(self, authenticated_client, offers):
        # Query investigated for performance issue in PC-23801
        query_args = {
            "limit": "100",
            "search-0-search_field": "EVENT_DATE",
            "search-0-operator": "DATE_FROM",
            "search-0-status": offers_models.OfferStatus.ACTIVE.value,
            "search-0-integer": "",
            "search-0-string": "",
            "search-0-date": (datetime.date.today() + datetime.timedelta(days=2)).isoformat(),
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    @pytest.mark.parametrize(
        "operator,criteria_indexes,expected_offer_indexes,has_provider",
        [
            ("IN", [0], [0, 2], True),
            ("NOT_IN", [0], [1, 3], False),
            ("NOT_IN", [1], [0, 1, 3], True),
            ("NOT_IN", [2], [0, 1, 2, 3], True),
            ("NOT_IN", [0, 1], [1, 3], False),
            ("NOT_EXIST", [], [1], False),
        ],
    )
    def test_list_offers_by_criterion(
        self, authenticated_client, criteria, offers, operator, criteria_indexes, expected_offer_indexes, has_provider
    ):
        query_args = {
            "search-3-search_field": "TAG",
            "search-3-operator": operator,
            "search-3-criteria": [criteria[criterion_index].id for criterion_index in criteria_indexes],
        }
        with assert_num_queries(
            (self.expected_num_queries_with_provider if has_provider else self.expected_num_queries_with_results)
            + int(bool(criteria_indexes))
        ):  # +1 because of reloading selected criterion in the form when criteria arg is set
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[index].id for index in expected_offer_indexes}

    def test_list_offers_by_in_and_not_in_criteria(self, authenticated_client, criteria, offers):
        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "IN",
            "search-0-criteria": criteria[0].id,
            "search-2-search_field": "TAG",
            "search-2-operator": "NOT_IN",
            "search-2-criteria": criteria[1].id,
        }
        with assert_num_queries(
            self.expected_num_queries_with_provider + 2
        ):  # +2 because of reloading selected criterion in the form, in both filters
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[0].id}

    def test_list_offers_by_category(self, authenticated_client, offers):
        query_args = {
            "search-3-search_field": "CATEGORY",
            "search-3-operator": "IN",
            "search-3-category": pro_categories.LIVRE.id,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id, offers[2].id}

    def test_list_offers_by_subcategory(self, authenticated_client, offers):
        query_args = {
            "search-3-search_field": "SUBCATEGORY",
            "search-3-operator": "IN",
            "search-3-subcategory": subcategories.LIVRE_PAPIER.id,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[2].id}

    def test_list_offers_by_price(self, authenticated_client, offers):
        query_args = {
            "search-3-search_field": "PRICE",
            "search-3-operator": "GREATER_THAN_OR_EQUAL_TO",
            "search-3-price": 12.20,
        }
        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offers[0].id)

    def test_list_offers_by_price_multiple_stocks(self, authenticated_client):
        offer_with_multiple_stocks_valid_and_not_valid = offers_factories.OfferFactory()

        offers_factories.StockFactory(
            price=15,
            offer=offer_with_multiple_stocks_valid_and_not_valid,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )

        offers_factories.StockFactory(
            price=16,
            offer=offer_with_multiple_stocks_valid_and_not_valid,
            beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=2),
            bookingLimitDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=1),
        )

        query_args = {
            "search-3-search_field": "PRICE",
            "search-3-operator": "EQUALS",
            "search-3-price": 16,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_with_multiple_stocks_valid_and_not_valid.id)

    def test_list_offers_by_price_no_offer_is_valid(self, authenticated_client, offers):
        query_args = {
            "search-3-search_field": "PRICE",
            "search-3-operator": "GREATER_THAN_OR_EQUAL_TO",
            "search-3-price": 120000.20,
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 0

    def test_list_free_offers(self, authenticated_client):
        free_offer = offers_factories.StockFactory(price=0).offer
        offers_factories.StockFactory(price=0.1)

        query_args = {
            "search-0-search_field": "PRICE",
            "search-0-operator": "EQUALS",
            "search-0-price": 0,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(free_offer.id)

    def test_list_offers_by_department(self, authenticated_client, offers):
        query_args = {
            "search-3-search_field": "DEPARTMENT",
            "search-3-operator": "IN",
            "search-3-department": ["74", "47", "971"],
        }

        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[0].id, offers[2].id}

    def test_list_offers_by_region(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "REGION",
            "search-0-operator": "IN",
            "search-0-region": ["La Réunion", "Auvergne-Rhône-Alpes"],
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert {int(row["ID"]) for row in rows} == {offers[1].id, offers[2].id}

    def test_list_offers_with_mediations(self, authenticated_client):
        mediation = offers_factories.MediationFactory()
        offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "MEDIATION",
            "search-0-operator": "NULLABLE",
            "search-0-boolean": "true",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert {int(row["ID"]) for row in rows} == {mediation.offerId}

    def test_list_offers_without_mediations(self, authenticated_client):
        offers_factories.MediationFactory()
        offer = offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "MEDIATION",
            "search-0-operator": "NULLABLE",
            "search-0-boolean": "false",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert {int(row["ID"]) for row in rows} == {offer.id}

    def test_list_offers_by_venue(self, authenticated_client, offers):
        venue_id = offers[1].venueId
        query_args = {
            "search-3-search_field": "VENUE",
            "search-3-operator": "IN",
            "search-3-venue": venue_id,
        }
        with assert_num_queries(
            self.expected_num_queries_with_results + 1
        ):  # +1 because of reloading selected venue in the form
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    def test_list_offers_by_venue_type(self, authenticated_client):
        offer = offers_factories.OfferFactory(venue__venueTypeCode=offerers_schemas.VenueTypeCode.BOOKSTORE)
        offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "VENUE_TYPE",
            "search-0-operator": "IN",
            "search-0-venue_type": offer.venue.venueTypeCode.name,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offer.id}

    def test_list_offers_by_address(self, authenticated_client, offers):
        for offer in offers[:3]:
            offer.offererAddress = offerers_factories.OffererAddressFactory(offerer=offer.venue.managingOfferer)
            db.session.add(offer)
        db.session.flush()

        address_id = offers[2].offererAddress.addressId
        query_args = {
            "search-0-search_field": "ADDRESS",
            "search-0-operator": "IN",
            "search-0-address": address_id,
        }
        with assert_num_queries(
            self.expected_num_queries_with_results + 1
        ):  # +1 because of reloading selected address in the form
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[2].id}

    def test_list_offers_by_status(self, authenticated_client, offers):
        offer = offers_factories.OfferFactory(isActive=False)

        query_args = {
            "search-0-search_field": "STATUS",
            "search-0-operator": "IN",
            "search-0-status": "INACTIVE",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offer.id}

    def test_list_offers_by_offerer(self, authenticated_client, offers):
        offerer_id = offers[1].venue.managingOffererId
        query_args = {
            "search-3-search_field": "OFFERER",
            "search-3-operator": "IN",
            "search-3-offerer": offerer_id,
        }
        with assert_num_queries(
            self.expected_num_queries_with_results + 1
        ):  # +1 because of reloading selected offerer in the form
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    def test_list_offers_by_validation(self, authenticated_client, offers):
        status = offers[2].validation
        query_args = {
            "search-3-search_field": "VALIDATION",
            "search-3-operator": "IN",
            "search-3-validation": status.value,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[2].id}
        assert rows[0]["État"] == "• Rejetée"

    def test_list_offers_by_four_filters(self, authenticated_client, criteria, offers):
        criterion_id = criteria[1].id
        venue_id = offers[2].venueId

        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "IN",
            "search-0-criteria": criterion_id,
            "search-1-search_field": "CATEGORY",
            "search-1-operator": "IN",
            "search-1-category": pro_categories.LIVRE.id,
            "search-2-search_field": "DEPARTMENT",
            "search-2-operator": "IN",
            "search-2-department": "74",
            "search-3-search_field": "VENUE",
            "search-3-operator": "IN",
            "search-3-venue": venue_id,
        }
        with assert_num_queries(
            self.expected_num_queries_with_results + 2
        ):  # +2 because of reloading selected criterion and venue
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[2].id}

    @pytest.mark.parametrize(
        "order,expected_list",
        [
            ("", ["Offre 4", "Offre 3", "Offre 2", "Offre 1"]),
            ("asc", ["Offre 4", "Offre 3", "Offre 2", "Offre 1"]),
            ("desc", ["Offre 1", "Offre 2", "Offre 3", "Offre 4"]),
        ],
    )
    def test_list_offers_pending_from_validated_offerers_sorted_by_date(
        self, authenticated_client, order, expected_list
    ):
        # test results when clicking on pending offers link (home page)
        offers_factories.OfferFactory(
            validation=offers_models.OfferValidationStatus.PENDING,
            venue__managingOfferer=offerers_factories.NewOffererFactory(),
        )

        validated_venue = offerers_factories.VenueFactory()
        for days_ago in (2, 4, 1, 3):
            offers_factories.OfferFactory(
                name=f"Offre {days_ago}",
                dateCreated=date_utils.get_naive_utc_now() - datetime.timedelta(days=days_ago),
                validation=offers_models.OfferValidationStatus.PENDING,
                venue=validated_venue,
            )

        query_args = {
            "only_validated_offerers": "on",
            "sort": "dateCreated",
            "order": order,
            "search-3-search_field": "VALIDATION",
            "search-3-operator": "IN",
            "search-3-validation": offers_models.OfferValidationStatus.PENDING.value,
            "search-2-search_field": "VALIDATED_OFFERER",
            "search-2-operator": "EQUALS",
            "search-2-boolean": "true",
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    **query_args,
                )
            )
            assert response.status_code == 200

        # must be sorted, older first
        rows = html_parser.extract_table_rows(response.data)
        assert [row["Nom de l'offre"] for row in rows] == expected_list

    def test_list_offers_with_flagging_rules(self, authenticated_client):
        rule_1 = offers_factories.OfferValidationRuleFactory(name="Règle magique")
        rule_2 = offers_factories.OfferValidationRuleFactory(name="Règle moldue")
        offer = offers_factories.OfferFactory(
            validation=offers_models.OfferValidationStatus.PENDING,
            flaggingValidationRules=[rule_1, rule_2],
        )

        prediction_reason = "Cette offre ne respecte pas les règles de conformité car elle a un prix trop bas."
        offers_factories.OfferComplianceFactory(
            offer=offer,
            compliance_score=50,
            compliance_reasons=["stock_price", "offer_subcategory_id", "offer_description"],
            validation_status_prediction=offers_models.ComplianceValidationStatusPrediction.REJECTED,
            validation_status_prediction_reason=prediction_reason,
        )

        query_args = {
            "search-0-search_field": "VALIDATION",
            "search-0-operator": "IN",
            "search-0-validation": offers_models.OfferValidationStatus.PENDING.value,
        }

        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    **query_args,
                )
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Règles de conformité"] == ", ".join([rule_1.name, rule_2.name])
        assert rows[0]["Score data"] == "50"
        assert rows[0]["Prédiction de validité"] == "À rejeter Prédiction générée par IA • À vérifier"

        # Check tooltip associated with "Score data"
        tooltips = html_parser.get_soup(response.data).find_all(
            "i", class_="bi-info-circle", attrs={"data-bs-toggle": "tooltip", "data-bs-html": "true"}
        )
        assert len(tooltips) == 2
        score_tooltip_html = tooltips[0].attrs.get("data-bs-title")
        assert (
            html_parser.content_as_text(score_tooltip_html)
            == "Raison de score faible : Prix Sous-catégorie Description de l'offre"
        )
        prediction_tooltip_html = tooltips[1].attrs.get("data-bs-title")
        expected_prediction_text = f"Explication de la prédiction : {prediction_reason}"
        assert html_parser.content_as_text(prediction_tooltip_html) == expected_prediction_text

    def test_list_offers_by_no_tags(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "NOT_EXIST",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    def test_list_offers_by_no_tags_and_validation(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "NOT_EXIST",
            "search-2-search_field": "VALIDATION",
            "search-2-operator": "IN",
            "search-2-validation": offers_models.OfferValidationStatus.APPROVED.value,
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID"]) for row in rows) == {offers[1].id}

    def test_list_offers_by_no_tags_and_other_validation(self, authenticated_client, offers):
        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "NOT_EXIST",
            "search-2-search_field": "VALIDATION",
            "search-2-operator": "IN",
            "search-2-validation": offers_models.OfferValidationStatus.PENDING.value,
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 0

    def test_list_offers_has_provider(self, authenticated_client):
        provider = providers_factories.ProviderFactory()
        offers_factories.OfferFactory(name="good", idAtProvider="pouet", lastProvider=provider)
        offers_factories.OfferFactory(name="bad")
        query_args = {
            "search-0-search_field": "SYNCHRONIZED",
            "search-0-operator": "NULLABLE",
            "search-0-boolean": "true",
        }
        with assert_num_queries(self.expected_num_queries_with_provider):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_has_no_provider(self, authenticated_client):
        provider = providers_factories.ProviderFactory()
        offers_factories.OfferFactory(name="bad", idAtProvider="pouet", lastProvider=provider)
        offers_factories.OfferFactory(name="good")
        query_args = {
            "search-0-search_field": "SYNCHRONIZED",
            "search-0-operator": "NULLABLE",
            "search-0-boolean": "false",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_by_provider(self, authenticated_client):
        provider = providers_factories.ProviderFactory()
        provider2 = providers_factories.ProviderFactory()
        offers_factories.OfferFactory(name="good", idAtProvider="pouet", lastProvider=provider)
        offers_factories.OfferFactory(
            idAtProvider="pouet2",
            lastProvider=provider2,
        )
        offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "PROVIDER",
            "search-0-operator": "IN",
            "search-0-provider": str(provider.id),
        }
        with assert_num_queries(
            self.expected_num_queries_with_provider + 1
        ):  # +1 because of reloading selected providers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_by_not_provider(self, authenticated_client):
        provider = providers_factories.ProviderFactory()
        provider2 = providers_factories.ProviderFactory()
        offers_factories.OfferFactory(name="good", idAtProvider="pouet", lastProvider=provider)
        offers_factories.OfferFactory(
            idAtProvider="pouet2",
            lastProvider=provider2,
        )
        offers_factories.OfferFactory()
        query_args = {
            "search-0-search_field": "PROVIDER",
            "search-0-operator": "NOT_IN",
            "search-0-provider": str(provider2.id),
        }
        with assert_num_queries(
            self.expected_num_queries_with_provider + 1
        ):  # +1 because of reloading selected providers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_has_headline(self, authenticated_client):
        offers_factories.HeadlineOfferFactory(offer__name="good")
        offers_factories.OfferFactory(name="bad")
        query_args = {
            "search-0-search_field": "HEADLINE",
            "search-0-operator": "EQUALS",
            "search-0-boolean": "true",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_has_no_headline(self, authenticated_client):
        offers_factories.HeadlineOfferFactory(offer__name="bad")
        offers_factories.OfferFactory(name="good")
        query_args = {
            "search-0-search_field": "HEADLINE",
            "search-0-operator": "EQUALS",
            "search-0-boolean": "false",
        }
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    def test_list_offers_by_highlight_request(self, authenticated_client):
        highlight_request = highlight_factories.HighlightRequestFactory(offer__name="good")
        offers_factories.OfferFactory(name="bad")
        query_args = {
            "search-0-search_field": "HIGHLIGHT_REQUEST",
            "search-0-operator": "IN",
            "search-0-boolean": "true",
            "search-0-highlight": highlight_request.highlightId,
        }
        with assert_num_queries(self.expected_num_queries_with_results + 1):  # fetch highlights (selectinload: 1 query)
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "good"

    # === Error cases ===

    def test_list_offers_by_invalid_field(self, authenticated_client):
        query_args = {
            "search-0-search_field": "CATEGRY",
            "search-0-operator": "IN",
            "search-0-category": pro_categories.LIVRE.id,
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert html_parser.extract_alert(response.data) == "Le filtre CATEGRY est invalide."

    def test_list_offers_by_invalid_operand(self, authenticated_client):
        query_args = {
            "search-0-search_field": "REGION",
            "search-0-operator": "EQUALS",
            "search-0-region": "Bretagne",
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert (
            html_parser.extract_alert(response.data)
            == "L'opérateur « est égal à » n'est pas supporté par le filtre Région de l'offre."
        )

    @pytest.mark.parametrize(
        "search_field,operator,operand",
        [
            ("PRICE", "EQUALS", "price"),
            ("EAN", "EQUALS", "string"),
            ("REGION", "IN", "region"),
        ],
    )
    def test_list_offers_by_empty_field(self, authenticated_client, search_field, operator, operand):
        query_args = {
            "search-0-search_field": search_field,
            "search-0-operator": operator,
            f"search-0-{operand}": "",
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert "est vide" in html_parser.extract_alert(response.data)

    def test_list_offers_by_category_and_missing_date(self, authenticated_client):
        query_args = {
            "search-0-search_field": "CATEGORY",
            "search-0-operator": "IN",
            "search-0-category": pro_categories.LIVRE.id,
            "search-2-search_field": "CREATION_DATE",
            "search-2-operator": "DATE_FROM",
            "search-4-search_field": "BOOKING_LIMIT_DATE",
            "search-4-operator": "DATE_TO",
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert (
            html_parser.extract_alert(response.data)
            == "Le filtre « Date de création » est vide. Le filtre « Date limite de réservation » est vide."
        )

    def test_list_offers_by_invalid_criteria(self, authenticated_client):
        query_args = {
            "search-0-search_field": "TAG",
            "search-0-operator": "IN",
            "search-0-criteria": "A",
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert html_parser.extract_alert(response.data) == "Le filtre « Tag » est vide."

    def test_list_offers_using_invalid_operator(self, authenticated_client):
        query_args = {
            "search-0-search_field": "CATEGORY",
            "search-0-operator": "OUT",
            "search-0-category": "13",
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

        assert html_parser.extract_alert(response.data) == "L'opérateur OUT n'est pas supporté par le filtre CATEGORY."

    @pytest.mark.parametrize(
        "search_field,value",
        [
            ("ID", "12, 34, A"),
            ("EAN", "123"),
            ("EAN", "1234567890ABC"),
            ("VISA", "1234567890123"),
            ("VISA", "1234567ABC"),
        ],
    )
    def test_list_offers_by_invalid_format(self, authenticated_client, search_field, value):
        query_args = {
            "search-0-search_field": search_field,
            "search-0-operator": "EQUALS",
            "search-0-string": value,
        }
        with assert_num_queries(self.expected_num_queries_with_no_result):  # + rollback - fetch offers
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 400

    # === Result content ===

    @pytest.mark.parametrize("first_quantity,second_quantity,expected_remaining", [(10, 7, 12), (5, 7, 7)])
    def test_list_offers_check_stock_limited(
        self, client, support_pro_n2_admin, first_quantity, second_quantity, expected_remaining
    ):
        offer = offers_factories.OfferFactory()
        offers_factories.StockFactory(offer=offer, quantity=first_quantity, dnBookedQuantity=5)
        offers_factories.StockFactory(offer=offer, quantity=second_quantity, dnBookedQuantity=0)

        query_args = self._get_query_args_by_id(offer.id)
        client = client.with_bo_session_auth(support_pro_n2_admin)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        row = html_parser.extract_table_rows(response.data)[0]
        assert row["Stock réservé"] == "5"
        assert row["Stock restant"] == str(expected_remaining)

    @pytest.mark.parametrize("first_quantity,second_quantity", [(None, None), (None, 10), (5, None)])
    def test_list_offers_check_stock_unlimited(self, client, support_pro_n2_admin, first_quantity, second_quantity):
        offer = offers_factories.OfferFactory()
        offers_factories.StockFactory(offer=offer, quantity=first_quantity, dnBookedQuantity=5)
        offers_factories.StockFactory(offer=offer, quantity=second_quantity, dnBookedQuantity=0)

        query_args = self._get_query_args_by_id(offer.id)
        client = client.with_bo_session_auth(support_pro_n2_admin)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        row = html_parser.extract_table_rows(response.data)[0]
        assert row["Stock réservé"] == "5"
        assert row["Stock restant"] == "Illimité"

    def test_list_offers_check_stock_sold_out(self, client, support_pro_n2_admin):
        offer = offers_factories.EventOfferFactory()
        offers_factories.EventStockFactory(offer=offer, quantity=5, dnBookedQuantity=5)
        offers_factories.EventStockFactory(offer=offer, quantity=7, dnBookedQuantity=7)

        query_args = self._get_query_args_by_id(offer.id)
        client = client.with_bo_session_auth(support_pro_n2_admin)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        row = html_parser.extract_table_rows(response.data)[0]
        assert row["Stock réservé"] == "12"
        assert row["Stock restant"] == "0"

    def test_list_offers_check_stock_expired(self, client, support_pro_n2_admin):
        offer = offers_factories.EventOfferFactory()
        offers_factories.EventStockFactory(
            offer=offer,
            quantity=7,
            dnBookedQuantity=2,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )
        offers_factories.EventStockFactory(
            offer=offer,
            quantity=9,
            dnBookedQuantity=3,
            bookingLimitDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )
        offers_factories.EventStockFactory(
            offer=offer,
            quantity=None,  # unlimited
            dnBookedQuantity=7,
            bookingLimitDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=1),
        )
        query_args = self._get_query_args_by_id(offer.id)
        client = client.with_bo_session_auth(support_pro_n2_admin)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        row = html_parser.extract_table_rows(response.data)[0]
        assert row["Stock réservé"] == "12"
        assert row["Stock restant"] == "0"

    def test_list_offers_check_stock_not_active(self, client, support_pro_n2_admin):
        for status in offers_models.OfferValidationStatus:
            if status != offers_models.OfferValidationStatus.APPROVED:
                offers_factories.StockFactory(offer__validation=status)
        offers_factories.StockFactory(offer__venue__managingOfferer=offerers_factories.NewOffererFactory())
        offers_factories.StockFactory(offer__venue__managingOfferer__isActive=False)

        query_args = {
            "search-0-search_field": "CATEGORY",
            "search-0-operator": "IN",
            "search-0-category": pro_categories.FILM.id,
        }

        client = client.with_bo_session_auth(support_pro_n2_admin)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == len(offers_models.OfferValidationStatus) - 1 + 2
        for row in rows:
            assert row["Stock réservé"] == "0"
            assert row["Stock restant"] == "-"

    @pytest.mark.parametrize(
        "venue_factory,expected_price_1,expected_price_2",
        [
            (offerers_factories.VenueFactory, "10,00 €", "10,00 € - 15,00 €"),
            (
                offerers_factories.CaledonianVenueFactory,
                "10,00 € (1195 CFP)",
                "10,00 € (1195 CFP) - 15,00 € (1790 CFP)",
            ),
        ],
    )
    def test_list_offers_price_with_different_stocks(
        self, client, pro_fraud_admin, venue_factory, expected_price_1, expected_price_2
    ):
        offer = offers_factories.OfferFactory(venue=venue_factory())

        client = client.with_bo_session_auth(pro_fraud_admin)
        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Tarif"] == "-"

        offers_factories.StockFactory(offer=offer, price=10)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Tarif"] == expected_price_1

        offers_factories.StockFactory(offer=offer, price=15)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Tarif"] == expected_price_2

    @pytest.mark.parametrize(
        "postal_code, expected_event_date_1, expected_event_date_2",
        [
            ("75001", "31/12/2023", "14/06/2023 → 31/12/2023"),
            ("97100", "30/12/2023 → 31/12/2023", "14/06/2023 → 31/12/2023"),
            ("98800", "31/12/2023 → 01/01/2024", "15/06/2023 → 01/01/2024"),
        ],
    )
    def test_list_offers_with_different_stock_beginning_datetimes(
        self, authenticated_client, postal_code, expected_event_date_1, expected_event_date_2
    ):
        venue = offerers_factories.VenueFactory()
        offer = offers_factories.OfferFactory(
            subcategoryId=subcategories.FESTIVAL_CINE.id,
            venue=venue,
            offererAddress=offerers_factories.OffererAddressFactory(
                offerer=venue.managingOfferer,
                address__postalCode=postal_code,
                address__departmentCode=get_department_code_from_city_code(postal_code),
            ),
        )

        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) de l'évènement"] == ""

        offers_factories.StockFactory(offer=offer, beginningDatetime=datetime.datetime(2023, 12, 31, 3, 0))
        offers_factories.StockFactory(offer=offer, beginningDatetime=datetime.datetime(2023, 12, 31, 21, 0))
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) de l'évènement"] == expected_event_date_1

        offers_factories.StockFactory(offer=offer, beginningDatetime=datetime.datetime(2023, 6, 14, 21, 0))
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) de l'évènement"] == expected_event_date_2

    def test_list_offers_with_different_booking_limit_datetimes(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.FESTIVAL_CINE.id)
        query_args = self._get_query_args_by_id(offer.id)

        offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=30)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) limite(s) de réservation"] == ""

        closest_stock = offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=3)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) limite(s) de réservation"] == closest_stock.bookingLimitDatetime.strftime("%d/%m/%Y")

        furthest_stock = offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=20)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert (
            rows[0]["Date(s) limite(s) de réservation"]
            == f"{closest_stock.bookingLimitDatetime.strftime('%d/%m/%Y')} → {furthest_stock.bookingLimitDatetime.strftime('%d/%m/%Y')}"
        )

    def test_list_offers_with_overseas_booking_limit_datetime(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.FESTIVAL_CINE.id)
        query_args = self._get_query_args_by_id(offer.id)

        offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=30)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) limite(s) de réservation"] == ""

        closest_stock = offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=3)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Date(s) limite(s) de réservation"] == closest_stock.bookingLimitDatetime.strftime("%d/%m/%Y")

        furthest_stock = offers_factories.StockFactory(
            offer=offer, bookingLimitDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=20)
        )
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert (
            rows[0]["Date(s) limite(s) de réservation"]
            == f"{closest_stock.bookingLimitDatetime.strftime('%d/%m/%Y')} → {furthest_stock.bookingLimitDatetime.strftime('%d/%m/%Y')}"
        )

    def test_list_offers_with_offerer_confidence_rule(self, client, pro_fraud_admin):
        rule = offerers_factories.ManualReviewOffererConfidenceRuleFactory(offerer__name="Offerer")
        offer = offers_factories.OfferFactory(venue__managingOfferer=rule.offerer, venue__name="Venue")

        client = client.with_bo_session_auth(pro_fraud_admin)
        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Entité juridique"] == "Offerer Revue manuelle"
        assert rows[0]["Partenaire culturel"] == "Venue"

    def test_list_offers_with_venue_confidence_rule(self, client, pro_fraud_admin):
        rule = offerers_factories.ManualReviewVenueConfidenceRuleFactory(
            venue__name="Venue", venue__managingOfferer__name="Offerer"
        )
        offer = offers_factories.OfferFactory(venue=rule.venue)

        client = client.with_bo_session_auth(pro_fraud_admin)
        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert rows[0]["Entité juridique"] == "Offerer"
        assert rows[0]["Partenaire culturel"] == "Venue Revue manuelle"

    def test_list_offers_with_top_acteur_offerer(self, client, pro_fraud_admin):
        offer = offers_factories.OfferFactory(
            venue__managingOfferer__name="Offerer",
            venue__managingOfferer__tags=[
                offerers_factories.OffererTagFactory(name=offerers_constants.TOP_ACTEUR_TAG_NAME, label="Top Acteur"),
                offerers_factories.OffererTagFactory(name="test", label="Test"),
            ],
        )

        client = client.with_bo_session_auth(pro_fraud_admin)
        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Entité juridique"] == "Offerer Top Acteur"

    def test_list_offers_with_product_image(self, authenticated_client):
        product_mediation = offers_factories.ProductMediationFactory()
        offer = offers_factories.OfferFactory(product=product_mediation.product)

        query_args = self._get_query_args_by_id(offer.id)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        expected = f'''style="background-image: url('{product_mediation.url}')"'''
        assert expected.encode() in response.data

    def test_list_offers_with_mediation_image(self, authenticated_client):
        mediation = offers_factories.MediationFactory()

        query_args = self._get_query_args_by_id(mediation.offerId)
        with assert_num_queries(self.expected_num_queries_with_results):
            response = authenticated_client.get(url_for(self.endpoint, **query_args))
            assert response.status_code == 200

        expected = f'''style="background-image: url('{mediation.thumbUrl}')"'''
        assert expected.encode() in response.data


class EditOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.edit_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    @patch("pcapi.core.search.reindex_offer_ids")
    def test_update_offer_tags(self, mock_reindex_offer_ids, legit_user, authenticated_client, criteria):
        offer_to_edit = offers_factories.OfferFactory(
            name="A Very Specific Name That Is Longer",
            criteria=[criteria[0]],
        )
        chosen_ranking_weight = 22
        base_form = {"criteria": [criteria[0].id, criteria[1].id], "rankingWeight": chosen_ranking_weight}

        response = self.post_to_endpoint(authenticated_client, offer_id=offer_to_edit.id, form=base_form)
        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.list_offers")
        assert response.location == expected_url

        db.session.refresh(offer_to_edit)
        assert offer_to_edit.rankingWeight == chosen_ranking_weight
        assert set(offer_to_edit.criteria) == {criteria[0], criteria[1]}

        # offer should be reindexed
        mock_reindex_offer_ids.assert_called_once_with([offer_to_edit.id])
        mock_reindex_offer_ids.reset_mock()

        # New Update without rankingWeight
        base_form = {"criteria": [criteria[2].id, criteria[1].id], "rankingWeight": ""}
        response = self.post_to_endpoint(authenticated_client, offer_id=offer_to_edit.id, form=base_form)
        assert response.status_code == 303

        db.session.refresh(offer_to_edit)
        assert offer_to_edit.rankingWeight is None
        assert set(offer_to_edit.criteria) == {criteria[1], criteria[2]}

        # offer should be reindexed
        mock_reindex_offer_ids.assert_called_once_with([offer_to_edit.id])

    @patch("pcapi.core.search.reindex_offer_ids")
    def test_update_offer_tags_from_htmx(self, mock_reindex_offer_ids, legit_user, authenticated_client, criteria):
        offer_to_edit = offers_factories.OfferFactory(
            name="A Very Specific Name That Is Longer",
            criteria=[criteria[0]],
        )
        chosen_ranking_weight = 22
        base_form = {"criteria": [criteria[0].id, criteria[1].id], "rankingWeight": chosen_ranking_weight}

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_edit.id,
            form=base_form,
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        # ensure that the row is rendered
        cells = html_parser.extract_plain_row(response.data, id=f"offer-row-{offer_to_edit.id}")
        assert len(cells) == 25
        i = count()
        assert cells[next(i)] == ""  # Checkbox
        assert cells[next(i)] == (  # Actions
            "Voir le détail de l’offre Valider l'offre Rejeter l'offre Mettre l'offre en pause Taguer / Pondérer"
        )
        assert cells[next(i)] == ""  # Image
        assert cells[next(i)] == str(offer_to_edit.id)  # ID
        assert cells[next(i)] == offer_to_edit.name  # Nom de l'offre
        assert cells[next(i)] == ""  # EAN / Allociné ID
        assert cells[next(i)] == offer_to_edit.category.pro_label  # Catégorie
        assert cells[next(i)] == offer_to_edit.subcategory.pro_label  # Sous-catégorie
        assert cells[next(i)] == ""  # Règles de conformité
        assert cells[next(i)] == ""  # Score data
        assert cells[next(i)] == ""  # Prédiction du validation_status (data)
        assert cells[next(i)] == "-"  # Tarif
        assert cells[next(i)] in {
            f"{criteria[0].name} {criteria[1].name}",
            f"{criteria[1].name} {criteria[0].name}",
        }  # Tag
        assert cells[next(i)] == ""  # Date(s) de l'évènement
        assert cells[next(i)] == ""  # Date(s) limite(s) de réservation
        assert cells[next(i)] == ""  # Créateur de l'offre
        assert cells[next(i)] == "22"  # Pondération
        assert cells[next(i)] == "• Validée"  # État
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Date de création
        assert cells[next(i)] == ""  # Dérnière validation
        assert cells[next(i)] == offer_to_edit.offererAddress.address.departmentCode  # Département
        assert cells[next(i)] == offer_to_edit.venue.managingOfferer.name  # Entité juridique
        assert cells[next(i)] == offer_to_edit.venue.name  # Partenaire culturel
        assert cells[next(i)] == "Voir toutes les offres"  # Offres du partenaire culturel
        assert cells[next(i)] == ""  # Partenaire technique


class GetEditOfferFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_edit_offer_form"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    def test_get_edit_form_test(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory()

        form_url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(2):  # session + tested_query
            response = authenticated_client.get(form_url)
            assert response.status_code == 200


class GetBatchEditOfferFormTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.get_batch_edit_offer_form"
    endpoint_kwargs = {"offer_ids": "1,2"}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    def test_get_edit_form_with_values_test(self, legit_user, authenticated_client, criteria):
        offers = offers_factories.OfferFactory.create_batch(
            3, subcategoryId=subcategories.LIVRE_PAPIER.id, criteria=[criteria[2]], rankingWeight=22
        )
        selected_offers = offers[:-1]  # select all but the last offer

        offer_ids = ",".join(str(offer.id) for offer in selected_offers)
        base_form = {
            "object_ids": offer_ids,
            "criteria": [criteria[2].id],
            "rankingWeight": "0",
        }  # as it is already get in the modal form

        response = self._update_offers_form(authenticated_client, base_form)
        assert response.status_code == 200

        # Edit N°1 - The two first offers
        base_form["criteria"].extend([criteria[0].id, criteria[1].id])
        response = self._update_offers(authenticated_client, base_form)
        assert response.status_code == 200
        for offer in offers[:-1]:
            assert set(offer.criteria) == set(criteria[:3])
            assert offer.rankingWeight is None

        # Edit N°2 - Only the last offer
        choosen_ranking_weight = 12
        base_form = {
            "criteria": [criteria[2].id, criteria[3].id],
            "rankingWeight": choosen_ranking_weight,
        }  # new set of criteria
        response = self._update_offer(authenticated_client, offers[-1], base_form)
        assert response.status_code == 303

        assert offers[0].rankingWeight is None
        assert offers[1].rankingWeight is None
        assert offers[2].rankingWeight == choosen_ranking_weight

        assert set(offers[0].criteria) == set(criteria[:3])
        assert set(offers[1].criteria) == set(criteria[:3])
        assert set(offers[2].criteria) == set(criteria[2:])

    def _update_offers_form(self, authenticated_client, form):
        edit_url = url_for("backoffice_web.offer.list_offers")
        authenticated_client.get(edit_url)

        url = url_for(self.endpoint)
        form["csrf_token"] = g.get("csrf_token", "")

        return authenticated_client.post(url, form=form)

    def _update_offers(self, authenticated_client, form):
        url = url_for("backoffice_web.offer.batch_edit_offer")
        form["csrf_token"] = g.get("csrf_token", "")

        return authenticated_client.post(url, form=form)

    def _update_offer(self, authenticated_client, offer, form):
        url = url_for("backoffice_web.offer.edit_offer", offer_id=offer.id)
        form["csrf_token"] = g.get("csrf_token", "")

        return authenticated_client.post(url, form=form)


class BatchEditOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.batch_edit_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_batch_edit_offer(self, mock_async_index, legit_user, authenticated_client, criteria):
        offers = offers_factories.OfferFactory.create_batch(3)
        parameter_ids = ",".join(str(offer.id) for offer in offers)
        chosen_ranking_weight = 2
        base_form = {
            "criteria": [criteria[0].id, criteria[1].id],
            "rankingWeight": chosen_ranking_weight,
            "object_ids": parameter_ids,
        }

        # Expected queries:
        # 1 x user_session + user
        # 1 x all offers
        # 1 x all criteria
        # 1 x update offers (for 3 offers)
        # 1 x insert into offer_criterion (for 3 insertions)
        # 1 x re-fetch to render updated rows (3 queries)
        response = self.post_to_endpoint(authenticated_client, form=base_form, expected_num_queries=8)
        assert response.status_code == 200
        # ensure rows are rendered
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[0].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[1].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[2].id}")

        for offer in offers:
            assert offer.rankingWeight == chosen_ranking_weight
            assert len(offer.criteria) == 2
            assert criteria[0] in offer.criteria
            assert criteria[2] not in offer.criteria

        mock_async_index.assert_called_once_with(
            [offer.id for offer in offers],
            reason=IndexationReason.OFFER_BATCH_UPDATE,
        )


class ListAlgoliaOffersTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.list_algolia_offers"
    needed_permission = perm_models.Permissions.READ_OFFERS

    # Use assert_num_queries() instead of assert_no_duplicated_queries() which does not detect one extra query caused
    # by a field added in the jinja template.
    # - fetch session + user (1 query)
    # - fetch offers with joinedload including extra data (1 query)
    expected_num_queries_with_no_result = 2
    # - fetch addresses (2 x selectinload: 2 queries)
    expected_num_queries_with_results = expected_num_queries_with_no_result + 2

    def test_list_offers_without_filter(self, authenticated_client, offers):
        # no filter => no query to fetch offers
        with assert_num_queries(self.expected_num_queries_with_no_result - 1):
            response = authenticated_client.get(url_for(self.endpoint))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_offers_by_query(self, authenticated_client):
        offer_to_display = offers_factories.OfferFactory(name="display")
        offers_factories.OfferFactory(name="hide")

        query_args = {"algolia_search": "display", "limit": 100}

        with patch(
            "pcapi.routes.backoffice.offers.blueprint.search_offer_ids", side_effect=((offer_to_display.id,),)
        ) as algolia_mock:
            with assert_num_queries(self.expected_num_queries_with_results):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200
            algolia_mock.assert_called_once_with(query="display", facetFilters=[], numericFilters=[], count=101)

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_by_query_with_filters(self, authenticated_client):
        offer_to_display = offers_factories.OfferFactory(name="display")
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "display",
            "limit": 1000,
            "search-0-search_field": "CATEGORY",
            "search-0-operator": "IN",
            "search-0-category": ["CARTE_JEUNES", "CINEMA"],
            "search-1-search_field": "EAN",
            "search-1-operator": "EQUALS",
            "search-1-string": "1234567890123",
        }

        with patch(
            "pcapi.routes.backoffice.offers.blueprint.search_offer_ids", side_effect=((offer_to_display.id,),)
        ) as algolia_mock:
            with assert_num_queries(self.expected_num_queries_with_results):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200
            algolia_mock.assert_called_once_with(
                query="display",
                count=1001,
                facetFilters=[
                    [
                        "offer.subcategoryId:CARTE_CINE_ILLIMITE",
                        "offer.subcategoryId:CARTE_CINE_MULTISEANCES",
                        "offer.subcategoryId:CARTE_JEUNES",
                        "offer.subcategoryId:CINE_PLEIN_AIR",
                        "offer.subcategoryId:CINE_VENTE_DISTANCE",
                        "offer.subcategoryId:EVENEMENT_CINE",
                        "offer.subcategoryId:FESTIVAL_CINE",
                        "offer.subcategoryId:SEANCE_CINE",
                    ],
                    "offer.ean:1234567890123",
                ],
                numericFilters=[],
            )

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_with_negative_filters(self, authenticated_client):
        offer_to_display = offers_factories.OfferFactory(name="display")
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "display",
            "limit": 1000,
            "search-0-search_field": "CATEGORY",
            "search-0-operator": "NOT_IN",
            "search-0-category": ["CARTE_JEUNES", "CINEMA"],
            "search-1-search_field": "EAN",
            "search-1-operator": "NOT_EQUALS",
            "search-1-string": "1234567890123",
        }

        with patch(
            "pcapi.routes.backoffice.offers.blueprint.search_offer_ids", side_effect=((offer_to_display.id,),)
        ) as algolia_mock:
            with assert_num_queries(self.expected_num_queries_with_results):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200
            algolia_mock.assert_called_once_with(
                query="display",
                count=1001,
                facetFilters=[
                    "offer.subcategoryId:-CARTE_CINE_ILLIMITE",
                    "offer.subcategoryId:-CARTE_CINE_MULTISEANCES",
                    "offer.subcategoryId:-CARTE_JEUNES",
                    "offer.subcategoryId:-CINE_PLEIN_AIR",
                    "offer.subcategoryId:-CINE_VENTE_DISTANCE",
                    "offer.subcategoryId:-EVENEMENT_CINE",
                    "offer.subcategoryId:-FESTIVAL_CINE",
                    "offer.subcategoryId:-SEANCE_CINE",
                    "offer.ean:-1234567890123",
                ],
                numericFilters=[],
            )

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)

    def test_list_offers_filter_on_price(self, authenticated_client):
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "",
            "limit": 1000,
            "search-0-search_field": "PRICE",
            "search-0-operator": "NUMBER_EQUALS",
            "search-0-price": "15",
            "search-1-search_field": "PRICE",
            "search-1-operator": "GREATER_THAN_OR_EQUAL_TO",
            "search-1-price": "15,12",
            "search-2-search_field": "PRICE",
            "search-2-operator": "LESS_THAN",
            "search-2-price": "12.34",
        }

        with patch("pcapi.routes.backoffice.offers.blueprint.search_offer_ids", return_value=[]) as algolia_mock:
            # no offers from algolia therefore no request to retrieve their fields
            with assert_num_queries(self.expected_num_queries_with_no_result - 1):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200

            algolia_mock.assert_called_once_with(
                query="",
                count=1001,
                facetFilters=[],
                numericFilters=[
                    "offer.prices=15",
                    "offer.prices>=15.12",
                    "offer.prices<12.34",
                ],
            )

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_offers_filter_on_date(self, authenticated_client):
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "",
            "limit": 1000,
            "search-0-search_field": "DATE",
            "search-0-operator": "DATE_FROM",
            "search-0-date": "2024-10-01",
            "search-1-search_field": "DATE",
            "search-1-operator": "DATE_TO",
            "search-1-date": "2024-10-01",
            "search-2-search_field": "DATE",
            "search-2-operator": "DATE_EQUALS",
            "search-2-date": "2024-10-01",
        }

        with patch("pcapi.routes.backoffice.offers.blueprint.search_offer_ids", return_value=[]) as algolia_mock:
            # no offers from algolia therefore no request to retrieve their fields
            with assert_num_queries(self.expected_num_queries_with_no_result - 1):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200

            algolia_mock.assert_called_once_with(
                query="",
                count=1001,
                facetFilters=[],
                numericFilters=[
                    "offer.dates>=1727733600",
                    "offer.dates<=1727820000",
                    "offer.dates>=1727733600",
                    "offer.dates<=1727820000",
                ],
            )

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_offers_filter_on_offerer(self, authenticated_client):
        venue1 = offerers_factories.VenueFactory()
        venue2 = offerers_factories.VenueFactory(managingOfferer=venue1.managingOfferer)
        offerers_factories.VenueFactory()
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "",
            "limit": 1000,
            "search-0-search_field": "OFFERER",
            "search-0-operator": "IN",
            "search-0-offerer": venue1.managingOffererId,
        }

        with patch("pcapi.routes.backoffice.offers.blueprint.search_offer_ids", return_value=[]) as algolia_mock:
            # one more request to get the venues from offerer id
            # one more request to cherck the offerer id
            # no offers from algolia therefore no request to retrieve their fields
            with assert_num_queries(self.expected_num_queries_with_no_result + 1):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200

            # handle venue order randomness to avoid flakyness
            try:
                algolia_mock.assert_called_once_with(
                    query="",
                    count=1001,
                    facetFilters=[[f"venue.id:{venue1.id}", f"venue.id:{venue2.id}"]],
                    numericFilters=[],
                )
            except AssertionError:
                algolia_mock.assert_called_once_with(
                    query="",
                    count=1001,
                    facetFilters=[[f"venue.id:{venue2.id}", f"venue.id:{venue1.id}"]],
                    numericFilters=[],
                )

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_offers_on_show_type(self, authenticated_client):
        offer_to_display = offers_factories.OfferFactory(name="display")
        offers_factories.OfferFactory(name="hide")

        query_args = {
            "algolia_search": "display",
            "limit": 1000,
            "search-0-search_field": "SHOW_TYPE",
            "search-0-operator": "IN",
            "search-0-show_type": ["Cirque"],
        }

        with patch(
            "pcapi.routes.backoffice.offers.blueprint.search_offer_ids", return_value=(offer_to_display.id,)
        ) as algolia_mock:
            with assert_num_queries(self.expected_num_queries_with_results):
                response = authenticated_client.get(url_for(self.endpoint, **query_args))
                assert response.status_code == 200
            algolia_mock.assert_called_once_with(
                query="display",
                count=1001,
                facetFilters=[
                    [
                        "offer.showType:Cirque",
                    ]
                ],
                numericFilters=[],
            )

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["ID"] == str(offer_to_display.id)


class ValidateOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.validate_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_validate_offer_with_stocks(self, legit_user, authenticated_client):
        offer_to_validate = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.REJECTED)
        offers_factories.StockFactory(offer=offer_to_validate, price=10.1)
        offers_factories.StockFactory(offer=offer_to_validate, price=1.01)

        response = self.post_to_endpoint(authenticated_client, offer_id=offer_to_validate.id)
        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.list_offers")
        assert response.location == expected_url

        db.session.refresh(offer_to_validate)
        assert offer_to_validate.isActive is True
        assert offer_to_validate.validation == offers_models.OfferValidationStatus.APPROVED
        assert offer_to_validate.lastValidationType == OfferValidationType.MANUAL
        assert offer_to_validate.lastValidationDate.date() == datetime.date.today()
        assert offer_to_validate.lastValidationPrice == decimal.Decimal("10.1")

    def test_validate_offer_with_stocks_using_htmx(self, legit_user, authenticated_client):
        offer_to_validate = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.REJECTED)
        offers_factories.StockFactory(offer=offer_to_validate, price=10.1)
        offers_factories.StockFactory(offer=offer_to_validate, price=1.01)

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_validate.id,
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        # ensure that the row is rendered
        cells = html_parser.extract_plain_row(response.data, id=f"offer-row-{offer_to_validate.id}")
        assert len(cells) == 25
        i = count()
        assert cells[next(i)] == ""  # Checkbox
        assert cells[next(i)] == (  # Actions
            "Voir le détail de l’offre Valider l'offre Rejeter l'offre Mettre l'offre en pause Taguer / Pondérer"
        )
        assert cells[next(i)] == ""  # Image
        assert cells[next(i)] == str(offer_to_validate.id)  # ID
        assert cells[next(i)] == offer_to_validate.name  # Nom de l'offre
        assert cells[next(i)] == ""  # EAN / Allociné ID
        assert cells[next(i)] == offer_to_validate.category.pro_label  # Catégorie
        assert cells[next(i)] == offer_to_validate.subcategory.pro_label  # Sous-catégorie
        assert cells[next(i)] == ""  # Règles de conformité
        assert cells[next(i)] == ""  # Score data
        assert cells[next(i)] == ""  # Predicition du validation_status (data)
        assert cells[next(i)] == "1,01 € - 10,10 €"  # Tarif
        assert cells[next(i)] == ""  # Tag
        assert cells[next(i)] == ""  # Date(s) de l'évènement
        assert cells[next(i)] == ""  # Date(s) limite(s) de réservation
        assert cells[next(i)] == ""  # Créateur de l'offre
        assert cells[next(i)] == ""  # Pondération
        assert cells[next(i)] == "• Validée"  # État
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Date de création
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Dérnière validation
        assert cells[next(i)] == offer_to_validate.offererAddress.address.departmentCode  # Département
        assert cells[next(i)] == offer_to_validate.venue.managingOfferer.name  # Entité juridique
        assert cells[next(i)] == offer_to_validate.venue.name  # Partenaire culturel
        assert cells[next(i)] == "Voir toutes les offres"  # Offres du partenaire culturel
        assert cells[next(i)] == ""  # Partenaire technique

        db.session.refresh(offer_to_validate)
        assert offer_to_validate.isActive is True
        assert offer_to_validate.validation == offers_models.OfferValidationStatus.APPROVED
        assert offer_to_validate.lastValidationType == OfferValidationType.MANUAL
        assert offer_to_validate.lastValidationDate.date() == datetime.date.today()
        assert offer_to_validate.lastValidationPrice == decimal.Decimal("10.1")

    def test_validate_offer_without_stocks(self, legit_user, authenticated_client):
        offer_to_validate = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.REJECTED)

        response = self.post_to_endpoint(authenticated_client, offer_id=offer_to_validate.id)
        assert response.status_code == 303

        assert offer_to_validate.isActive is True
        assert offer_to_validate.lastValidationType == OfferValidationType.MANUAL
        assert offer_to_validate.lastValidationPrice is None
        assert offer_to_validate.validation == offers_models.OfferValidationStatus.APPROVED


class GetValidateOfferFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_validate_offer_form"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_get_validate_form_test(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory()

        form_url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(2):  # session + tested_query
            response = authenticated_client.get(form_url)
            assert response.status_code == 200


class RejectOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.reject_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_reject_offer(self, legit_user, authenticated_client):
        offer_to_reject = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.APPROVED)
        confirmed_booking = bookings_factories.BookingFactory(
            user=users_factories.BeneficiaryFactory(),
            stock__offer=offer_to_reject,
            status=BookingStatus.CONFIRMED,
        )

        response = self.post_to_endpoint(authenticated_client, offer_id=offer_to_reject.id)
        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.list_offers")
        assert response.location == expected_url

        assert offer_to_reject.isActive is False
        assert offer_to_reject.validation == offers_models.OfferValidationStatus.REJECTED
        assert offer_to_reject.lastValidationType == OfferValidationType.MANUAL
        assert offer_to_reject.lastValidationDate.date() == datetime.date.today()
        assert offer_to_reject.lastValidationPrice is None

        assert len(mails_testing.outbox) == 2
        assert mails_testing.outbox[0]["To"] == confirmed_booking.user.email
        assert mails_testing.outbox[0]["template"] == dataclasses.asdict(
            TransactionalEmail.BOOKING_CANCELLATION_BY_PRO_TO_BENEFICIARY.value
        )
        assert mails_testing.outbox[0]["params"]["REJECTED"] is True
        assert mails_testing.outbox[1]["To"] == offer_to_reject.venue.bookingEmail
        assert mails_testing.outbox[1]["template"] == dataclasses.asdict(
            TransactionalEmail.OFFER_VALIDATED_TO_REJECTED_TO_PRO.value
        )

    def test_reject_offer_with_htmx(self, legit_user, authenticated_client):
        offer_to_reject = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.APPROVED)
        confirmed_booking = bookings_factories.BookingFactory(
            user=users_factories.BeneficiaryFactory(),
            stock__offer=offer_to_reject,
            status=BookingStatus.CONFIRMED,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_reject.id,
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        # ensure that the row is rendered
        cells = html_parser.extract_plain_row(response.data, id=f"offer-row-{offer_to_reject.id}")
        assert len(cells) == 25
        i = count()
        assert cells[next(i)] == ""  # Checkbox
        assert cells[next(i)] == (  # Actions
            "Voir le détail de l’offre Valider l'offre Rejeter l'offre Publier l'offre Taguer / Pondérer"
        )
        assert cells[next(i)] == ""  # Image
        assert cells[next(i)] == str(offer_to_reject.id)  # ID
        assert cells[next(i)] == offer_to_reject.name  # Nom de l'offre
        assert cells[next(i)] == ""  # EAN / Allociné ID
        assert cells[next(i)] == offer_to_reject.category.pro_label  # Catégorie
        assert cells[next(i)] == offer_to_reject.subcategory.pro_label  # Sous-catégorie
        assert cells[next(i)] == ""  # Règles de conformité
        assert cells[next(i)] == ""  # Score data
        assert cells[next(i)] == ""  # Predicition du validation_status (data)
        assert cells[next(i)] == "10,10 €"  # Tarif
        assert cells[next(i)] == ""  # Tag
        assert cells[next(i)] == ""  # Date(s) de l'évènement
        assert cells[next(i)] == ""  # Date(s) limite(s) de réservation
        assert cells[next(i)] == ""  # Créateur de l'offre
        assert cells[next(i)] == ""  # Pondération
        assert cells[next(i)] == "• Rejetée"  # État
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Date de création
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Dérnière validation
        assert cells[next(i)] == offer_to_reject.offererAddress.address.departmentCode  # Département
        assert cells[next(i)] == offer_to_reject.venue.managingOfferer.name  # Entité juridique
        assert cells[next(i)] == offer_to_reject.venue.name  # Partenaire culturel
        assert cells[next(i)] == "Voir toutes les offres"  # Offres du partenaire culturel
        assert cells[next(i)] == ""  # Partenaire technique

        assert offer_to_reject.isActive is False
        assert offer_to_reject.validation == offers_models.OfferValidationStatus.REJECTED
        assert offer_to_reject.lastValidationType == OfferValidationType.MANUAL
        assert offer_to_reject.lastValidationDate.date() == datetime.date.today()
        assert offer_to_reject.lastValidationPrice is None

        assert len(mails_testing.outbox) == 2
        assert mails_testing.outbox[0]["To"] == confirmed_booking.user.email
        assert mails_testing.outbox[0]["template"] == dataclasses.asdict(
            TransactionalEmail.BOOKING_CANCELLATION_BY_PRO_TO_BENEFICIARY.value
        )
        assert mails_testing.outbox[0]["params"]["REJECTED"] is True
        assert mails_testing.outbox[1]["To"] == offer_to_reject.venue.bookingEmail
        assert mails_testing.outbox[1]["template"] == dataclasses.asdict(
            TransactionalEmail.OFFER_VALIDATED_TO_REJECTED_TO_PRO.value
        )


class GetRejectOfferFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_reject_offer_form"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_get_edit_form_test(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory()

        form_url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(2):  # session + tested_query
            response = authenticated_client.get(form_url)
            assert response.status_code == 200


class BatchOfferValidateTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.batch_validate_offers"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    @pytest.mark.parametrize(
        "venue_email, pro_email, expected_recipient, additional_queries",
        [
            ("venue@example.com", "pro.user@example.com", "venue@example.com", False),
            (None, "pro.user@example.com", "pro.user@example.com", True),
        ],
    )
    def test_batch_validate_offers(
        self, legit_user, authenticated_client, venue_email, pro_email, expected_recipient, additional_queries
    ):
        user_offerer = offerers_factories.UserOffererFactory(user__email=pro_email)
        offerers_factories.NewUserOffererFactory(offerer=user_offerer.offerer)  # not attached
        venue = offerers_factories.VenueFactory(managingOfferer=user_offerer.offerer, bookingEmail=venue_email)
        offers = offers_factories.OfferFactory.create_batch(
            3, validation=offers_models.OfferValidationStatus.DRAFT, venue=venue
        )

        yesterday = datetime.datetime.now(datetime.UTC) - datetime.timedelta(hours=24)
        offers.append(
            offers_factories.OfferFactory(
                validation=offers_models.OfferValidationStatus.REJECTED, venue=venue, publicationDatetime=yesterday
            )
        )

        for offer in offers:
            offers_factories.StockFactory(offer=offer, price=10.1)
        parameter_ids = ",".join(str(offer.id) for offer in offers)

        expected_queries = 1  # user + session
        expected_queries += 1  # select offers
        expected_queries += 1  # update offers
        expected_queries += 1  # select offerer
        expected_queries += 1  # select offerer_address
        expected_queries += 1  # select venue

        if additional_queries:
            expected_queries += len(offers)  # for each offer, fetch its recipients
        response = self.post_to_endpoint(
            authenticated_client, form={"object_ids": parameter_ids}, expected_num_queries=expected_queries
        )

        assert response.status_code == 200

        for offer in offers:
            db.session.refresh(offer)

            # ensure rows are rendered
            html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offer.id}")

            assert offer.lastValidationDate.strftime("%d/%m/%Y") == datetime.date.today().strftime("%d/%m/%Y")
            assert offer.isActive is True
            assert offer.lastValidationType is OfferValidationType.MANUAL
            assert offer.validation is offers_models.OfferValidationStatus.APPROVED
            assert offer.lastValidationAuthor == legit_user
            assert offer.lastValidationPrice == decimal.Decimal("10.1")

        assert len(mails_testing.outbox) == len(offers)
        for email_data in mails_testing.outbox:
            assert email_data["To"] == expected_recipient
            assert email_data["template"] == dataclasses.asdict(TransactionalEmail.OFFER_APPROVAL_TO_PRO.value)
            assert email_data["params"]["VENUE_NAME"] == venue.name


class BatchOfferRejectTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.batch_reject_offers"
    needed_permission = perm_models.Permissions.PRO_FRAUD_ACTIONS

    def test_batch_reject_offers(self, legit_user, authenticated_client):
        beneficiary = users_factories.BeneficiaryFactory()
        draft_offer = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.DRAFT)
        pending_offer = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.PENDING)
        confirmed_offer = offers_factories.OfferFactory(validation=offers_models.OfferValidationStatus.APPROVED)
        confirmed_booking = bookings_factories.BookingFactory(
            user=beneficiary, stock__offer=confirmed_offer, status=BookingStatus.CONFIRMED
        )
        parameter_ids = ",".join(str(offer.id) for offer in [draft_offer, pending_offer, confirmed_offer])

        assert confirmed_booking.status == BookingStatus.CONFIRMED

        response = self.post_to_endpoint(authenticated_client, form={"object_ids": parameter_ids})

        assert confirmed_booking.status == BookingStatus.CANCELLED
        assert response.status_code == 200
        for offer in [draft_offer, pending_offer, confirmed_offer]:
            db.session.refresh(offer)
            assert offer.lastValidationDate.strftime("%d/%m/%Y") == datetime.date.today().strftime("%d/%m/%Y")
            assert offer.isActive is False
            assert offer.lastValidationType is OfferValidationType.MANUAL
            assert offer.validation is offers_models.OfferValidationStatus.REJECTED
            assert offer.lastValidationAuthor == legit_user

        assert len(mails_testing.outbox) == 4
        emails_dict = {email_data["To"]: email_data for email_data in mails_testing.outbox}
        assert confirmed_booking.user.email in emails_dict
        assert emails_dict[confirmed_booking.user.email]["template"] == dataclasses.asdict(
            TransactionalEmail.BOOKING_CANCELLATION_BY_PRO_TO_BENEFICIARY.value
        )
        assert emails_dict[confirmed_booking.user.email]["params"]["REJECTED"] is True
        assert draft_offer.venue.bookingEmail in emails_dict
        assert emails_dict[draft_offer.venue.bookingEmail]["template"] == dataclasses.asdict(
            TransactionalEmail.OFFER_REJECTION_TO_PRO.value
        )
        assert pending_offer.venue.bookingEmail in emails_dict
        assert emails_dict[pending_offer.venue.bookingEmail]["template"] == dataclasses.asdict(
            TransactionalEmail.OFFER_PENDING_TO_REJECTED_TO_PRO.value
        )
        assert confirmed_offer.venue.bookingEmail in emails_dict
        assert emails_dict[confirmed_offer.venue.bookingEmail]["template"] == dataclasses.asdict(
            TransactionalEmail.OFFER_VALIDATED_TO_REJECTED_TO_PRO.value
        )


class IndexOfferButtonTest(button_helpers.ButtonHelper):
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    button_label = "Resync. Algolia"

    @property
    def path(self):
        offer = offers_factories.OfferFactory()
        return url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)


class MoveOfferVenueButtonTest(button_helpers.ButtonHelper):
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    button_label = "Modifier le partenaire culturel"

    @property
    def path(self):
        venue = offerers_factories.VenueFactory()
        offerers_factories.VenueFactory(managingOfferer=venue.managingOfferer, pricing_point=venue)
        offer = offers_factories.EventOfferFactory(venue=venue)
        offers_factories.EventStockFactory(offer=offer)
        offers_factories.EventStockFactory(
            offer=offer,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=3),
            isSoftDeleted=True,
        )
        return url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)


class IndexOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.reindex"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT

    @mock.patch("pcapi.core.search.async_index_offer_ids")
    def test_index_offer(self, mocked_async_index_offer_ids, authenticated_client):
        offer = offers_factories.OfferFactory()

        response = self.post_to_endpoint(authenticated_client, offer_id=offer.id)
        assert response.status_code == 303

        redirected_response = authenticated_client.get(response.headers["location"])
        assert "La resynchronisation de l'offre a été demandée." in html_parser.extract_alert(redirected_response.data)

        mocked_async_index_offer_ids.assert_called()
        assert set(mocked_async_index_offer_ids.call_args[0][0]) == {offer.id}


@pytest.fixture(scope="function", name="venues_in_same_offerer")
def venues_in_same_offerer_fixture() -> tuple[offerers_models.Venue]:
    offerer = offerers_factories.OffererFactory()
    source_venue = offerers_factories.VenueFactory(managingOfferer=offerer, pricing_point="self")
    venue_with_same_pricing_point = offerers_factories.VenueFactory(managingOfferer=offerer, pricing_point=source_venue)
    venue_with_own_pricing_point = offerers_factories.VenueFactory(managingOfferer=offerer, pricing_point="self")
    venue_without_pricing_point = offerers_factories.VenueFactory(managingOfferer=offerer, pricing_point=None)
    return source_venue, venue_with_same_pricing_point, venue_with_own_pricing_point, venue_without_pricing_point


class EditOfferVenueTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.edit_offer_venue"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT

    def _test_move_event(
        self,
        mocked_async_index_offer_ids,
        authenticated_client,
        *,
        source_venue: offerers_models.Venue,
        destination_venue: offerers_models.Venue,
        move_offer_address: bool,
        notify_beneficiary: bool,
        with_pricings: bool = True,
        expected_error: str | None = None,
    ):
        offer = offers_factories.EventOfferFactory(venue=source_venue)
        original_address_id = offer.offererAddress.address.id

        stock1 = offers_factories.EventStockFactory(offer=offer)
        booking1_1 = bookings_factories.BookingFactory(stock=stock1)
        booking1_2 = bookings_factories.UsedBookingFactory(stock=stock1)
        finance_event1_2 = finance_factories.UsedBookingFinanceEventFactory(booking=booking1_2)
        if with_pricings:
            finance_factories.PricingFactory(event=finance_event1_2, booking=booking1_2)

        stock2 = offers_factories.EventStockFactory(offer=offer)
        booking2_1 = bookings_factories.UsedBookingFactory(stock=stock2)
        finance_event2_1 = finance_factories.UsedBookingFinanceEventFactory(booking=booking2_1)
        if with_pricings:
            finance_factories.PricingFactory(event=finance_event2_1, booking=booking2_1)
        booking2_2 = bookings_factories.CancelledBookingFactory(stock=stock2)

        # other objects to validate queries filters
        offers_factories.EventStockFactory(
            offer=offer,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=7),
            isSoftDeleted=True,
        )
        offers_factories.EventStockFactory(
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=7)
        )
        bookings_factories.ReimbursedBookingFactory()

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            form={
                "venue": destination_venue.id,
                "move_offer_address": "on" if move_offer_address else "",
                "notify_beneficiary": "on" if notify_beneficiary else "",
            },
        )
        assert response.status_code == 303

        if not expected_error:
            assert (
                html_parser.extract_alert(authenticated_client.get(response.location).data)
                == f"L'offre a été déplacée vers le partenaire culturel {destination_venue.name}"
            )

            mocked_async_index_offer_ids.assert_called_once_with(
                {offer.id}, reason=IndexationReason.OFFER_UPDATE, log_extra={"changes": {"venueId"}}
            )

            # Mail is sent only for in-going booking
            assert len(mails_testing.outbox) == (1 if notify_beneficiary else 0)

            expected_venue = destination_venue
        else:
            assert html_parser.extract_alert(authenticated_client.get(response.location).data) == expected_error

            mocked_async_index_offer_ids.assert_not_called()
            assert len(mails_testing.outbox) == 0

            expected_venue = source_venue

        assert offer.venueId == expected_venue.id
        assert booking1_1.venueId == expected_venue.id
        assert booking1_2.venueId == expected_venue.id
        assert booking2_1.venueId == expected_venue.id
        assert booking2_2.venueId == expected_venue.id
        assert finance_event1_2.venueId == expected_venue.id
        assert finance_event1_2.pricingPointId == expected_venue.current_pricing_point.id
        assert finance_event2_1.venueId == expected_venue.id
        assert finance_event2_1.pricingPointId == expected_venue.current_pricing_point.id
        if move_offer_address:
            assert offer.offererAddress.address.id == expected_venue.offererAddress.address.id
        else:
            assert offer.offererAddress.id != expected_venue.offererAddress.id
            assert offer.offererAddress.id != source_venue.offererAddress.id
            assert offer.offererAddress.address.id == original_address_id

    @pytest.mark.parametrize("move_offer_address, notify_beneficiary", [(False, False), (True, False), (True, True)])
    @patch("pcapi.core.search.async_index_offer_ids")
    def test_move_event_offer_when_venue_has_same_pricing_point(
        self,
        mocked_async_index_offer_ids,
        authenticated_client,
        venues_in_same_offerer,
        move_offer_address,
        notify_beneficiary,
    ):
        source_venue, venue_with_same_pricing_point, _, _ = venues_in_same_offerer
        self._test_move_event(
            mocked_async_index_offer_ids,
            authenticated_client,
            source_venue=source_venue,
            destination_venue=venue_with_same_pricing_point,
            move_offer_address=move_offer_address,
            notify_beneficiary=notify_beneficiary,
        )

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_move_event_offer_without_pricing_when_venue_has_different_pricing_point(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        source_venue, _, venue_with_own_pricing_point, _ = venues_in_same_offerer
        self._test_move_event(
            mocked_async_index_offer_ids,
            authenticated_client,
            source_venue=source_venue,
            destination_venue=venue_with_own_pricing_point,
            move_offer_address=True,
            notify_beneficiary=True,
            with_pricings=False,
        )

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_move_event_offer_with_pricings_when_venue_has_different_pricing_point(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        source_venue, _, venue_with_own_pricing_point, _ = venues_in_same_offerer
        self._test_move_event(
            mocked_async_index_offer_ids,
            authenticated_client,
            source_venue=source_venue,
            destination_venue=venue_with_own_pricing_point,
            move_offer_address=True,
            notify_beneficiary=True,
            expected_error="Le partenaire culturel de cette offre ne peut pas être modifié : "
            "Il existe des réservations valorisées sur un autre point de valorisation que celui du nouveau partenaire culturel",
        )

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_cant_move_when_destination_venue_has_no_pricing_point(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        # This case should not happen if the user only selects from venues dropdown list
        source_venue, _, _, venue_without_pricing_point = venues_in_same_offerer
        self._test_move_event(
            mocked_async_index_offer_ids,
            authenticated_client,
            source_venue=source_venue,
            destination_venue=venue_without_pricing_point,
            move_offer_address=True,
            notify_beneficiary=True,
            with_pricings=False,
            expected_error="Le partenaire culturel de cette offre ne peut pas être modifié : Ce partenaire culturel n'est pas éligible au transfert de l'offre",
        )

    def test_cant_move_when_offer_is_not_an_event(self, authenticated_client, venues_in_same_offerer):
        source_venue, destination_venue, _, _ = venues_in_same_offerer
        offer = offers_factories.ThingStockFactory(offer__venue=source_venue).offer

        response = self.post_to_endpoint(authenticated_client, offer_id=offer.id, form={"venue": destination_venue.id})

        assert response.status_code == 303
        assert (
            html_parser.extract_alert(authenticated_client.get(response.location).data)
            == "Le partenaire culturel de cette offre ne peut pas être modifié : L'offre n'est pas un évènement"
        )

    def test_cant_move_when_event_is_in_the_past(self, authenticated_client, venues_in_same_offerer):
        source_venue, destination_venue, _, _ = venues_in_same_offerer
        offer = offers_factories.EventOfferFactory(venue=source_venue)
        offers_factories.EventStockFactory.create_batch(
            2, offer=offer, beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(days=1)
        )
        offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(days=1)
        )

        response = self.post_to_endpoint(authenticated_client, offer_id=offer.id, form={"venue": destination_venue.id})

        assert response.status_code == 303
        assert (
            html_parser.extract_alert(authenticated_client.get(response.location).data)
            == "Le partenaire culturel de cette offre ne peut pas être modifié : L'évènement a déjà eu lieu pour 2 stocks"
        )

    def test_cant_move_when_reimbursed_bookings(self, authenticated_client, venues_in_same_offerer):
        source_venue, destination_venue, _, _ = venues_in_same_offerer
        stock = offers_factories.EventStockFactory(offer__venue=source_venue)
        bookings_factories.ReimbursedBookingFactory(stock=stock)

        response = self.post_to_endpoint(
            authenticated_client, offer_id=stock.offer.id, form={"venue": destination_venue.id}
        )

        assert response.status_code == 303
        assert (
            html_parser.extract_alert(authenticated_client.get(response.location).data)
            == "Le partenaire culturel de cette offre ne peut pas être modifié : 1 réservation est déjà remboursée sur cette offre"
        )

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_move_event_offer_with_price_category(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        source_venue, destination_venue, _, _ = venues_in_same_offerer

        offer = offers_factories.EventOfferFactory(venue=source_venue)
        gold_label = offers_factories.PriceCategoryLabelFactory(label="Gold", venue=offer.venue)
        gold_stock = offers_factories.EventStockFactory(
            offer=offer,
            priceCategory=offers_factories.PriceCategoryFactory(offer=offer, priceCategoryLabel=gold_label),
        )
        silver_label = offers_factories.PriceCategoryLabelFactory(label="Silver", venue=offer.venue)
        silver_stock = offers_factories.EventStockFactory(
            offer=offer,
            priceCategory=offers_factories.PriceCategoryFactory(offer=offer, priceCategoryLabel=silver_label),
        )
        destination_silver_label = offers_factories.PriceCategoryLabelFactory(label="Silver", venue=destination_venue)

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            form={"venue": destination_venue.id, "notify_beneficiary": ""},
        )
        assert response.status_code == 303

        assert offer.venueId == destination_venue.id
        assert gold_stock.priceCategory.priceCategoryLabel.venue == destination_venue
        assert gold_stock.priceCategory.priceCategoryLabel != gold_label
        assert gold_label.venue == source_venue
        assert silver_stock.priceCategory.priceCategoryLabel.venue == destination_venue
        assert silver_stock.priceCategory.priceCategoryLabel == destination_silver_label
        assert db.session.query(offers_models.PriceCategoryLabel).count() == 4

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_sould_move_event_offerer_address(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        source_venue, _, _, _ = venues_in_same_offerer
        destination_venue = offerers_factories.VenueFactory(
            managingOfferer=source_venue.managingOfferer, pricing_point=source_venue
        )

        self._test_move_event(
            mocked_async_index_offer_ids,
            authenticated_client,
            source_venue=source_venue,
            destination_venue=destination_venue,
            move_offer_address=True,
            notify_beneficiary=False,
        )

    @patch("pcapi.core.search.async_index_offer_ids")
    def test_sould_keep_event_offerer_address(
        self, mocked_async_index_offer_ids, authenticated_client, venues_in_same_offerer
    ):
        source_venue, destination_venue, _, _ = venues_in_same_offerer
        offerer_address = offerers_factories.OffererAddressFactory(offerer=source_venue.managingOfferer)
        offer = offers_factories.EventOfferFactory(venue=source_venue, offererAddress=offerer_address)

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            form={"venue": destination_venue.id, "move_offer_address": "", "notify_beneficiary": ""},
            follow_redirects=True,
        )

        assert response.status_code == 200
        assert (
            html_parser.extract_alert(response.data)
            == f"L'offre a été déplacée vers le partenaire culturel {destination_venue.name}"
        )

        assert offer.venueId == destination_venue.id
        assert offer.offererAddress == offerer_address


class GetOfferStockEditFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_offer_stock_edit_form"
    endpoint_kwargs = {"offer_id": 1, "stock_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    # session + check is_stock_editable + retrieve stock + get booking prices
    expected_num_queries = 4

    def test_get_stock_edit_form(self, authenticated_client):
        booking = bookings_factories.UsedBookingFactory(stock__offer__subcategoryId=subcategories.CONFERENCE.id)

        form_url = url_for(self.endpoint, offer_id=booking.stock.offer.id, stock_id=booking.stock.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(form_url)
            assert response.status_code == 200

    def test_get_stock_edit_form_non_event_offer(self, authenticated_client):
        stock = offers_factories.StockFactory()

        form_url = url_for(self.endpoint, offer_id=stock.offer.id, stock_id=stock.id)

        with assert_num_queries(self.expected_num_queries - 2):  # -2 no form generated
            response = authenticated_client.get(form_url)
            assert response.status_code == 200

        assert b"Ce stock n&#39;est pas \xc3\xa9ditable." in response.data

    def test_get_stock_edit_form_no_eligible_bookings(self, authenticated_client, app):
        booking = bookings_factories.BookingFactory(
            stock__offer__subcategoryId=subcategories.CONFERENCE.id, status=BookingStatus.CANCELLED
        )

        form_url = url_for(self.endpoint, offer_id=booking.stock.offer.id, stock_id=booking.stock.id)

        try:
            response = authenticated_client.get(form_url)
        finally:
            app.redis_client.delete(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK)

        assert response.status_code == 200
        assert b"Ce stock n&#39;est pas \xc3\xa9ditable." in response.data

    def test_get_stock_edit_form_running_cashflow_script_test(self, authenticated_client, app):
        booking = bookings_factories.UsedBookingFactory(stock__offer__subcategoryId=subcategories.CONFERENCE.id)

        form_url = url_for(self.endpoint, offer_id=booking.stock.offer.id, stock_id=booking.stock.id)
        app.redis_client.set(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK, "1", 600)

        try:
            response = authenticated_client.get(form_url)
        finally:
            app.redis_client.delete(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK)

        assert response.status_code == 200
        assert (
            "Le script de génération des cashflows est en cours, veuillez réessayer plus tard.".encode("utf-8")
            in response.data
        )


class ConfirmOfferStockTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.confirm_offer_stock"
    endpoint_kwargs = {"offer_id": 1, "stock_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    def test_confirm_edit_bookings_by_value(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("10"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("10"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("10"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": 2.1}
        )

        assert (
            '<span class="fw-bold">2 réservations :</span><span class="text-nowrap">10,00 €</span> → <span class="text-nowrap">2,10 €</span>'.encode()
            in response.data
        )
        assert (
            '<span class="fw-bold">1 réservation :</span><span class="text-nowrap">123,45 €</span> → <span class="text-nowrap">2,10 €</span>'.encode()
            in response.data
        )

    def test_confirm_edit_bookings_by_percent(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("10"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("10"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("10"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"percent": 12}
        )

        assert (
            '<span class="fw-bold">2 réservations :</span><span class="text-nowrap">10,00 €</span> → <span class="text-nowrap">8,80 €</span>'.encode()
            in response.data
        )
        assert (
            '<span class="fw-bold">1 réservation :</span><span class="text-nowrap">123,45 €</span> → <span class="text-nowrap">108,64 €</span>'.encode()
            in response.data
        )

    def test_confirm_edit_bookings_with_value_higher_than_bookings(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("10"),
        )
        bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("5"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": 7.5}
        )

        assert (
            "Cette modification amènerait à augmenter le prix de certaines réservations. Celles-ci ne seront pas changées".encode()
            in response.data
        )
        assert (
            '<span class="fw-bold">1 réservation :</span><span class="text-nowrap">5,00 €</span> → <span class="text-nowrap">5,00 €</span>'.encode()
            in response.data
        )

    def test_offer_stock_edit_cashfow_script_running(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )
        app.redis_client.set(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK, "1", 600)

        try:
            response = self.post_to_endpoint(
                authenticated_client,
                offer_id=offer.id,
                stock_id=event.booking.stock.id,
                form={"price": 50.1},
            )
        finally:
            app.redis_client.delete(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK)

        assert response.status_code == 200
        assert (
            html_parser.extract_alert(response.data)
            == "Le script de génération des cashflows est en cours, veuillez réessayer plus tard."
        )

    def test_offer_stock_edit_offer_not_an_event(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.LIVRE_PAPIER.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={"price": 50.1},
        )

        db.session.refresh(event.booking)

        assert response.status_code == 200
        assert html_parser.extract_alert(response.data) == "Ce stock n'est pas éditable."


class EditOfferStockTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.edit_offer_stock"
    endpoint_kwargs = {"offer_id": 1, "stock_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_OFFERS

    def test_offer_stock_edit_used_booking(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        booking_to_edit = bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )
        stock_untouched = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        booking_untouched = bookings_factories.UsedBookingFactory(
            stock=stock_untouched,
            amount=decimal.Decimal("123.45"),
        )
        cancelled_event = finance_factories.FinanceEventFactory(
            booking=booking_to_edit, venue=venue, pricingPoint=venue
        )
        cancelled_pricing = finance_factories.PricingFactory(
            event=cancelled_event,
            pricingPoint=venue,
        )
        later_booking = bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            stock__offer__subcategoryId=subcategories.CONFERENCE.id,
            stock__beginningDatetime=date_utils.get_naive_utc_now() + datetime.timedelta(hours=2),
        )
        later_event = finance_factories.FinanceEventFactory(
            booking=later_booking,
            venue=venue,
            pricingPoint=venue,
            pricingOrderingDate=later_booking.stock.beginningDatetime,
        )
        later_pricing = finance_factories.PricingFactory(
            event=later_event,
            pricingPoint=venue,
        )
        later_pricing_id = later_pricing.id

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": 50.1}
        )

        db.session.refresh(booking_to_edit)
        db.session.refresh(booking_untouched)
        db.session.refresh(stock_to_edit)
        db.session.refresh(stock_untouched)
        db.session.refresh(cancelled_event)
        db.session.refresh(later_event)

        assert response.status_code == 303
        assert stock_untouched.price == decimal.Decimal("123.45")
        assert booking_untouched.amount == decimal.Decimal("123.45")
        assert cancelled_event.status == finance_models.FinanceEventStatus.READY
        assert cancelled_pricing.status == finance_models.PricingStatus.CANCELLED

        assert stock_to_edit.price == decimal.Decimal("50.1")
        assert booking_to_edit.amount == decimal.Decimal("50.1")
        assert later_event.status == finance_models.FinanceEventStatus.READY
        assert db.session.query(finance_models.Pricing).filter_by(id=later_pricing_id).count() == 0

    def test_offer_stock_edit_with_french_decimal(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        booking_to_edit = bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": "50,1"}
        )

        db.session.refresh(booking_to_edit)
        db.session.refresh(stock_to_edit)

        assert response.status_code == 303

        assert stock_to_edit.price == decimal.Decimal("50.1")
        assert booking_to_edit.amount == decimal.Decimal("50.1")

    def test_offer_stock_edit_confirmed_booking(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        booking_to_edit = bookings_factories.BookingFactory(
            status=BookingStatus.CONFIRMED,
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": 50.1}
        )

        db.session.refresh(booking_to_edit)
        db.session.refresh(stock_to_edit)

        assert response.status_code == 303
        assert stock_to_edit.price == decimal.Decimal("50.1")
        assert booking_to_edit.amount == decimal.Decimal("50.1")

    def test_offer_stock_edit_confirmed_booking_percent(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        booking_to_edit = bookings_factories.BookingFactory(
            status=BookingStatus.CONFIRMED,
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"percent": "12.34"}
        )

        db.session.refresh(booking_to_edit)
        db.session.refresh(stock_to_edit)

        assert response.status_code == 303
        assert stock_to_edit.price == decimal.Decimal("108.22")
        assert booking_to_edit.amount == decimal.Decimal("108.22")

    def test_offer_stock_edit_cashfow_script_running(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )
        app.redis_client.set(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK, "1", 600)

        try:
            response = self.post_to_endpoint(
                authenticated_client,
                offer_id=offer.id,
                stock_id=event.booking.stock.id,
                form={"price": 50.1},
            )
        finally:
            app.redis_client.delete(finance_conf.REDIS_GENERATE_CASHFLOW_LOCK)

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert (
            html_parser.extract_alert(response.data)
            == "Le script de génération des cashflows est en cours, veuillez réessayer plus tard."
        )

        assert event.booking.stock.price == decimal.Decimal("123.45")

    def test_offer_stock_edit_no_used_or_confirmed_bookings(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        stock = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
        )
        bookings_factories.BookingFactory(
            status=BookingStatus.CANCELLED,
            stock=stock,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=stock.id,
            form={"price": 50.1},
        )

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert html_parser.extract_alert(response.data) == "Ce stock n'est pas éditable."

    def test_offer_stock_edit_offer_not_an_event(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.LIVRE_PAPIER.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={"price": 50.1},
        )

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert html_parser.extract_alert(response.data) == "Ce stock n'est pas éditable."

        assert event.booking.stock.price == decimal.Decimal("123.45")

    def test_offer_stock_edit_raising_price(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={"price": 200.0},
        )

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert (
            "Les données envoyées comportent des erreurs. Nouveau prix : Le prix doit être positif et inférieur à 123.45 €."
            in html_parser.extract_alert(response.data)
        )

        assert event.booking.stock.price == decimal.Decimal("123.45")

    def test_offer_stock_edit_raising_price_percent(self, authenticated_client, app):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={"percent": 200.0},
        )

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert (
            "Les données envoyées comportent des erreurs. Réduction en pourcentage (%) : La réduction du prix doit être entre 0 % et 100 %."
            in html_parser.extract_alert(response.data)
        )

        assert event.booking.stock.price == decimal.Decimal("123.45")

    def test_offer_stock_edit_both_price_percent(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("123.45"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={
                "percent": 10.0,
                "price": 100,
            },
        )

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200
        assert (
            "Les données envoyées comportent des erreurs. Nouveau prix : Un seul des deux champs est utilisable à la fois ; Réduction en pourcentage (%) : Un seul des deux champs est utilisable à la fois ;"
            in html_parser.extract_alert(response.data)
        )

        assert event.booking.stock.price == decimal.Decimal("123.45")

    def test_edit_stock_withprice_higher_than_booking(self, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.CONFERENCE.id)
        venue = offer.venue
        event = finance_factories.FinanceEventFactory(
            booking__amount=decimal.Decimal("5.00"),
            booking__stock__offer=offer,
            booking__stock__price=decimal.Decimal("123.45"),
            venue=venue,
            pricingPoint=venue,
        )
        finance_factories.PricingFactory(
            event=event,
            pricingPoint=venue,
        )

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer.id,
            stock_id=event.booking.stock.id,
            form={"price": 10.0},
        )

        db.session.refresh(event.booking)

        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id)
        assert response.location == expected_url

        response = authenticated_client.get(url_for("backoffice_web.offer.get_offer_details", offer_id=offer.id))
        assert response.status_code == 200

        assert event.booking.stock.price == decimal.Decimal("10.00")
        assert event.booking.amount == decimal.Decimal("5.00")

    def test_offer_stock_edit_with_price_category(self, authenticated_client):
        price_category = offers_factories.PriceCategoryFactory(
            offer__subcategoryId=subcategories.CONFERENCE.id,
            price=decimal.Decimal("123.45"),
        )
        label = price_category.priceCategoryLabel.label
        offer = price_category.offer
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
            priceCategory=price_category,
        )
        booking_to_edit = bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": "50.1"}
        )

        db.session.expire_all()

        assert response.status_code == 303

        assert stock_to_edit.price == decimal.Decimal("50.1")
        assert booking_to_edit.amount == decimal.Decimal("50.1")
        assert price_category.price == decimal.Decimal("50.1")
        assert price_category.priceCategoryLabel.label == label
        assert stock_to_edit.priceCategory == price_category

    def test_offer_stock_edit_with_price_category_with_multiple_stocks(self, authenticated_client):
        price_category = offers_factories.PriceCategoryFactory(
            offer__subcategoryId=subcategories.CONFERENCE.id,
            price=decimal.Decimal("123.45"),
        )
        label = price_category.priceCategoryLabel.label
        offer = price_category.offer
        stock_to_edit = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
            priceCategory=price_category,
        )
        other_stock = offers_factories.StockFactory(
            offer=offer,
            price=decimal.Decimal("123.45"),
            priceCategory=price_category,
        )
        booking_to_edit = bookings_factories.UsedBookingFactory(
            stock=stock_to_edit,
            amount=decimal.Decimal("123.45"),
        )

        response = self.post_to_endpoint(
            authenticated_client, offer_id=offer.id, stock_id=stock_to_edit.id, form={"price": "50.1"}
        )

        db.session.expire_all()

        assert response.status_code == 303
        expected_label = f"{label} - Revalorisation du {datetime.date.today().strftime('%d/%m/%Y')}"
        assert stock_to_edit.price == decimal.Decimal("50.1")
        assert booking_to_edit.amount == decimal.Decimal("50.1")
        assert stock_to_edit.priceCategory != price_category
        assert stock_to_edit.priceCategory.price == decimal.Decimal("50.1")
        assert stock_to_edit.priceCategory.priceCategoryLabel.label == expected_label

        assert other_stock.price == decimal.Decimal("123.45")
        assert other_stock.priceCategory == price_category
        assert other_stock.priceCategory.price == decimal.Decimal("123.45")
        assert other_stock.priceCategory.priceCategoryLabel.label == label


class DownloadBookingsCSVTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.download_bookings_csv"
    endpoint_kwargs = {"offer_id": 1, "stock_id": 1}
    needed_permission = perm_models.Permissions.READ_OFFERS

    # session + bookings
    expected_num_queries = 2

    def test_download_bookings_csv(self, legit_user, authenticated_client):
        offerer = offerers_factories.UserOffererFactory().offerer  # because of join on UserOfferers
        offer = offers_factories.ThingOfferFactory(venue__managingOfferer=offerer)
        bookings_factories.UsedBookingFactory(stock__offer=offer)
        bookings_factories.ReimbursedBookingFactory(stock__offer=offer)
        bookings_factories.BookingFactory(stock__offer=offer)
        bookings_factories.BookingFactory()  # other offer

        url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        assert len(response.data.split(b"\n")) == 1 + 3 + 1


class DownloadBookingsXLSXTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.download_bookings_xlsx"
    endpoint_kwargs = {"offer_id": 1, "stock_id": 1}
    needed_permission = perm_models.Permissions.READ_OFFERS

    # session + bookings
    expected_num_queries = 2

    def reader_from_response(self, response):
        wb = openpyxl.load_workbook(BytesIO(response.data))
        return wb.active

    def test_download_bookings_xlsx(self, authenticated_client):
        offerer = offerers_factories.UserOffererFactory().offerer  # because of join on UserOfferers
        offer = offers_factories.EventOfferFactory(venue__managingOfferer=offerer)
        booking1 = bookings_factories.UsedBookingFactory(stock__offer=offer)
        booking2 = bookings_factories.ReimbursedBookingFactory(stock__offer=offer)
        bookings_factories.BookingFactory()  # other offer

        url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        sheet = self.reader_from_response(response)
        assert sheet.cell(row=1, column=1).value == "Structure"
        assert sheet.cell(row=2, column=1).value == booking1.venue.name
        assert sheet.cell(row=3, column=1).value == booking2.venue.name
        assert sheet.cell(row=4, column=1).value == None


class ActivateOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.activate_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offer
    # update offers
    expected_num_queries = 3

    def test_activate_offer_with_stocks(self, legit_user, authenticated_client):
        offer_to_activate = offers_factories.OfferFactory(isActive=False)

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_activate.id,
            expected_num_queries=self.expected_num_queries,
        )
        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.list_offers")
        assert response.location == expected_url

        db.session.refresh(offer_to_activate)
        assert offer_to_activate.isActive is True

    def test_activate_offer_with_stocks_using_htmx(self, legit_user, authenticated_client):
        offer_to_activate = offers_factories.OfferFactory(isActive=False)

        expected_num_queries = self.expected_num_queries
        expected_num_queries += 3  # re-fetch to render updated offer + venue address + offer address

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_activate.id,
            expected_num_queries=expected_num_queries,
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        # ensure that the row is rendered
        cells = html_parser.extract_plain_row(response.data, id=f"offer-row-{offer_to_activate.id}")
        assert len(cells) == 25
        i = count()
        assert cells[next(i)] == ""  # Checkbox
        assert cells[next(i)] == (  # Actions
            "Voir le détail de l’offre Valider l'offre Rejeter l'offre Mettre l'offre en pause Taguer / Pondérer"
        )
        assert cells[next(i)] == ""  # Image
        assert cells[next(i)] == str(offer_to_activate.id)  # ID
        assert cells[next(i)] == offer_to_activate.name  # Nom de l'offre
        assert cells[next(i)] == ""  # EAN / Allociné ID
        assert cells[next(i)] == offer_to_activate.category.pro_label  # Catégorie
        assert cells[next(i)] == offer_to_activate.subcategory.pro_label  # Sous-catégorie
        assert cells[next(i)] == ""  # Règles de conformité
        assert cells[next(i)] == ""  # Score data
        assert cells[next(i)] == ""  # Predicition du validation_status (data)
        assert cells[next(i)] == "-"  # Tarif
        assert cells[next(i)] == ""  # Tag
        assert cells[next(i)] == ""  # Date(s) de l'évènement
        assert cells[next(i)] == ""  # Date(s) limite(s) de réservation
        assert cells[next(i)] == ""  # Créateur de l'offre
        assert cells[next(i)] == ""  # Pondération
        assert cells[next(i)] == "• Validée"  # État
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Date de création
        assert cells[next(i)] == ""  # Dérnière validation
        assert cells[next(i)] == offer_to_activate.offererAddress.address.departmentCode  # Département
        assert cells[next(i)] == offer_to_activate.venue.managingOfferer.name  # Entité juridique
        assert cells[next(i)] == offer_to_activate.venue.name  # Partenaire culturel
        assert cells[next(i)] == "Voir toutes les offres"  # Offres du partenaire culturel
        assert cells[next(i)] == ""  # Partenaire technique

        db.session.refresh(offer_to_activate)
        assert offer_to_activate.isActive is True


class GetActivateOfferFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_activate_offer_form"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offer
    expected_num_queries = 2

    def test_get_activate_form_test(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory()

        form_url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(form_url)
            assert response.status_code == 200


class DeactivateOfferTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.deactivate_offer"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offer
    # update offers
    expected_num_queries = 3

    def test_deactivate_offer_with_stocks(self, legit_user, authenticated_client):
        offer_to_deactivate = offers_factories.OfferFactory(isActive=True)

        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_deactivate.id,
            expected_num_queries=self.expected_num_queries,
        )
        assert response.status_code == 303

        expected_url = url_for("backoffice_web.offer.list_offers")
        assert response.location == expected_url

        db.session.refresh(offer_to_deactivate)
        assert offer_to_deactivate.isActive is False

    def test_deactivate_offer_with_stocks_using_htmx(self, legit_user, authenticated_client):
        offer_to_deactivate = offers_factories.OfferFactory(isActive=True)

        expected_num_queries = self.expected_num_queries
        # re-fetch offer to render again + venue address + offer address
        expected_num_queries += 3
        response = self.post_to_endpoint(
            authenticated_client,
            offer_id=offer_to_deactivate.id,
            expected_num_queries=expected_num_queries,
            headers={"hx-request": "true"},
        )
        assert response.status_code == 200
        # ensure that the row is rendered
        cells = html_parser.extract_plain_row(response.data, id=f"offer-row-{offer_to_deactivate.id}")
        assert len(cells) == 25
        i = count()
        assert cells[next(i)] == ""  # Checkbox
        assert cells[next(i)] == (  # Actions
            "Voir le détail de l’offre Valider l'offre Rejeter l'offre Publier l'offre Taguer / Pondérer"
        )
        assert cells[next(i)] == ""  # Image
        assert cells[next(i)] == str(offer_to_deactivate.id)  # ID
        assert cells[next(i)] == offer_to_deactivate.name  # Nom de l'offre
        assert cells[next(i)] == ""  # EAN / Allociné ID
        assert cells[next(i)] == offer_to_deactivate.category.pro_label  # Catégorie
        assert cells[next(i)] == offer_to_deactivate.subcategory.pro_label  # Sous-catégorie
        assert cells[next(i)] == ""  # Règles de conformité
        assert cells[next(i)] == ""  # Score data
        assert cells[next(i)] == ""  # Predicition du validation_status (data)
        assert cells[next(i)] == "-"  # Tarif
        assert cells[next(i)] == ""  # Tag
        assert cells[next(i)] == ""  # Date(s) de l'évènement
        assert cells[next(i)] == ""  # Date(s) limite(s) de réservation
        assert cells[next(i)] == ""  # Créateur de l'offre
        assert cells[next(i)] == ""  # Pondération
        assert cells[next(i)] == "• Validée"  # État
        assert cells[next(i)] == datetime.date.today().strftime("%d/%m/%Y")  # Date de création
        assert cells[next(i)] == ""  # Dérnière validation
        assert cells[next(i)] == offer_to_deactivate.offererAddress.address.departmentCode  # Département
        assert cells[next(i)] == offer_to_deactivate.venue.managingOfferer.name  # Entité juridique
        assert cells[next(i)] == offer_to_deactivate.venue.name  # Partenaire culturel
        assert cells[next(i)] == "Voir toutes les offres"  # Offres du partenaire culturel
        assert cells[next(i)] == ""  # Partenaire technique

        db.session.refresh(offer_to_deactivate)
        assert offer_to_deactivate.isActive is False


class GetDeactivateOfferFormTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_deactivate_offer_form"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offer
    expected_num_queries = 2

    def test_get_deactivate_form_test(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory()

        form_url = url_for(self.endpoint, offer_id=offer.id)

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(form_url)
            assert response.status_code == 200


class BatchOfferActivateTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.batch_activate_offers"
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offers
    # update offers
    # re-fetch to render updated offer rows (3 queries)
    expected_num_queries = 6

    def test_batch_activate_offers(self, legit_user, authenticated_client):
        offers = offers_factories.OfferFactory.create_batch(3, isActive=False)
        parameter_ids = ",".join(str(offer.id) for offer in offers)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            expected_num_queries=self.expected_num_queries,
        )

        assert response.status_code == 200
        # ensure rows are rendered
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[0].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[1].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[2].id}")
        for offer in offers:
            db.session.refresh(offer)
            assert offer.isActive is True


class BatchOfferDeactivateTest(PostEndpointHelper):
    endpoint = "backoffice_web.offer.batch_deactivate_offers"
    needed_permission = perm_models.Permissions.ADVANCED_PRO_SUPPORT
    # session + user
    # get offers
    # update offers
    # re-fetch to render updated offer rows (3 queries)
    expected_num_queries = 6

    def test_batch_deactivate_offers(self, legit_user, authenticated_client):
        offers = offers_factories.OfferFactory.create_batch(3)
        parameter_ids = ",".join(str(offer.id) for offer in offers)

        response = self.post_to_endpoint(
            authenticated_client,
            form={"object_ids": parameter_ids},
            expected_num_queries=self.expected_num_queries,
        )

        assert response.status_code == 200
        # ensure rows are rendered
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[0].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[1].id}")
        html_parser.get_tag(response.data, tag="tr", id=f"offer-row-{offers[2].id}")
        for offer in offers:
            db.session.refresh(offer)
            assert offer.isActive is False


class GetOfferDetailsTest(GetEndpointHelper):
    endpoint = "backoffice_web.offer.get_offer_details"
    endpoint_kwargs = {"offer_id": 1}
    needed_permission = perm_models.Permissions.READ_OFFERS

    # session + offer with joined data
    expected_num_queries = 2

    def test_get_detail_offer(self, authenticated_client):
        offer = offers_factories.OfferFactory(
            description="Une offre pour tester",
            withdrawalDetails="Demander à la caisse",
            bookingContact="contact@example.com",
            bookingEmail="offre@example.com",
            ean="1234567891234",
            extraData={"author": "Author", "editeur": "Editor", "gtl_id": "08010000"},
        )
        offers_factories.OfferMetaDataFactory(offer=offer, videoUrl="https://www.youtube.com/watch?v=ixSI7qD-Z1s")
        offers_factories.OfferComplianceFactory(
            offer=offer,
            compliance_score=55,
            compliance_reasons=["stock_price", "offer_subcategory_id", "offer_description"],
            validation_status_prediction=offers_models.ComplianceValidationStatusPrediction.APPROVED,
            validation_status_prediction_reason="Cette offre est conforme aux règles de conformité",
        )
        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)

        assert str(offer.id) == descriptions["Offer ID"]
        assert descriptions["Catégorie"] == "Films, vidéos"
        assert descriptions["Sous-catégorie"] == "Support physique (DVD, Blu-ray...)"
        assert descriptions["Statut"] == "Épuisée"
        assert descriptions["Score data"] == "55"
        assert descriptions["Prédiction de validité"] == "À valider Prédiction générée par IA • À vérifier"
        assert descriptions["Raison de score faible"] == "Prix Sous-catégorie Description de l'offre"
        assert descriptions["Entité juridique"].startswith("Le Petit Rintintin Management")
        assert descriptions["Partenaire culturel"].startswith("Le Petit Rintintin")
        assert "Utilisateur de la dernière validation" not in descriptions
        assert "Date de dernière validation" not in descriptions
        assert descriptions["Auteur"] == "Author"
        assert descriptions["EAN"] == "1234567891234"
        assert descriptions["Éditeur"] == "Editor"
        assert descriptions["Description"] == "Une offre pour tester"
        assert descriptions["Vidéo"] == "https://www.youtube.com/watch?v=ixSI7qD-Z1s"
        assert descriptions["Informations de retrait"] == "Demander à la caisse"
        assert descriptions["Email de contact"] == "contact@example.com"
        assert descriptions["Email auquel envoyer les notifications"] == "offre@example.com"

        buttons = html_parser.extract(response.data, "button")
        assert "Resync. Algolia" in buttons

        badges = html_parser.extract_badges(response.data)
        assert "• Validée" in badges

        assert html_parser.count_table_rows(response.data) == 0

    @pytest.mark.parametrize(
        "subcategory_id",
        (
            subcategories.SUPPORT_PHYSIQUE_FILM.id,
            subcategories.SUPPORT_PHYSIQUE_MUSIQUE_VINYLE.id,
            subcategories.SUPPORT_PHYSIQUE_MUSIQUE_CD.id,
        ),
    )
    def test_get_detail_offer_display_music_type_only_for_music_support(self, subcategory_id, authenticated_client):
        offer = offers_factories.OfferFactory(
            subcategoryId=subcategory_id,
            extraData={"gtl_id": "08010000"},
        )

        url = url_for(self.endpoint, offer_id=offer.id, _external=True)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        if subcategory_id in (
            subcategories.SUPPORT_PHYSIQUE_MUSIQUE_VINYLE.id,
            subcategories.SUPPORT_PHYSIQUE_MUSIQUE_CD.id,
        ):
            assert descriptions["Type de musique"] == "Alternatif"
        else:
            assert "Type de musique" not in descriptions

    def test_get_detail_offer_with_product(self, authenticated_client):
        product = offers_factories.ProductFactory(subcategoryId=subcategories.LIVRE_PAPIER.id, name="good book")
        offer = offers_factories.OfferFactory(product=product)
        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

    def test_get_detail_event_offer(self, authenticated_client):
        product = offers_factories.ProductFactory(
            name="good movie",
            subcategoryId=subcategories.SEANCE_CINE.id,
            durationMinutes=133,
            description="description",
            extraData={
                "cast": [
                    "first actor",
                    "second actor",
                    "third actor",
                ],
                "type": "FEATURE_FILM",
                "visa": "123456",
                "genres": [
                    "ADVENTURE",
                    "ANIMATION",
                    "DRAMA",
                ],
                "theater": {
                    "allocine_room_id": "W1234",
                    "allocine_movie_id": 654321,
                },
                "companies": [
                    {
                        "company": {
                            "name": "Company1 Name",
                        },
                        "activity": "InternationalDistributionExports",
                    },
                    {
                        "company": {
                            "name": "Company2 Name",
                        },
                        "activity": "Distribution",
                    },
                    {
                        "company": {
                            "name": "Company3 Name",
                        },
                        "activity": "Production",
                    },
                    {
                        "company": {"name": "Company4 Name"},
                        "activity": "Production",
                    },
                    {
                        "company": {"name": "Company5 Name"},
                        "activity": "PrAgency",
                    },
                ],
                "countries": [
                    "Never Land",
                ],
                "releaseDate": "2023-04-12",
                "stageDirector": "Georges Méliès",
                "diffusionVersion": "VO",
                "performer": "John Doe",
            },
        )
        offer = offers_factories.OfferFactory(
            product=product,
            offererAddress=None,
            idAtProvider="pouet provider",
            isActive=True,
            isDuo=False,
            audioDisabilityCompliant=True,
            mentalDisabilityCompliant=False,
            motorDisabilityCompliant=None,
            visualDisabilityCompliant=False,
        )

        url = url_for(self.endpoint, offer_id=offer.id)

        response = authenticated_client.get(url)
        assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Offer ID"] == str(offer.id)
        assert descriptions["Catégorie"] == "Cinéma"
        assert descriptions["Sous-catégorie"] == "Séance de cinéma"
        assert descriptions["Produit"] == "good movie"
        assert descriptions["Genres"] == "ADVENTURE, ANIMATION, DRAMA"
        assert descriptions["Statut"] == "Épuisée"
        assert descriptions["Entité juridique"].startswith("Le Petit Rintintin Management")
        assert descriptions["Partenaire culturel"].startswith("Le Petit Rintintin")
        assert "Adresse" not in descriptions
        assert "Utilisateur de la dernière validation" not in descriptions
        assert "Date de la dernière validation" not in descriptions

        assert descriptions["Identifiant chez le fournisseur"] == "pouet provider"
        assert descriptions["Langue"] == "VO"
        assert descriptions["Durée"] == "133 minutes"
        assert descriptions["Description"] == "description"
        assert descriptions["Interprète"] == "John Doe"

        accessibility_badges = html_parser.extract_accessibility_badges(response.data)
        assert accessibility_badges["Handicap auditif"] is True
        assert accessibility_badges["Handicap mental"] is False
        assert accessibility_badges["Handicap moteur"] is False
        assert accessibility_badges["Handicap visuel"] is False

        badges = html_parser.extract_badges(response.data)
        assert "• Validée" in badges

        buttons = html_parser.extract(response.data, "button")
        assert "Resync. Algolia" in buttons
        assert "Modifier le partenaire culturel" not in buttons

        assert html_parser.count_table_rows(response.data) == 0

    def test_get_detail_validated_offer(self, legit_user, authenticated_client):
        validation_date = date_utils.get_naive_utc_now()
        offer = offers_factories.OfferFactory(
            lastValidationDate=validation_date,
            validation=offers_models.OfferValidationStatus.APPROVED,
            lastValidationAuthor=legit_user,
        )

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Utilisateur de la dernière validation"] == legit_user.full_name
        assert descriptions["Date de la dernière validation"] == format_date(validation_date, "%d/%m/%Y à %Hh%M")

    def test_get_detail_offer_without_show_subtype(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory(
            withdrawalDetails="Demander à la caisse",
            extraData={"showType": 1510},
        )

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

    def test_get_detail_offer_display_modify_offer_button(self, client):
        offer = offers_factories.OfferFactory()
        manage_offers = (
            db.session.query(perm_models.Permission).filter_by(name=perm_models.Permissions.MANAGE_OFFERS.name).one()
        )
        read_offers = (
            db.session.query(perm_models.Permission).filter_by(name=perm_models.Permissions.READ_OFFERS.name).one()
        )
        role = perm_factories.RoleFactory(permissions=[read_offers, manage_offers])
        user = users_factories.AdminFactory()
        user.backoffice_profile.roles = [role]
        db.session.flush()

        authenticated_client = client.with_bo_session_auth(user)
        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        buttons = html_parser.extract(response.data, "button")
        assert "Taguer/Pondérer" in buttons
        assert "Valider l'offre" not in buttons
        assert "Rejeter l'offre" not in buttons

    def test_get_detail_offer_display_validation_buttons_fraud(self, client):
        offer = offers_factories.OfferFactory()
        pro_fraud_actions = (
            db.session.query(perm_models.Permission)
            .filter_by(name=perm_models.Permissions.PRO_FRAUD_ACTIONS.name)
            .one()
        )
        read_offers = (
            db.session.query(perm_models.Permission).filter_by(name=perm_models.Permissions.READ_OFFERS.name).one()
        )
        role = perm_factories.RoleFactory(permissions=[read_offers, pro_fraud_actions])
        user = users_factories.AdminFactory()
        user.backoffice_profile.roles = [role]
        db.session.flush()

        authenticated_client = client.with_bo_session_auth(user)
        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        buttons = html_parser.extract(response.data, "button")
        assert "Taguer/Pondérer" not in buttons
        assert "Valider" in buttons
        assert "Rejeter" in buttons

    def test_get_detail_rejected_offer(self, legit_user, authenticated_client):
        validation_date = date_utils.get_naive_utc_now()
        offer = offers_factories.OfferFactory(
            lastValidationDate=validation_date,
            validation=offers_models.OfferValidationStatus.REJECTED,
            lastValidationAuthor=legit_user,
        )

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Utilisateur de la dernière validation"] == legit_user.full_name
        assert descriptions["Date de la dernière validation"] == format_date(validation_date, "%d/%m/%Y à %Hh%M")

    def test_get_offer_details_overseas(self, legit_user, authenticated_client):
        offerer_address = offerers_factories.OffererAddressFactory(
            address=geography_factories.AddressFactory(
                postalCode="97200", departmentCode="972", timezone="America/Martinique"
            )
        )
        venue = offerers_factories.VenueFactory(managingOfferer=offerer_address.offerer)
        offer = offers_factories.OfferFactory(
            venue=venue, subcategoryId=subcategories.CONCERT.id, offererAddress=offerer_address
        )

        stock_id = offers_factories.EventStockFactory(
            offer=offer,
            beginningDatetime=datetime.datetime(2027, 1, 1, 2),
            bookingLimitDatetime=datetime.datetime(2026, 12, 25, 16),
            price=30,
        ).id

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 3  # check_can_move_event_offer (2 x count + pricing point)

        url = url_for(self.endpoint, offer_id=offer.id, _external=True)
        db.session.expunge_all()

        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 1
        assert stocks_rows[0]["ID"] == str(stock_id)
        assert stocks_rows[0]["Prix"] == "30,00 €"
        assert stocks_rows[0]["Date / Heure (Martinique)"] == "31/12/2026 à 22h00"
        assert stocks_rows[0]["Fin des réservations (Martinique)"] == "25/12/2026 à 12h00"

    def test_get_offer_details_with_one_expired_stock(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.SEANCE_CINE.id)

        expired_stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(hours=1), price=6.66
        )

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 1  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 1
        assert stocks_rows[0]["ID"] == str(expired_stock.id)
        assert stocks_rows[0]["Stock réservé"] == "0"
        assert stocks_rows[0]["Stock restant"] == "0"
        assert stocks_rows[0]["Prix"] == "6,66 €"
        assert stocks_rows[0]["Date / Heure (Paris)"] == format_date(
            expired_stock.beginningDatetime, "%d/%m/%Y à %Hh%M"
        )

    def test_get_offer_details_with_two_expired_stocks(self, legit_user, authenticated_client):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.SEANCE_CINE.id)

        expired_stock_1 = offers_factories.EventStockFactory(
            offer=offer,
            quantity=100,
            dnBookedQuantity=70,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(hours=2),
        )
        expired_stock_2 = offers_factories.EventStockFactory(
            offer=offer,
            quantity=None,
            dnBookedQuantity=25,
            beginningDatetime=date_utils.get_naive_utc_now() - datetime.timedelta(hours=1),
        )

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 1  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 2
        assert stocks_rows[1]["ID"] == str(expired_stock_1.id)
        assert stocks_rows[1]["Stock réservé"] == "70"
        assert stocks_rows[1]["Stock restant"] == "0"
        assert stocks_rows[1]["Prix"] == "10,10 €"
        assert stocks_rows[1]["Date / Heure (Paris)"] == format_date(
            expired_stock_1.beginningDatetime, "%d/%m/%Y à %Hh%M"
        )

        assert stocks_rows[0]["ID"] == str(expired_stock_2.id)
        assert stocks_rows[0]["Stock réservé"] == "25"
        assert stocks_rows[0]["Stock restant"] == "0"
        assert stocks_rows[0]["Prix"] == "10,10 €"
        assert stocks_rows[0]["Date / Heure (Paris)"] == format_date(
            expired_stock_2.beginningDatetime, "%d/%m/%Y à %Hh%M"
        )

    @pytest.mark.parametrize(
        "quantity,booked_quantity,expected_remaining,venue_factory,expected_price,expected_timezone",
        [
            (1000, 0, "1000", offerers_factories.VenueFactory, "10,10 €", "Paris"),
            (1000, 50, "950", offerers_factories.VenueFactory, "10,10 €", "Paris"),
            (1000, 1000, "0", offerers_factories.VenueFactory, "10,10 €", "Paris"),
            (None, 0, "Illimité", offerers_factories.VenueFactory, "10,10 €", "Paris"),
            (None, 50, "Illimité", offerers_factories.CaledonianVenueFactory, "10,10 € (1205 CFP)", "Noumea"),
        ],
    )
    def test_get_offer_details_with_one_bookable_stock(
        self,
        legit_user,
        authenticated_client,
        quantity,
        booked_quantity,
        expected_remaining,
        venue_factory,
        expected_price,
        expected_timezone,
    ):
        offer = offers_factories.OfferFactory(subcategoryId=subcategories.SEANCE_CINE.id, venue=venue_factory())
        stock = offers_factories.EventStockFactory(offer=offer, quantity=quantity, dnBookedQuantity=booked_quantity)

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 3  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 1
        assert stocks_rows[0]["ID"] == str(stock.id)
        assert stocks_rows[0]["Stock réservé"] == str(booked_quantity)
        assert stocks_rows[0]["Stock restant"] == expected_remaining
        assert stocks_rows[0]["Prix"] == expected_price
        assert stocks_rows[0][f"Date / Heure ({expected_timezone})"] == format_date(
            stock.beginningDatetime, "%d/%m/%Y à %Hh%M", address=offer.offererAddress.address
        )

    def test_get_offer_details_with_soft_deleted_stock(self, authenticated_client):
        stock = offers_factories.EventStockFactory(quantity=0, dnBookedQuantity=0, isSoftDeleted=True)

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 3  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=stock.offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 1
        assert stocks_rows[0]["ID"] == str(stock.id)
        assert stocks_rows[0]["Stock réservé"] == "0"
        assert stocks_rows[0]["Stock restant"] == "supprimé"
        assert stocks_rows[0]["Prix"] == "10,10 €"
        assert stocks_rows[0]["Date / Heure (Paris)"] == format_date(stock.beginningDatetime, "%d/%m/%Y à %Hh%M")

    @pytest.mark.parametrize(
        "venue_factory,expected_price_1,expected_price_2,expected_price_3,expected_price_4",
        [
            (offerers_factories.VenueFactory, "0,00 €", "13,00 €", "42,00 €", "66,60 €"),
            (
                offerers_factories.CaledonianVenueFactory,
                "0,00 € (0 CFP)",
                "13,00 € (1550 CFP)",
                "42,00 € (5010 CFP)",
                "66,60 € (7945 CFP)",
            ),
        ],
    )
    def test_get_offer_details_with_price_categories(
        self,
        authenticated_client,
        venue_factory,
        expected_price_1,
        expected_price_2,
        expected_price_3,
        expected_price_4,
    ):
        venue = venue_factory()
        offer = offers_factories.EventOfferFactory(venue=venue)
        price_gold = offers_factories.PriceCategoryFactory(
            offer=offer, priceCategoryLabel__label="OR", price=66.6, priceCategoryLabel__venue=venue
        )
        price_silver = offers_factories.PriceCategoryFactory(
            offer=offer, priceCategoryLabel__label="ARGENT", price=42, priceCategoryLabel__venue=venue
        )
        price_bronze = offers_factories.PriceCategoryFactory(
            offer=offer, priceCategoryLabel__label="BRONZE", price=13, priceCategoryLabel__venue=venue
        )
        price_free = offers_factories.PriceCategoryFactory(
            offer=offer, priceCategoryLabel__label="GRATUIT", price=0, priceCategoryLabel__venue=venue
        )

        offers_factories.EventStockFactory(offer=offer, priceCategory=price_gold)
        offers_factories.EventStockFactory(offer=offer, priceCategory=price_silver)
        offers_factories.EventStockFactory(offer=offer, priceCategory=price_bronze)
        offers_factories.EventStockFactory(offer=offer, priceCategory=price_free)

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 3  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert len(stocks_rows) == 4
        assert stocks_rows[0]["Tarif"] == "GRATUIT"
        assert stocks_rows[0]["Prix"] == expected_price_1
        assert stocks_rows[1]["Tarif"] == "BRONZE"
        assert stocks_rows[1]["Prix"] == expected_price_2
        assert stocks_rows[2]["Tarif"] == "ARGENT"
        assert stocks_rows[2]["Prix"] == expected_price_3
        assert stocks_rows[3]["Tarif"] == "OR"
        assert stocks_rows[3]["Prix"] == expected_price_4

    def test_get_offer_details_stocks_sorted_by_event_date_desc(self, authenticated_client):
        now = date_utils.get_naive_utc_now()
        offer = offers_factories.EventOfferFactory()
        stock1 = offers_factories.EventStockFactory(offer=offer, beginningDatetime=now + datetime.timedelta(days=5))
        stock2 = offers_factories.EventStockFactory(offer=offer, beginningDatetime=now + datetime.timedelta(days=9))
        stock3 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=now + datetime.timedelta(days=7), isSoftDeleted=True
        )

        query_count = self.expected_num_queries
        query_count += 1  # _get_editable_stock
        query_count += 3  # check_can_move_event_offer

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(query_count):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        stocks_rows = html_parser.extract_table_rows(response.data)
        assert [row["ID"] for row in stocks_rows] == [str(stock2.id), str(stock3.id), str(stock1.id)]

    def test_get_event_offer(self, legit_user, authenticated_client):
        venue = offerers_factories.VenueFactory()
        offerers_factories.VenueFactory.create_batch(2, managingOfferer=venue.managingOfferer, pricing_point=venue)
        offer = offers_factories.EventOfferFactory(venue=venue)

        url = url_for(self.endpoint, offer_id=offer.id)
        # Additional queries to check if "Modifier le partenaire culturel" should be displayed or not":
        # - _get_editable_stock
        # - count stocks with beginningDatetime in the past
        # - count reimbursed bookings
        # - fetch destination venue candidates
        with assert_num_queries(self.expected_num_queries + 4):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        buttons = html_parser.extract(response.data, "button")
        assert "Modifier le partenaire culturel" in buttons

    def test_get_offer_details(self, authenticated_client):
        address = geography_factories.AddressFactory(
            street="1v Place Jacques Rueff",
            postalCode="75007",
            city="Paris",
            latitude=48.85605,
            longitude=2.298,
            inseeCode="75107",
            banId="75107_4803_00001_v",
        )
        offerer_adress = offerers_factories.OffererAddressFactory(label="Champ de Mars", address=address)
        offer = offers_factories.OfferFactory(
            offererAddressId=offerer_adress.id, venue__managingOfferer=offerer_adress.offerer
        )

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Localisation"] == "Champ de Mars 1v Place Jacques Rueff 75007 Paris 48.85605, 2.29800"

    def test_get_offer_details_with_offerer_confidence_rule(self, authenticated_client):
        rule = offerers_factories.ManualReviewOffererConfidenceRuleFactory(offerer__name="Offerer")
        offer = offers_factories.OfferFactory(venue__managingOfferer=rule.offerer)

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Entité juridique"] == "Offerer Revue manuelle"

    def test_get_offer_details_with_venue_confidence_rule(self, authenticated_client):
        rule = offerers_factories.ManualReviewVenueConfidenceRuleFactory(venue__name="Venue")
        offer = offers_factories.OfferFactory(venue=rule.venue)

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Partenaire culturel"] == "Venue Revue manuelle"

    def test_collective_offer_with_top_acteur_offerer(self, authenticated_client):
        offer = offers_factories.OfferFactory(
            venue__managingOfferer__name="Offerer",
            venue__managingOfferer__tags=[
                offerers_factories.OffererTagFactory(name=offerers_constants.TOP_ACTEUR_TAG_NAME, label="Top Acteur")
            ],
        )

        url = url_for(self.endpoint, offer_id=offer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url)
            assert response.status_code == 200

        descriptions = html_parser.extract_descriptions(response.data)
        assert descriptions["Entité juridique"] == "Offerer Top Acteur"
