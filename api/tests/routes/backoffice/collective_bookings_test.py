import datetime
from io import BytesIO

from flask import url_for
import openpyxl
import pytest

from pcapi.core.categories import subcategories_v2 as subcategories
from pcapi.core.educational import factories as educational_factories
from pcapi.core.educational import models as educational_models
from pcapi.core.finance import factories as finance_factories
from pcapi.core.finance import models as finance_models
from pcapi.core.offerers import factories as offerers_factories
from pcapi.core.permissions import models as perm_models
from pcapi.core.testing import assert_num_queries
from pcapi.models import db

from .helpers import html_parser
from .helpers.get import GetEndpointHelper
from .helpers.post import PostEndpointHelper


pytestmark = [
    pytest.mark.usefixtures("db_session"),
    pytest.mark.backoffice,
]


@pytest.fixture(scope="function", name="collective_bookings")
def collective_bookings_fixture() -> tuple:
    institution1 = educational_factories.EducationalInstitutionFactory(name="Collège Pépin le Bref")
    institution2 = educational_factories.EducationalInstitutionFactory(name="Collège Bertrade de Laon")
    institution3 = educational_factories.EducationalInstitutionFactory(name="Lycée Charlemagne")
    venue = offerers_factories.VenueFactory()  # same venue and offerer for 2 bookings
    # 0
    pending = educational_factories.PendingCollectiveBookingFactory(
        educationalInstitution=institution2,
        collectiveStock__collectiveOffer__name="Offer n°1",
        collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.CONFERENCE_RENCONTRE],
        collectiveStock__beginningDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=6),
        dateCreated=datetime.datetime.utcnow() - datetime.timedelta(days=4),
        venue=venue,
    )
    offerers_factories.UserOffererFactory(offerer=pending.offerer)
    # 1
    confirmed = educational_factories.CollectiveBookingFactory(
        educationalInstitution=institution1,
        educationalRedactor=educational_factories.EducationalRedactorFactory(firstName="Pépin", lastName="d'Italie"),
        collectiveStock__price=1234,
        collectiveStock__collectiveOffer__name="Visite des locaux primitifs du pass Culture",
        collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.VISITE_GUIDEE],
        collectiveStock__bookingLimitDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=2),
        collectiveStock__beginningDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=3),
        dateCreated=datetime.datetime.utcnow() - datetime.timedelta(days=3),
    )
    # 2
    cancelled = educational_factories.CancelledCollectiveBookingFactory(
        educationalInstitution=institution2,
        educationalRedactor=educational_factories.EducationalRedactorFactory(firstName="Louis", lastName="Le Pieux"),
        collectiveStock__price=567.8,
        collectiveStock__collectiveOffer__name="Offer n°2",
        collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.CONCERT],
        collectiveStock__bookingLimitDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=3),
        collectiveStock__beginningDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=7),
        dateCreated=datetime.datetime.utcnow() - datetime.timedelta(days=2),
        venue=venue,
    )
    # 3
    used = educational_factories.UsedCollectiveBookingFactory(
        educationalInstitution=institution3,
        collectiveStock__collectiveOffer__name="Offer n°3",
        collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.PROJECTION_AUDIOVISUELLE],
        collectiveStock__beginningDatetime=datetime.datetime.utcnow() + datetime.timedelta(days=5),
        dateCreated=datetime.datetime.utcnow() - datetime.timedelta(days=1),
    )
    # 4
    reimbursed = educational_factories.ReimbursedCollectiveBookingFactory(
        educationalInstitution=institution3,
        collectiveStock__collectiveOffer__name="Offer n°4",
        collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.ATELIER_DE_PRATIQUE],
        dateCreated=datetime.datetime.utcnow() - datetime.timedelta(days=5),
    )

    return pending, confirmed, cancelled, used, reimbursed


class ListCollectiveBookingsTest(GetEndpointHelper):
    endpoint = "backoffice_web.collective_bookings.list_collective_bookings"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # Use assert_num_queries() instead of assert_no_duplicated_queries() which does not detect one extra query caused
    # by a field added in the jinja template.
    # - fetch session (1 query)
    # - fetch user (1 query)
    # - fetch collective bookings with extra data (1 query)
    # - check finance incident feature flag
    expected_num_queries = 4

    def test_list_bookings_without_filter(self, authenticated_client, collective_bookings):
        with assert_num_queries(self.expected_num_queries - 2):
            response = authenticated_client.get(url_for(self.endpoint))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    @pytest.mark.parametrize(
        "incident_status, display_alert",
        [
            (finance_models.IncidentStatus.VALIDATED, True),
            (finance_models.IncidentStatus.CREATED, False),
            (finance_models.IncidentStatus.CANCELLED, False),
        ],
    )
    def test_display_incident_alert(self, authenticated_client, incident_status, display_alert):
        collective_booking = educational_factories.ReimbursedCollectiveBookingFactory(
            incidents=[
                finance_factories.CollectiveBookingFinanceIncidentFactory(
                    incident=finance_factories.FinanceIncidentFactory(status=incident_status)
                )
            ]
        )

        search_query = collective_booking.id
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=search_query))
            assert response.status_code == 200

        if display_alert:
            assert "bi-exclamation-triangle-fill" in str(response.data)
        else:
            assert "bi-exclamation-triangle-fill" not in str(response.data)

    def test_list_bookings_by_id(self, authenticated_client, collective_bookings):
        searched_id = str(collective_bookings[1].id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=searched_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # Warning: test may return more than 1 row when an offer id or institution id is the same as booking id
        assert len(rows) >= 1
        for row_index, row in enumerate(rows):
            if row["ID résa"] == searched_id:
                break
        else:
            row_index, row = 0, {}  # for pylint
            assert False, f"Expected booking {searched_id} not found in results"
        assert row["Établissement"].startswith("Collège Pépin le Bref")
        assert row["Enseignant"] == "Pépin d'Italie"
        assert row["Nom de l'offre"] == "Visite des locaux primitifs du pass Culture"
        assert row["ID offre"].isdigit()
        assert row["Montant"] == "1 234,00 €"
        assert row["Statut"] == "Confirmée"
        assert row["Date de réservation"].startswith(
            (datetime.date.today() - datetime.timedelta(days=3)).strftime("%d/%m/%Y à ")
        )
        assert row["Date de l'évènement"].startswith(
            (datetime.date.today() + datetime.timedelta(days=3)).strftime("%d/%m/%Y à ")
        )
        assert row["Structure"] == collective_bookings[1].offerer.name
        assert row["Lieu"] == collective_bookings[1].venue.name

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[row_index]
        assert f"Formats : {subcategories.EacFormat.VISITE_GUIDEE.value}" in extra_data
        assert "Date de validation" not in extra_data
        assert (
            f"Date de confirmation de réservation : {(datetime.date.today() - datetime.timedelta(days=1)).strftime('%d/%m/%Y')} à "
            in extra_data
        )
        assert (
            f"Date limite de réservation : {(datetime.date.today() + datetime.timedelta(days=2)).strftime('%d/%m/%Y')} à "
            in extra_data
        )
        assert "Date d'annulation" not in extra_data

        assert html_parser.extract_pagination_info(response.data) == (1, 1, len(rows))

    @pytest.mark.parametrize(
        "query_args",
        [
            {},
            {
                "from_date": datetime.datetime(1970, 1, 1),
            },
            {
                "to_date": datetime.datetime(2037, 12, 31),
            },
            {
                "from_date": datetime.datetime(1970, 1, 1),
                "to_date": datetime.datetime(2037, 12, 31),
            },
        ],
    )
    def test_display_download_link(self, authenticated_client, collective_bookings, query_args):
        venue_id = [collective_bookings[0].venueId]
        kwargs = {**query_args, "venue_id": venue_id}
        response = authenticated_client.get(url_for(self.endpoint, **kwargs))
        assert b"pc-clipboard" in response.data

    def test_list_bookings_by_name(self, authenticated_client, collective_bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q="Visite des locaux primitifs du pass Culture"))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1
        assert rows[0]["Nom de l'offre"] == "Visite des locaux primitifs du pass Culture"
        assert html_parser.extract_pagination_info(response.data) == (1, 1, 1)

    def test_list_bookings_by_id_not_found(self, authenticated_client, collective_bookings):
        search_query = str(collective_bookings[-1].id * 1000)
        with assert_num_queries(self.expected_num_queries - 1):
            response = authenticated_client.get(url_for(self.endpoint, q=search_query))
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 0

    def test_list_bookings_by_offer_id(self, authenticated_client, collective_bookings):
        searched_id = str(collective_bookings[2].collectiveStock.collectiveOffer.id)
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, q=searched_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # Warning: test may return more than 1 row when a booking id or institution id is the same as offer id
        assert len(rows) >= 1
        for row_index, row in enumerate(rows):
            if row["ID offre"] == searched_id:
                break
        else:
            row_index, row = 0, {}  # for pylint
            assert False, f"Expected offer {searched_id} not found in results"
        assert row["ID résa"] == str(collective_bookings[2].id)
        assert row["Établissement"].startswith("Collège Bertrade de Laon")
        assert row["Enseignant"] == "Louis Le Pieux"
        assert row["Nom de l'offre"] == collective_bookings[2].collectiveStock.collectiveOffer.name
        assert row["Montant"] == "567,80 €"
        assert row["Statut"] == "Annulée"
        assert row["Date de réservation"].startswith(
            (datetime.date.today() - datetime.timedelta(days=2)).strftime("%d/%m/%Y à ")
        )
        assert row["Date de l'évènement"].startswith(
            (datetime.date.today() + datetime.timedelta(days=7)).strftime("%d/%m/%Y à ")
        )
        assert row["Structure"] == collective_bookings[2].offerer.name
        assert row["Lieu"] == collective_bookings[2].venue.name

        extra_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[row_index]
        assert f"Formats : {subcategories.EacFormat.CONCERT.value}" in extra_data
        assert "Date de validation" not in extra_data
        assert (
            f"Date de confirmation de réservation : {(datetime.date.today() - datetime.timedelta(days=1)).strftime('%d/%m/%Y')} à "
            in extra_data
        )
        assert (
            f"Date limite de réservation : {(datetime.date.today() + datetime.timedelta(days=3)).strftime('%d/%m/%Y')} à "
            in extra_data
        )
        assert f"Date d'annulation : {datetime.date.today().strftime('%d/%m/%Y')} à " in extra_data

        assert html_parser.extract_pagination_info(response.data) == (1, 1, len(rows))

    def test_list_bookings_by_cashflow_batch(self, authenticated_client):
        cashflows = finance_factories.CashflowFactory.create_batch(
            3,
            bankAccount=finance_factories.BankAccountFactory(),
        )

        finance_factories.CollectivePricingFactory(
            collectiveBooking=educational_factories.UsedCollectiveBookingFactory()
        )
        finance_factories.CollectivePricingFactory(
            collectiveBooking=educational_factories.ReimbursedCollectiveBookingFactory(), cashflows=[cashflows[1]]
        )

        reimbursed_pricing1 = finance_factories.CollectivePricingFactory(
            collectiveBooking=educational_factories.ReimbursedCollectiveBookingFactory(), cashflows=[cashflows[0]]
        )
        reimbursed_pricing3 = finance_factories.CollectivePricingFactory(
            collectiveBooking=educational_factories.ReimbursedCollectiveBookingFactory(), cashflows=[cashflows[2]]
        )

        searched_cashflow_batches = [cashflows[0].batch.id, cashflows[2].batch.id]
        # one more query because of cashflow_batches validation
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, cashflow_batches=searched_cashflow_batches))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {
            reimbursed_pricing1.collectiveBookingId,
            reimbursed_pricing3.collectiveBookingId,
        }

    @pytest.mark.parametrize(
        "status, expected_idx",
        [
            ([educational_models.CollectiveBookingStatus.PENDING.name], (0,)),
            ([educational_models.CollectiveBookingStatus.CONFIRMED.name], (1,)),
            ([educational_models.CollectiveBookingStatus.CANCELLED.name], (2,)),
            ([educational_models.CollectiveBookingStatus.USED.name], (3,)),
            ([educational_models.CollectiveBookingStatus.REIMBURSED.name], (4,)),
            (
                [
                    educational_models.CollectiveBookingStatus.CONFIRMED.name,
                    educational_models.CollectiveBookingStatus.USED.name,
                ],
                (1, 3),
            ),
        ],
    )
    def test_list_bookings_by_status(self, authenticated_client, collective_bookings, status, expected_idx):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, status=status))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["ID résa"] for row in rows) == {str(collective_bookings[idx].id) for idx in expected_idx}

    def test_list_bookings_by_date(self, authenticated_client, collective_bookings):
        date_from = datetime.date.today() - datetime.timedelta(days=3)
        date_to = datetime.date.today() - datetime.timedelta(days=2)
        separator = " - "

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    from_to_date=f"{date_from.strftime('%d/%m/%Y')}{separator}{date_to.strftime('%d/%m/%Y')}",
                )
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {collective_bookings[1].id, collective_bookings[2].id}

    @pytest.mark.parametrize(
        "from_date, to_date, expected_offers_name",
        [
            (
                (datetime.date.today() + datetime.timedelta(days=5)).isoformat(),
                None,
                {"Offer n°1", "Offer n°3", "Offer n°2"},
            ),
            (
                None,
                (datetime.date.today() + datetime.timedelta(days=6)).isoformat(),
                {"Offer n°1", "Offer n°3", "Offer n°4", "Visite des locaux primitifs du pass Culture"},
            ),
            (
                (datetime.date.today() + datetime.timedelta(days=5)).isoformat(),
                (datetime.date.today() + datetime.timedelta(days=6)).isoformat(),
                {"Offer n°3", "Offer n°1"},
            ),
        ],
    )
    def test_list_bookings_by_event_date(
        self, authenticated_client, collective_bookings, from_date, to_date, expected_offers_name
    ):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(
                    self.endpoint,
                    event_from_date=from_date,
                    event_to_date=to_date,
                )
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(row["Nom de l'offre"] for row in rows) == expected_offers_name

    def test_list_bookings_by_offerer(self, authenticated_client, collective_bookings):
        offerer_ids = [collective_bookings[1].offererId, collective_bookings[3].offererId]
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, offerer=offerer_ids))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {collective_bookings[1].id, collective_bookings[3].id}

    def test_list_bookings_by_venue(self, authenticated_client, collective_bookings):
        venue_id = collective_bookings[0].venueId
        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {collective_bookings[0].id, collective_bookings[2].id}

    def test_list_bookings_by_educational_institution(self, authenticated_client, collective_bookings):
        institution_id = collective_bookings[1].educationalInstitution.id

        with assert_num_queries(self.expected_num_queries + 1):
            response = authenticated_client.get(url_for(self.endpoint, institution=institution_id))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert set(int(row["ID résa"]) for row in rows) == {collective_bookings[1].id}

    def test_list_bookings_more_than_max(self, authenticated_client):
        educational_factories.CollectiveBookingFactory.create_batch(
            25, status=educational_models.CollectiveBookingStatus.CONFIRMED
        )

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(self.endpoint, status=educational_models.CollectiveBookingStatus.CONFIRMED.name, limit=20)
            )
            assert response.status_code == 200

        assert html_parser.count_table_rows(response.data) == 2 * 20  # extra data in second row for each booking
        assert "Il y a plus de 20 résultats dans la base de données" in html_parser.extract_alert(response.data)

    def test_additional_data_when_reimbursed(self, authenticated_client, collective_bookings):
        reimbursed = collective_bookings[4]
        pricing_venue = offerers_factories.VenueFactory()
        bank_account = finance_factories.BankAccountFactory()
        offerers_factories.VenueBankAccountLinkFactory(venue=pricing_venue, bankAccount=bank_account)
        pricing = finance_factories.CollectivePricingFactory(
            collectiveBooking=reimbursed,
            status=finance_models.PricingStatus.INVOICED,
            venue=pricing_venue,
        )
        cashflow = finance_factories.CashflowFactory(bankAccount=bank_account, pricings=[pricing])

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(self.endpoint, status=educational_models.CollectiveBookingStatus.REIMBURSED.name)
            )
            assert response.status_code == 200

        reimbursement_data = html_parser.extract(response.data, tag="tr", class_="collapse accordion-collapse")[0]
        assert "Total payé par l'utilisateur : 100,00 €" in reimbursement_data
        assert f"Date de remboursement : {reimbursed.reimbursementDate.strftime('%d/%m/%Y')} à " in reimbursement_data
        assert "Montant remboursé : 100,00 €" in reimbursement_data
        assert f"N° de virement : {cashflow.batch.label}" in reimbursement_data
        assert "Taux de remboursement : 100,0 %" in reimbursement_data

    def test_list_collective_bookings_by_formats(self, authenticated_client):
        target_format = subcategories.EacFormat.CONCERT
        target_booking = educational_factories.CollectiveBookingFactory(
            collectiveStock__collectiveOffer__formats=[target_format]
        )

        educational_factories.CollectiveBookingFactory(
            collectiveStock__collectiveOffer__formats=[subcategories.EacFormat.ATELIER_DE_PRATIQUE]
        )

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, formats=str(target_format.name)))
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        assert len(rows) == 1

        assert rows[0]["ID résa"] == str(target_booking.id)

    def test_sort_collective_bookings_by_event_date(self, authenticated_client, collective_bookings):
        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(
                url_for(self.endpoint, status=[s.name for s in educational_models.CollectiveBookingStatus])
            )
            assert response.status_code == 200

        rows = html_parser.extract_table_rows(response.data)
        # beginningDatetime at J+7, J+6, J+5, J+3, J+1
        assert [row["ID résa"] for row in rows] == [str(collective_bookings[idx].id) for idx in (2, 0, 3, 1, 4)]


class MarkCollectiveBookingAsUsedTest(PostEndpointHelper):
    endpoint = "backoffice_web.collective_bookings.mark_booking_as_used"
    endpoint_kwargs = {"collective_booking_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_uncancel_and_mark_as_used(self, authenticated_client, collective_bookings):
        cancelled = collective_bookings[2]

        response = self.post_to_endpoint(authenticated_client, collective_booking_id=cancelled.id)

        assert response.status_code == 303

        db.session.refresh(cancelled)
        assert cancelled.status is educational_models.CollectiveBookingStatus.CONFIRMED

        redirected_response = authenticated_client.get(response.headers["location"])
        assert f"La réservation {cancelled.id} a été validée" in html_parser.extract_alert(redirected_response.data)

    def test_uncancel_non_cancelled_booking(self, authenticated_client, collective_bookings):
        non_cancelled = collective_bookings[3]
        old_status = non_cancelled.status

        response = self.post_to_endpoint(authenticated_client, collective_booking_id=non_cancelled.id)

        assert response.status_code == 303

        db.session.refresh(non_cancelled)
        assert non_cancelled.status == old_status

        redirected_response = authenticated_client.get(response.headers["location"])
        assert "Impossible de valider une réservation qui n'est pas annulée" in html_parser.extract_alert(
            redirected_response.data
        )


class CancelCollectiveBookingTest(PostEndpointHelper):
    endpoint = "backoffice_web.collective_bookings.mark_booking_as_cancelled"
    endpoint_kwargs = {"collective_booking_id": 1}
    needed_permission = perm_models.Permissions.MANAGE_BOOKINGS

    def test_cancel_booking(self, authenticated_client, collective_bookings):
        confirmed = collective_bookings[1]

        response = self.post_to_endpoint(
            authenticated_client,
            collective_booking_id=confirmed.id,
            form={"reason": educational_models.CollectiveBookingCancellationReasons.BACKOFFICE.value},
        )

        assert response.status_code == 303

        db.session.refresh(confirmed)
        assert confirmed.status is educational_models.CollectiveBookingStatus.CANCELLED
        assert confirmed.cancellationReason == educational_models.CollectiveBookingCancellationReasons.BACKOFFICE

        redirected_response = authenticated_client.get(response.headers["location"])
        assert f"La réservation {confirmed.id} a été annulée" in html_parser.extract_alert(redirected_response.data)

    def test_cant_cancel_reimbursed_booking(self, authenticated_client, collective_bookings):
        reimbursed = collective_bookings[4]
        old_status = reimbursed.status
        booking_id = reimbursed.id

        response = self.post_to_endpoint(
            authenticated_client,
            collective_booking_id=booking_id,
            form={"reason": educational_models.CollectiveBookingCancellationReasons.BACKOFFICE.value},
        )

        assert response.status_code == 303

        booking = educational_models.CollectiveBooking.query.filter_by(id=booking_id).one()
        assert booking.status == old_status

        redirected_response = authenticated_client.get(response.headers["location"])
        assert (
            "Cette réservation est en train d’être remboursée, il est impossible de l’invalider"
            in html_parser.extract_alert(redirected_response.data)
        )

    def test_cant_cancel_cancelled_booking(self, authenticated_client, collective_bookings):
        cancelled = collective_bookings[2]
        old_status = cancelled.status
        booking_id = cancelled.id

        response = self.post_to_endpoint(
            authenticated_client,
            collective_booking_id=booking_id,
            form={"reason": educational_models.CollectiveBookingCancellationReasons.BACKOFFICE.value},
        )

        assert response.status_code == 303

        booking = educational_models.CollectiveBooking.query.filter_by(id=booking_id).one()
        assert booking.status == old_status

        redirected_response = authenticated_client.get(response.headers["location"])

        assert "Impossible d'annuler une réservation déjà annulée" in html_parser.extract_alert(
            redirected_response.data
        )

    def test_cant_cancel_booking_without_reason(self, authenticated_client, collective_bookings):
        confirmed = collective_bookings[1]

        response = self.post_to_endpoint(authenticated_client, collective_booking_id=confirmed.id)

        assert response.status_code == 303

        db.session.refresh(confirmed)
        assert confirmed.status == educational_models.CollectiveBookingStatus.CONFIRMED
        assert confirmed.cancellationReason is None

        redirected_response = authenticated_client.get(response.headers["location"])

        assert (
            "Les données envoyées comportent des erreurs. Raison : Information obligatoire"
            in html_parser.extract_alert(redirected_response.data)
        )


class GetCollectiveBookingCSVDownloadTest(GetEndpointHelper):
    endpoint = "backoffice_web.collective_bookings.get_collective_booking_csv_download"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # session + current user + list of bookings
    expected_num_queries = 3

    def test_csv_length(self, authenticated_client, collective_bookings):
        venue_id = collective_bookings[0].venueId

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_id))
            assert response.status_code == 200

        expected_length = 1  # headers
        expected_length += 1  # on booking
        expected_length += 1  # empty line

        assert len(response.data.split(b"\n")) == expected_length


class GetCollectiveBookingXLSXDownloadTest(GetEndpointHelper):
    endpoint = "backoffice_web.collective_bookings.get_collective_booking_xlsx_download"
    needed_permission = perm_models.Permissions.READ_BOOKINGS

    # session + current user + list of bookings
    expected_num_queries = 3

    def reader_from_response(self, response):
        wb = openpyxl.load_workbook(BytesIO(response.data))
        return wb.active

    def test_xlsx_length(self, authenticated_client, collective_bookings):
        venue_id = collective_bookings[0].venueId

        with assert_num_queries(self.expected_num_queries):
            response = authenticated_client.get(url_for(self.endpoint, venue=venue_id))
            assert response.status_code == 200

        sheet = self.reader_from_response(response)

        assert sheet.cell(row=1, column=1).value == "Lieu"
        assert sheet.cell(row=2, column=1).value == collective_bookings[0].venue.name
        assert sheet.cell(row=3, column=1).value == None
