import csv
from datetime import date
from datetime import datetime
from datetime import time
from datetime import timedelta
from io import BytesIO
from io import StringIO

from dateutil import tz
import openpyxl
import pytest
from pytest import fixture

import pcapi.core.bookings.factories as bookings_factories
from pcapi.core.bookings.models import Booking
from pcapi.core.bookings.models import BookingExportType
from pcapi.core.bookings.models import BookingStatus
from pcapi.core.bookings.models import BookingStatusFilter
import pcapi.core.bookings.repository as booking_repository
from pcapi.core.bookings.repository import get_bookings_from_deposit
from pcapi.core.categories import subcategories_v2 as subcategories
import pcapi.core.offerers.factories as offerers_factories
from pcapi.core.offerers.models import Venue
import pcapi.core.offers.factories as offers_factories
from pcapi.core.offers.models import Offer
from pcapi.core.testing import assert_no_duplicated_queries
from pcapi.core.testing import assert_num_queries
import pcapi.core.users.factories as users_factories
from pcapi.core.users.models import User
from pcapi.domain.booking_recap import utils as booking_recap_utils
from pcapi.utils.date import utc_to_local_datetime


pytestmark = pytest.mark.usefixtures("db_session")


NOW = datetime.utcnow()
YESTERDAY = NOW - timedelta(days=1)
TWO_DAYS_AGO = NOW - timedelta(days=2)
THREE_DAYS_AGO = NOW - timedelta(days=3)
FOUR_DAYS_AGO = NOW - timedelta(days=4)
FIVE_DAYS_AGO = NOW - timedelta(days=5)
ONE_WEEK_FROM_NOW = NOW + timedelta(weeks=1)

HEADERS = [
    "Lieu",
    "Nom de l’offre",
    "Date de l'évènement",
    "EAN",
    "Nom et prénom du bénéficiaire",
    "Email du bénéficiaire",
    "Téléphone du bénéficiaire",
    "Date et heure de réservation",
    "Date et heure de validation",
    "Contremarque",
    "Intitulé du tarif",
    "Prix de la réservation",
    "Statut de la contremarque",
    "Date et heure de remboursement",
    "Type d'offre",
    "Code postal du bénéficiaire",
    "Duo",
]


def test_find_all_ongoing_bookings(app):
    # Given
    user = users_factories.BeneficiaryGrant18Factory()
    stock = offers_factories.StockFactory(price=0, quantity=10)
    bookings_factories.CancelledBookingFactory(user=user, stock=stock)
    bookings_factories.UsedBookingFactory(user=user, stock=stock)
    ongoing_booking = bookings_factories.BookingFactory(user=user, stock=stock)

    # When
    all_ongoing_bookings = booking_repository.find_ongoing_bookings_by_stock(stock.id)

    # Then
    assert all_ongoing_bookings == [ongoing_booking]


def test_find_not_cancelled_bookings_by_stock(app):
    # Given
    user = users_factories.BeneficiaryGrant18Factory()
    stock = offers_factories.ThingStockFactory(price=0)
    bookings_factories.CancelledBookingFactory(user=user, stock=stock)
    used_booking = bookings_factories.UsedBookingFactory(user=user, stock=stock)
    not_cancelled_booking = bookings_factories.BookingFactory(user=user, stock=stock)

    # When
    all_not_cancelled_bookings = booking_repository.find_not_cancelled_bookings_by_stock(stock.id)

    # Then
    assert set(all_not_cancelled_bookings) == {used_booking, not_cancelled_booking}


default_booking_date = date.today()
one_year_before_booking = default_booking_date - timedelta(weeks=52)
one_year_after_booking = default_booking_date + timedelta(weeks=52)


class FindByProUserTest:
    def test_should_return_only_expected_booking_attributes(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
            amount=12,
        )

        # When
        bookings_query, total = booking_repository.find_by_pro_user(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )

        bookings = bookings_query.all()
        # Then
        assert total == 1
        assert len(bookings) == 1
        expected_booking = bookings[0]
        assert expected_booking.offerId == stock.offer.id
        assert expected_booking.offerName == offer.name
        assert expected_booking.beneficiaryFirstName == "Ron"
        assert expected_booking.beneficiaryLastName == "Weasley"
        assert expected_booking.beneficiaryEmail == "beneficiary@example.com"
        assert expected_booking.bookedAt == booking_date  # .astimezone(tz.gettz("Europe/Paris"))
        assert expected_booking.bookingToken == "ABCDEF"
        assert expected_booking.usedAt
        assert not expected_booking.cancelledAt
        assert not expected_booking.reimbursedAt
        assert expected_booking.quantity == 1
        assert expected_booking.bookingAmount == 12

    def test_should_return_only_validated_bookings_for_requested_period(self, app: fixture):
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)

        booking_date = datetime(2020, 1, 1, 10, 0, 0)

        bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, dateUsed=(booking_date + timedelta(days=1))
        )
        used_booking_2 = bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, dateUsed=(booking_date + timedelta(days=4))
        )
        bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, dateUsed=(booking_date + timedelta(days=8))
        )

        bookings_query, total = booking_repository.find_by_pro_user(
            user=pro,
            booking_period=((booking_date + timedelta(2)), (booking_date + timedelta(5))),
            status_filter=BookingStatusFilter.VALIDATED,
        )
        bookings = bookings_query.all()

        assert total == 1
        assert len(bookings) == 1
        assert bookings[0].bookingToken == used_booking_2.token

    def test_should_return_only_reimbursed_bookings_for_requested_period(self, app: fixture):
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        stock = offers_factories.ThingStockFactory(offer__venue=venue, price=10)

        booking_date = datetime(2020, 1, 1, 10, 0, 0)

        bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, reimbursementDate=(booking_date + timedelta(days=1))
        )
        reimbursed_booking_1 = bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, reimbursementDate=(booking_date + timedelta(days=2))
        )
        bookings_factories.UsedBookingFactory(
            stock=stock, quantity=1, dateCreated=booking_date, reimbursementDate=(booking_date + timedelta(days=4))
        )

        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro,
            booking_period=((booking_date + timedelta(2)), (booking_date + timedelta(3))),
            status_filter=BookingStatusFilter.REIMBURSED,
        )
        bookings = bookings_query.all()

        assert len(bookings) == 1
        assert bookings[0].bookingToken == reimbursed_booking_1.token

    def test_should_return_booking_as_duo_when_quantity_is_two(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock, quantity=2)

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 2
        assert bookings[0].quantity == 2

    def test_should_not_duplicate_bookings_when_user_is_admin_and_bookings_offerer_has_multiple_user(
        self, app: fixture
    ):
        # Given
        admin = users_factories.AdminFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(offerer=offerer)
        offerers_factories.UserOffererFactory(offerer=offerer)
        offerers_factories.UserOffererFactory(offerer=offerer)

        bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=offerer, quantity=2)

        # When
        bookings_recap_paginated_query, _ = booking_repository.find_by_pro_user(
            user=admin, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_recap_paginated_query.all()

        # Then
        assert len(bookings) == 2
        assert bookings[0].quantity == 2

    def test_should_return_event_booking_when_booking_is_on_an_event(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(
            offer=offer, price=0, beginningDatetime=datetime.utcnow() + timedelta(hours=98)
        )
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.BookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=yesterday,
            token="ABCDEF",
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        expected_booking = bookings[0]
        assert expected_booking.offerId == stock.offer.id
        assert expected_booking.offerName == stock.offer.name
        assert expected_booking.beneficiaryFirstName == "Ron"
        assert expected_booking.beneficiaryLastName == "Weasley"
        assert expected_booking.beneficiaryEmail == "beneficiary@example.com"
        assert expected_booking.bookedAt == yesterday
        assert expected_booking.bookingToken == "ABCDEF"
        assert not expected_booking.usedAt
        assert not expected_booking.cancelledAt
        assert not expected_booking.reimbursedAt
        assert not expected_booking.isConfirmed
        assert expected_booking.stockBeginningDatetime == stock.beginningDatetime

    def test_should_return_event_confirmed_booking_when_booking_is_on_an_event_in_confirmation_period(
        self, app: fixture
    ):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(
            offer=offer, price=0, beginningDatetime=datetime.utcnow() + timedelta(hours=98)
        )
        more_than_two_days_ago = datetime.utcnow() - timedelta(days=3)
        bookings_factories.BookingFactory(
            user=beneficiary, stock=stock, dateCreated=more_than_two_days_ago, token="ABCDEF"
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert bookings[0].isConfirmed

    def test_should_return_cancellation_date_when_booking_has_been_cancelled(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=5)
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.CancelledBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=yesterday,
            token="ABCDEF",
            amount=5,
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].cancelledAt

    def test_should_return_validation_date_when_booking_has_been_used_and_not_cancelled_not_reimbursed(
        self, app: fixture
    ):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        product = offers_factories.EventProductFactory()
        offer = offers_factories.EventOfferFactory(venue=venue, product=product)
        stock = offers_factories.EventStockFactory(offer=offer, price=5)
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=yesterday,
            token="ABCDEF",
            amount=5,
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        expected_booking = bookings[0]
        assert expected_booking.usedAt
        assert not expected_booking.cancelledAt
        assert not expected_booking.reimbursedAt

    def test_should_return_correct_number_of_matching_offerers_bookings_linked_to_user(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        today = datetime.utcnow()
        bookings_factories.BookingFactory(user=beneficiary, stock=stock, dateCreated=today, token="ABCD")

        offerer2 = offerers_factories.OffererFactory(siren="8765432")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer2)

        venue2 = offerers_factories.VenueFactory(managingOfferer=offerer, siret="8765432098765")
        offer2 = offers_factories.ThingOfferFactory(venue=venue2)
        stock2 = offers_factories.ThingStockFactory(offer=offer2, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock2, dateCreated=today, token="FGHI")

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 2

    def test_should_return_bookings_from_first_page(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(email="beneficiary@example.com")
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.EventOfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, price=0)
        today = datetime.utcnow()
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock, dateCreated=yesterday, token="ABCD")
        booking2 = bookings_factories.BookingFactory(user=beneficiary, stock=stock, dateCreated=today, token="FGHI")

        # When
        bookings_query, total = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking), page=1, per_page_limit=1
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].bookingToken == booking2.token
        assert total == 2

    def test_should_not_return_bookings_when_offerer_link_is_not_validated(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.NotValidatedUserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue, extraData=dict({"ean": "9876543234"}))
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock)

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )
        bookings = bookings_query.all()

        # Then
        assert bookings == []

    def test_should_return_one_booking_recap_item_when_quantity_booked_is_one(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.EventOfferFactory(venue=venue, isDuo=True)
        stock = offers_factories.EventStockFactory(offer=offer, price=0, beginningDatetime=datetime.utcnow())
        today = datetime.utcnow()
        booking = bookings_factories.BookingFactory(user=beneficiary, stock=stock, dateCreated=today, token="FGHI")

        # When
        bookings_query, total = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking), page=1, per_page_limit=4
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].bookingToken == booking.token
        assert total == 1

    def test_should_return_two_booking_recap_items_when_quantity_booked_is_two(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.EventOfferFactory(venue=venue, isDuo=True)
        stock = offers_factories.EventStockFactory(offer=offer, price=0, beginningDatetime=datetime.utcnow())
        today = datetime.utcnow()
        booking = bookings_factories.BookingFactory(
            user=beneficiary, stock=stock, dateCreated=today, token="FGHI", quantity=2
        )

        # When
        bookings_query, total = booking_repository.find_by_pro_user(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking), page=1, per_page_limit=4
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 2
        assert bookings[0].bookingToken == booking.token
        assert bookings[1].bookingToken == booking.token
        assert total == 2

    def test_should_return_booking_date_with_offerer_timezone_when_venue_is_digital(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].bookedAt == booking_date

    def test_should_return_booking_ean_when_information_is_available(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue, extraData=dict({"ean": "9876543234"}))
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )
        bookings = bookings_query.all()

        # Then
        assert bookings[0].offerEan == "9876543234"

    def test_should_return_only_booking_for_requested_venue(self, app: fixture):
        # Given
        pro_user = users_factories.ProFactory()
        user_offerer = offerers_factories.UserOffererFactory(user=pro_user)

        bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=user_offerer.offerer)
        booking_two = bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=user_offerer.offerer)

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=pro_user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            venue_id=booking_two.venue.id,
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        expected_booking = bookings[0]
        assert expected_booking.offerId == booking_two.stock.offer.id
        assert expected_booking.offerName == booking_two.stock.offer.name
        assert expected_booking.bookingAmount == booking_two.amount

    def test_should_return_only_booking_for_requested_event_date(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        event_date = datetime(2020, 12, 24, 10, 30)
        expected_booking = bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=event_date, offer__venue__managingOfferer=user_offerer.offerer
            )
        )
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(offer__venue__managingOfferer=user_offerer.offerer)
        )
        bookings_factories.BookingFactory(
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer)
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            event_date=event_date.date(),
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].bookingToken == expected_booking.token

    def should_consider_venue_locale_datetime_when_filtering_by_event_date(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        event_datetime = datetime(2020, 4, 21, 20, 00)

        offer_in_cayenne = offers_factories.OfferFactory(
            venue__postalCode="97300", venue__managingOfferer=user_offerer.offerer
        )
        cayenne_event_datetime = datetime(2020, 4, 22, 2, 0)
        stock_in_cayenne = offers_factories.EventStockFactory(
            offer=offer_in_cayenne, beginningDatetime=cayenne_event_datetime
        )
        cayenne_booking = bookings_factories.BookingFactory(stock=stock_in_cayenne)

        offer_in_mayotte = offers_factories.OfferFactory(
            venue__postalCode="97600", venue__managingOfferer=user_offerer.offerer
        )
        mayotte_event_datetime = datetime(2020, 4, 20, 22, 0)
        stock_in_mayotte = offers_factories.EventStockFactory(
            offer=offer_in_mayotte, beginningDatetime=mayotte_event_datetime
        )
        mayotte_booking = bookings_factories.BookingFactory(stock=stock_in_mayotte)

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            event_date=event_datetime.date(),
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 2
        bookings_tokens = [booking_recap.bookingToken for booking_recap in bookings]
        assert cayenne_booking.token in bookings_tokens
        assert mayotte_booking.token in bookings_tokens

    def test_should_return_only_bookings_for_requested_booking_period(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        booking_beginning_period = datetime(2020, 12, 24, 10, 30).date()
        booking_ending_period = datetime(2020, 12, 26, 15, 00).date()
        booking_status_filter = BookingStatusFilter.BOOKED
        expected_booking = bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 26, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )
        bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 29, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )
        bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 22, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(booking_beginning_period, booking_ending_period),
            status_filter=booking_status_filter,
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 1
        assert bookings[0].bookedAt == expected_booking.dateCreated

    def should_consider_venue_locale_datetime_when_filtering_by_booking_period(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        requested_booking_period_beginning = datetime(2020, 4, 21, 20, 00).date()
        requested_booking_period_ending = datetime(2020, 4, 22, 20, 00).date()

        offer_in_cayenne = offers_factories.OfferFactory(
            venue__postalCode="97300", venue__managingOfferer=user_offerer.offerer
        )
        cayenne_booking_datetime = datetime(2020, 4, 22, 2, 0)
        stock_in_cayenne = offers_factories.EventStockFactory(
            offer=offer_in_cayenne,
        )
        cayenne_booking = bookings_factories.BookingFactory(
            stock=stock_in_cayenne, dateCreated=cayenne_booking_datetime
        )

        offer_in_mayotte = offers_factories.OfferFactory(
            venue__postalCode="97600", venue__managingOfferer=user_offerer.offerer
        )
        mayotte_booking_datetime = datetime(2020, 4, 20, 23, 0)
        stock_in_mayotte = offers_factories.EventStockFactory(
            offer=offer_in_mayotte,
        )
        mayotte_booking = bookings_factories.BookingFactory(
            stock=stock_in_mayotte, dateCreated=mayotte_booking_datetime
        )

        # When
        bookings_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(requested_booking_period_beginning, requested_booking_period_ending),
        )
        bookings = bookings_query.all()

        # Then
        assert len(bookings) == 2
        bookings_tokens = [booking_recap.bookingToken for booking_recap in bookings]
        assert cayenne_booking.token in bookings_tokens
        assert mayotte_booking.token in bookings_tokens

    def test_should_return_only_bookings_for_requested_offer_type(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        bookings_factories.BookingFactory(
            dateCreated=default_booking_date,
            stock__offer__venue__managingOfferer=user_offerer.offerer,
        )

        # When
        individual_bookings_recap_paginated_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
        )
        individual_bookings_recap_paginated = individual_bookings_recap_paginated_query.all()
        all_bookings_recap_paginated_query, _ = booking_repository.find_by_pro_user(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
        )
        all_bookings_recap_paginated = all_bookings_recap_paginated_query.all()

        # Then
        assert len(individual_bookings_recap_paginated) == 1
        assert len(all_bookings_recap_paginated) == 1


class GetOfferBookingsByStatusCSVTest:
    def _validate_csv_row(
        self, data_dict: dict, beneficiary: User, offer: Offer, venue: Venue, booking: Booking, status: str, duo: str
    ):
        assert data_dict["Lieu"] == venue.name
        assert data_dict["Nom de l’offre"] == offer.name
        assert data_dict["Date de l'évènement"] == str(
            utc_to_local_datetime(booking.stock.beginningDatetime, venue.timezone)
        )
        assert data_dict["EAN"] == ((offer.extraData or {}).get("ean") or "")
        assert data_dict["Nom et prénom du bénéficiaire"] == " ".join((beneficiary.lastName, beneficiary.firstName))
        assert data_dict["Email du bénéficiaire"] == beneficiary.email
        assert data_dict["Téléphone du bénéficiaire"] == (beneficiary.phoneNumber or "")
        assert data_dict["Date et heure de réservation"] == str(
            booking.dateCreated.astimezone(tz.gettz("Europe/Paris"))
        )
        if booking.dateUsed:
            assert data_dict["Date et heure de validation"] == str(
                booking.dateUsed.astimezone(tz.gettz("Europe/Paris"))
            )
        else:
            assert data_dict["Date et heure de validation"] == ""
        token = booking_recap_utils.get_booking_token(
            booking.token,
            booking.status,
            booking.isExternal,
            booking.stock.beginningDatetime,
        )
        if token:
            assert data_dict["Contremarque"] == token
        else:
            assert data_dict["Contremarque"] == ""
        assert data_dict["Intitulé du tarif"] == booking.stock.priceCategory.label
        assert data_dict["Prix de la réservation"] == f"{booking.amount:.2f}"
        assert data_dict["Statut de la contremarque"] == status
        if booking.reimbursementDate:
            assert data_dict["Date et heure de remboursement"] == str(
                booking.reimbursementDate.astimezone(tz.gettz("Europe/Paris"))
            )
        else:
            assert data_dict["Date et heure de remboursement"] == ""
        assert data_dict["Type d'offre"] == "offre grand public"
        assert data_dict["Code postal du bénéficiaire"] == beneficiary.postalCode
        assert data_dict["Duo"] == duo

    def should_return_validated_bookings_for_offer(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=10)
        )

        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=40)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary_2)
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        queries = 0
        queries += 1  # Get bookings

        offer_id = offer.id
        with assert_num_queries(queries):
            bookings_csv = booking_repository.export_bookings_by_offer_id(
                offer_id=offer_id,
                event_date=date.today() + timedelta(days=10),
                export_type=BookingExportType.CSV,
                validated=True,
            )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 2
        self._validate_csv_row(
            dict(zip(headers, data[0])), beneficiary, offer, venue, validated_booking, "validé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[1])), beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non"
        )

    def should_return_validated_bookings_for_offer_with_old_cancelled_booking(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=10)
        )

        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)
        bookings_factories.CancelledBookingFactory(
            stock=stock, user=beneficiary_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=2)
        )

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=40)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary_2)
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        queries = 1  # Get bookings

        offer_id = offer.id
        with assert_num_queries(queries):
            bookings_csv = booking_repository.export_bookings_by_offer_id(
                offer_id=offer_id,
                event_date=date.today() + timedelta(days=10),
                export_type=BookingExportType.CSV,
                validated=True,
            )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 2
        self._validate_csv_row(
            dict(zip(headers, data[0])), beneficiary, offer, venue, validated_booking, "validé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[1])), beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non"
        )

    def should_return_validated_bookings_for_offer_with_duo(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5))
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary, quantity=2)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=40)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary_2)
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        bookings_csv = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=5),
            export_type=BookingExportType.CSV,
            validated=True,
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 3
        self._validate_csv_row(
            dict(zip(headers, data[0])), beneficiary, offer, venue, validated_booking, "validé", "DUO 1"
        )
        self._validate_csv_row(
            dict(zip(headers, data[1])), beneficiary, offer, venue, validated_booking, "validé", "DUO 2"
        )
        self._validate_csv_row(
            dict(zip(headers, data[2])), beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non"
        )

    def should_return_all_bookings_for_offer(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        beneficiary_3 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary3@example.com", firstName="Hermione", lastName="Granger", postalCode="97300"
        )
        beneficiary_4 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary4@example.com", firstName="severus", lastName="Snape", postalCode="93000"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=10)
        )
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        reimbursed_booking = bookings_factories.ReimbursedBookingFactory(user=beneficiary_3, stock=stock)
        new_booking = bookings_factories.BookingFactory(user=beneficiary_4, stock=stock)

        # When
        bookings_csv = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=10),
            export_type=BookingExportType.CSV,
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 4
        self._validate_csv_row(
            dict(zip(headers, data[0])), beneficiary, offer, venue, validated_booking, "validé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[1])), beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[2])), beneficiary_3, offer, venue, reimbursed_booking, "remboursé", "Non"
        )
        self._validate_csv_row(dict(zip(headers, data[3])), beneficiary_4, offer, venue, new_booking, "réservé", "Non")

    def should_return_all_bookings_for_offer_with_duo(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5))
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary, quantity=2)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        reimbursed_booking = bookings_factories.ReimbursedBookingFactory(user=beneficiary, stock=stock)
        new_booking = bookings_factories.BookingFactory(user=beneficiary_2, stock=stock, quantity=2)

        # When
        bookings_csv = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=5),
            export_type=BookingExportType.CSV,
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 6
        self._validate_csv_row(
            dict(zip(headers, data[0])), beneficiary, offer, venue, validated_booking, "validé", "DUO 1"
        )
        self._validate_csv_row(
            dict(zip(headers, data[1])), beneficiary, offer, venue, validated_booking, "validé", "DUO 2"
        )
        self._validate_csv_row(
            dict(zip(headers, data[2])), beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[3])), beneficiary, offer, venue, reimbursed_booking, "remboursé", "Non"
        )
        self._validate_csv_row(
            dict(zip(headers, data[4])), beneficiary_2, offer, venue, new_booking, "réservé", "DUO 1"
        )
        self._validate_csv_row(
            dict(zip(headers, data[5])), beneficiary_2, offer, venue, new_booking, "réservé", "DUO 2"
        )


class GetOfferBookingsByStatusExcelTest:
    def _validate_excel_row(
        self,
        bookings_excel,
        row: int,
        beneficiary: User,
        offer: Offer,
        venue: Venue,
        booking: Booking,
        status: str,
        duo: str,
    ):
        wb = openpyxl.load_workbook(BytesIO(bookings_excel))
        sheet = wb.active

        # Then
        # Headers
        for i in range(1, len(HEADERS)):
            assert sheet.cell(row=1, column=i).value == HEADERS[i - 1]
        # Lieu
        assert sheet.cell(row=row, column=1).value == venue.name
        # Nom de l’offre
        assert sheet.cell(row=row, column=2).value == offer.name
        # Date de l'évènement
        assert sheet.cell(row=row, column=3).value == str(
            utc_to_local_datetime(booking.stock.beginningDatetime, venue.timezone)
        )
        # EAN
        assert sheet.cell(row=row, column=4).value == ((offer.extraData or {}).get("ean") or None)
        # Nom et prénom du bénéficiaire
        assert sheet.cell(row=row, column=5).value == " ".join((beneficiary.lastName, beneficiary.firstName))
        # Email du bénéficiaire
        assert sheet.cell(row=row, column=6).value == beneficiary.email
        # Téléphone du bénéficiaire
        assert sheet.cell(row=row, column=7).value == (beneficiary.phoneNumber or None)
        # Date et heure de réservation
        assert sheet.cell(row=row, column=8).value == str(booking.dateCreated.astimezone(tz.gettz("Europe/Paris")))
        # Date et heure de validation
        if booking.dateUsed:
            assert sheet.cell(row=row, column=9).value == str(booking.dateUsed.astimezone(tz.gettz("Europe/Paris")))
        else:
            assert sheet.cell(row=row, column=9).value == "None"
        # Contremarque
        token = booking_recap_utils.get_booking_token(
            booking.token,
            booking.status,
            booking.isExternal,
            booking.stock.beginningDatetime,
        )
        if token:
            assert sheet.cell(row=row, column=10).value == token
        else:
            assert sheet.cell(row=row, column=10).value == None
        # Intitulé du tarif
        assert sheet.cell(row=row, column=11).value == booking.stock.priceCategory.label
        # Prix de la réservation
        assert sheet.cell(row=row, column=12).value == float(str(booking.amount).rstrip("0"))
        # Statut de la contremarque
        assert sheet.cell(row=row, column=13).value == status
        # Date et heure de remboursement
        if booking.reimbursementDate:
            assert sheet.cell(row=row, column=14).value == str(
                booking.reimbursementDate.astimezone(tz.gettz("Europe/Paris"))
            )
        else:
            assert sheet.cell(row=row, column=14).value == "None"
        # Type d'offre
        assert sheet.cell(row=row, column=15).value == "offre grand public"
        # Code postal du bénéficiaire
        assert sheet.cell(row=row, column=16).value == beneficiary.postalCode
        # Duo
        assert sheet.cell(row=row, column=17).value == duo

    def should_return_validated_bookings_for_offer(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=3))
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary)
        bookings_factories.BookingFactory(
            stock=stock_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        with assert_num_queries(2):
            bookings_excel = booking_repository.export_bookings_by_offer_id(
                offer_id=offer.id,
                event_date=date.today() + timedelta(days=3),
                export_type=BookingExportType.EXCEL,
                validated=True,
            )

        # Then
        self._validate_excel_row(bookings_excel, 2, beneficiary, offer, venue, validated_booking, "validé", "Non")
        self._validate_excel_row(bookings_excel, 3, beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non")

    def should_return_validated_bookings_for_offer_with_duo(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=3))
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary, quantity=2)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary)
        bookings_factories.BookingFactory(
            stock=stock_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        bookings_excel = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=3),
            export_type=BookingExportType.EXCEL,
            validated=True,
        )

        # Then
        self._validate_excel_row(bookings_excel, 2, beneficiary, offer, venue, validated_booking, "validé", "DUO 1")
        self._validate_excel_row(bookings_excel, 3, beneficiary, offer, venue, validated_booking, "validé", "DUO 2")
        self._validate_excel_row(bookings_excel, 4, beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non")

    def should_return_all_bookings_for_offer(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        beneficiary_3 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary3@example.com", firstName="Hermione", lastName="Granger", postalCode="97300"
        )
        beneficiary_4 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary4@example.com", firstName="severus", lastName="Snape", postalCode="93000"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=30)
        )
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        reimbursed_booking = bookings_factories.ReimbursedBookingFactory(user=beneficiary_3, stock=stock)
        new_booking = bookings_factories.BookingFactory(user=beneficiary_4, stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary)
        bookings_factories.BookingFactory(
            stock=stock_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        bookings_excel = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=30),
            export_type=BookingExportType.EXCEL,
        )

        # Then
        self._validate_excel_row(bookings_excel, 2, beneficiary, offer, venue, validated_booking, "validé", "Non")
        self._validate_excel_row(bookings_excel, 3, beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non")
        self._validate_excel_row(bookings_excel, 4, beneficiary_3, offer, venue, reimbursed_booking, "remboursé", "Non")
        self._validate_excel_row(bookings_excel, 5, beneficiary_4, offer, venue, new_booking, "réservé", "Non")

    def should_return_validated_bookings_for_offer_with_duo_offerer_with_mutiple_users(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        pro_2 = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        offerers_factories.UserOffererFactory(user=pro_2, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=3))
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary, quantity=2)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary)
        bookings_factories.BookingFactory(
            stock=stock_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        bookings_excel = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=3),
            export_type=BookingExportType.EXCEL,
            validated=True,
        )

        # Then
        self._validate_excel_row(bookings_excel, 2, beneficiary, offer, venue, validated_booking, "validé", "DUO 1")
        self._validate_excel_row(bookings_excel, 3, beneficiary, offer, venue, validated_booking, "validé", "DUO 2")
        self._validate_excel_row(bookings_excel, 4, beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non")

    def should_return_all_bookings_for_offer_with_duo(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        beneficiary_2 = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary2@example.com", firstName="Harry", lastName="Potter", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.OfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=30)
        )
        validated_booking = bookings_factories.UsedBookingFactory(stock=stock, user=beneficiary, quantity=2)
        validated_booking_2 = bookings_factories.BookingFactory(
            stock=stock, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        reimbursed_booking = bookings_factories.ReimbursedBookingFactory(user=beneficiary, stock=stock)
        new_booking = bookings_factories.BookingFactory(user=beneficiary_2, stock=stock, quantity=2)

        stock_2 = offers_factories.EventStockFactory(
            offer=offer, beginningDatetime=datetime.utcnow() + timedelta(days=5)
        )
        bookings_factories.UsedBookingFactory(stock=stock_2, user=beneficiary)
        bookings_factories.BookingFactory(
            stock=stock_2, cancellation_limit_date=datetime.utcnow() - timedelta(days=1), user=beneficiary_2
        )
        bookings_factories.BookingFactory(stock=stock_2)

        # When
        bookings_excel = booking_repository.export_bookings_by_offer_id(
            offer_id=offer.id,
            event_date=date.today() + timedelta(days=30),
            export_type=BookingExportType.EXCEL,
        )

        # Then
        self._validate_excel_row(bookings_excel, 2, beneficiary, offer, venue, validated_booking, "validé", "DUO 1")
        self._validate_excel_row(bookings_excel, 3, beneficiary, offer, venue, validated_booking, "validé", "DUO 2")
        self._validate_excel_row(bookings_excel, 4, beneficiary_2, offer, venue, validated_booking_2, "confirmé", "Non")
        self._validate_excel_row(bookings_excel, 5, beneficiary, offer, venue, reimbursed_booking, "remboursé", "Non")
        self._validate_excel_row(bookings_excel, 6, beneficiary_2, offer, venue, new_booking, "réservé", "DUO 1")
        self._validate_excel_row(bookings_excel, 7, beneficiary_2, offer, venue, new_booking, "réservé", "DUO 2")


class GetCsvReportTest:
    def test_should_return_only_expected_booking_attributes(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley", postalCode="97300"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        booking = bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
            amount=12,
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Lieu"] == venue.name
        assert data_dict["Nom de l’offre"] == offer.name
        assert data_dict["Date de l'évènement"] == ""
        assert data_dict["EAN"] == ((offer.extraData or {}).get("ean") or "")
        assert data_dict["Nom et prénom du bénéficiaire"] == " ".join((beneficiary.lastName, beneficiary.firstName))
        assert data_dict["Email du bénéficiaire"] == beneficiary.email
        assert data_dict["Téléphone du bénéficiaire"] == (beneficiary.phoneNumber or "")
        assert data_dict["Date et heure de réservation"] == str(
            booking.dateCreated.astimezone(tz.gettz("Europe/Paris"))
        )
        assert data_dict["Date et heure de validation"] == str(booking.dateUsed.astimezone(tz.gettz("Europe/Paris")))
        assert data_dict["Contremarque"] == booking.token
        assert data_dict["Intitulé du tarif"] == ""
        assert data_dict["Prix de la réservation"] == f"{booking.amount:.2f}"
        assert data_dict["Statut de la contremarque"] == booking_repository.BOOKING_STATUS_LABELS[booking.status]
        assert data_dict["Date et heure de remboursement"] == ""
        assert data_dict["Type d'offre"] == "offre grand public"
        assert data_dict["Code postal du bénéficiaire"] == beneficiary.postalCode
        assert data_dict["Duo"] == "Non"

    def test_should_not_return_token_for_non_used_goods(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        booking = bookings_factories.BookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
            amount=12,
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Lieu"] == venue.name
        assert data_dict["Nom de l’offre"] == offer.name
        assert data_dict["Date de l'évènement"] == ""
        assert data_dict["EAN"] == ((offer.extraData or {}).get("ean") or "")
        assert data_dict["Nom et prénom du bénéficiaire"] == " ".join((beneficiary.lastName, beneficiary.firstName))
        assert data_dict["Email du bénéficiaire"] == beneficiary.email
        assert data_dict["Téléphone du bénéficiaire"] == (beneficiary.phoneNumber or "")
        assert data_dict["Date et heure de réservation"] == str(
            booking.dateCreated.astimezone(tz.gettz("Europe/Paris"))
        )
        assert data_dict["Date et heure de validation"] == ""
        assert data_dict["Contremarque"] == ""
        assert data_dict["Intitulé du tarif"] == ""
        assert data_dict["Prix de la réservation"] == f"{booking.amount:.2f}"
        assert data_dict["Statut de la contremarque"] == booking_repository.BOOKING_STATUS_LABELS[booking.status]
        assert data_dict["Date et heure de remboursement"] == ""
        assert data_dict["Code postal du bénéficiaire"] == ""
        assert data_dict["Duo"] == "Non"

    def test_should_return_only_validated_bookings_for_requested_period(self, app: fixture):
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        booking_date = datetime(2020, 1, 1, 10, 0, 0)

        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            dateUsed=(booking_date + timedelta(days=1)),
            stock__offer__name="Harry Potter Vol 1",
        )
        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            dateUsed=(booking_date + timedelta(days=4)),
            stock__offer__name="Harry Potter Vol 2",
        )
        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            dateUsed=(booking_date + timedelta(days=8)),
            stock__offer__name="Harry Potter Vol 3",
        )
        bookings_csv = booking_repository.get_export(
            user=pro,
            booking_period=((booking_date + timedelta(2)), (booking_date + timedelta(5))),
            status_filter=BookingStatusFilter.VALIDATED,
        )

        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        # Check bookings offer name
        assert data[0][1] == "Harry Potter Vol 2"

    def test_should_return_only_reimbursed_bookings_for_requested_period(self, app: fixture):
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        booking_date = datetime(2020, 1, 1, 10, 0, 0)

        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            reimbursementDate=(booking_date + timedelta(days=1)),
            stock__offer__name="Harry Potter Vol 1",
        )
        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            reimbursementDate=(booking_date + timedelta(days=2)),
            stock__offer__name="Harry Potter Vol 2",
        )
        bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
            quantity=1,
            dateCreated=booking_date,
            reimbursementDate=(booking_date + timedelta(days=4)),
            stock__offer__name="Harry Potter Vol 3",
        )

        bookings_csv = booking_repository.get_export(
            user=pro,
            booking_period=((booking_date + timedelta(2)), (booking_date + timedelta(3))),
            status_filter=BookingStatusFilter.REIMBURSED,
        )

        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        # Check bookings offer name
        assert data[0][1] == "Harry Potter Vol 2"

    def test_should_return_booking_as_duo_when_quantity_is_two(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock, quantity=2)

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2
        assert data[0][16] == "DUO 1"
        assert data[1][16] == "DUO 2"

    def test_should_not_duplicate_bookings_when_user_is_admin_and_bookings_offerer_has_multiple_user(
        self, app: fixture
    ):
        # Given
        admin = users_factories.AdminFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(offerer=offerer)
        offerers_factories.UserOffererFactory(offerer=offerer)
        offerers_factories.UserOffererFactory(offerer=offerer)

        bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=offerer, quantity=2)

        # When
        bookings_csv = booking_repository.get_export(
            user=admin, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2

    def test_should_return_event_booking_when_booking_is_on_an_event(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        offer = offers_factories.EventOfferFactory(venue=venue)
        stock = offers_factories.EventStockFactory(
            offer=offer, price=0, beginningDatetime=datetime.utcnow() + timedelta(hours=98)
        )
        yesterday = datetime.utcnow() - timedelta(days=1)
        booking = bookings_factories.BookingFactory(
            user=beneficiary, stock=stock, dateCreated=yesterday, token="ABCDEF"
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert headers == HEADERS
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Lieu"] == venue.name
        assert data_dict["Nom de l’offre"] == offer.name
        assert data_dict["Date de l'évènement"] == str(stock.beginningDatetime.astimezone(tz.gettz("Europe/Paris")))
        assert data_dict["EAN"] == ((offer.extraData or {}).get("ean") or "")
        assert data_dict["Nom et prénom du bénéficiaire"] == " ".join((beneficiary.lastName, beneficiary.firstName))
        assert data_dict["Email du bénéficiaire"] == beneficiary.email
        assert data_dict["Téléphone du bénéficiaire"] == (beneficiary.phoneNumber or "")
        assert data_dict["Date et heure de réservation"] == str(
            booking.dateCreated.astimezone(tz.gettz("Europe/Paris"))
        )
        assert data_dict["Date et heure de validation"] == ""
        assert data_dict["Contremarque"] == booking.token
        assert data_dict["Intitulé du tarif"] == "Tarif unique"
        assert data_dict["Prix de la réservation"] == f"{booking.amount:.2f}"
        assert data_dict["Statut de la contremarque"] == booking_repository.BOOKING_STATUS_LABELS[booking.status]
        assert data_dict["Date et heure de remboursement"] == ""
        assert data_dict["Code postal du bénéficiaire"] == ""
        assert data_dict["Duo"] == "Non"

    def test_should_return_event_confirmed_booking_when_booking_is_on_an_event_in_confirmation_period(
        self, app: fixture
    ):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(
            offer=offer, price=0, beginningDatetime=datetime.utcnow() + timedelta(hours=98)
        )
        more_than_two_days_ago = datetime.utcnow() - timedelta(days=3)
        bookings_factories.BookingFactory(
            user=beneficiary, stock=stock, dateCreated=more_than_two_days_ago, token="ABCDEF"
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Statut de la contremarque"] == "confirmé"

    def test_should_return_cancelled_status_when_booking_has_been_cancelled(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=5)
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.CancelledBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=yesterday,
            token="ABCDEF",
            amount=5,
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert (
            data_dict["Statut de la contremarque"] == booking_repository.BOOKING_STATUS_LABELS[BookingStatus.CANCELLED]
        )

    def test_should_return_validation_date_when_booking_has_been_used_and_not_cancelled_not_reimbursed(
        self, app: fixture
    ):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)
        product = offers_factories.EventProductFactory()
        offer = offers_factories.EventOfferFactory(venue=venue, product=product)
        stock = offers_factories.EventStockFactory(offer=offer, price=5)
        yesterday = datetime.utcnow() - timedelta(days=1)
        booking = bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=yesterday,
            token="ABCDEF",
            amount=5,
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Statut de la contremarque"] == booking_repository.BOOKING_STATUS_LABELS[BookingStatus.USED]
        assert data_dict["Date et heure de validation"] == str(booking.dateUsed.astimezone(tz.gettz("Europe/Paris")))

    def test_should_return_correct_number_of_matching_offerers_bookings_linked_to_user(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        today = datetime.utcnow()
        bookings_factories.BookingFactory(user=beneficiary, stock=stock, dateCreated=today, token="ABCD")

        offerer2 = offerers_factories.OffererFactory(siren="8765432")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer2)

        venue2 = offerers_factories.VenueFactory(managingOfferer=offerer, siret="8765432098765")
        offer2 = offers_factories.ThingOfferFactory(venue=venue2)
        stock2 = offers_factories.ThingStockFactory(offer=offer2, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock2, dateCreated=today, token="FGHI")

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2

    def test_should_not_return_bookings_when_offerer_link_is_not_validated(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.NotValidatedUserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue, extraData=dict({"ean": "9876543234"}))
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        bookings_factories.BookingFactory(user=beneficiary, stock=stock)

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        _, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 0

    def test_should_return_booking_date_with_offerer_timezone_when_venue_is_digital(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Date et heure de réservation"] == str(booking_date.astimezone(tz.gettz("America/Cayenne")))

    def test_should_return_booking_ean_when_information_is_available(self, app: fixture):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory(postalCode="97300")
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer, isVirtual=True, siret=None)

        offer = offers_factories.ThingOfferFactory(venue=venue, extraData=dict({"ean": "9876543234"}))
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365))
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["EAN"] == "9876543234"

    def test_should_return_booking_with_venue_name_when_public_name_is_not_provided(self, app):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue_for_event = offerers_factories.VenueFactory(
            managingOfferer=offerer, name="Lieu pour un évènement", siret="11816909600069"
        )
        product = offers_factories.EventProductFactory(name="Shutter Island")
        offer_for_event = offers_factories.EventOfferFactory(venue=venue_for_event, product=product)
        stock_for_event = offers_factories.EventStockFactory(offer=offer_for_event, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_event,
            dateCreated=(default_booking_date + timedelta(days=1)),
            token="BBBBBB",
        )

        venue_for_book = offerers_factories.VenueFactory(
            managingOfferer=offerer, name="Lieu pour un livre", siret="41816609600069"
        )
        product_book = offers_factories.ThingProductFactory(name="Harry Potter")
        offer_for_book = offers_factories.ThingOfferFactory(
            venue=venue_for_book, product=product_book, extraData=dict({"ean": "9876543234"})
        )
        stock_for_book = offers_factories.ThingStockFactory(offer=offer_for_book, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_book,
            dateCreated=default_booking_date,
            token="AAAAAA",
        )

        venue_for_thing = offerers_factories.VenueFactory(
            managingOfferer=offerer, name="Lieu pour un bien qui n'est pas un livre", siret="83994784300018"
        )
        product_thing = offers_factories.ThingProductFactory(name="Harry Potter")
        offer_for_thing = offers_factories.ThingOfferFactory(venue=venue_for_thing, product=product_thing)
        stock_for_thing = offers_factories.ThingStockFactory(offer=offer_for_thing, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_thing,
            dateCreated=(default_booking_date - timedelta(days=1)),
            token="ABCDEF",
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 3
        data_dicts = [dict(zip(headers, line)) for line in data]
        venues = [data_dict["Lieu"] for data_dict in data_dicts]
        assert sorted(venues) == [venue_for_thing.name, venue_for_book.name, venue_for_event.name]

    def test_should_return_booking_with_venue_public_name_when_public_name_is_provided(self, app):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory()
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue_for_event = offerers_factories.VenueFactory(
            managingOfferer=offerer, name="Opéra paris", publicName="Super Opéra de Paris", siret="11816909600069"
        )
        product = offers_factories.EventProductFactory(name="Shutter Island")
        offer_for_event = offers_factories.EventOfferFactory(venue=venue_for_event, product=product)
        stock_for_event = offers_factories.EventStockFactory(offer=offer_for_event, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_event,
            dateCreated=(default_booking_date + timedelta(days=1)),
            token="BBBBBB",
        )

        venue_for_book = offerers_factories.VenueFactory(
            managingOfferer=offerer, name="Lieu pour un livre", publicName="Librairie Châtelet", siret="41816609600069"
        )
        product_book = offers_factories.ThingProductFactory(name="Harry Potter")
        offer_for_book = offers_factories.ThingOfferFactory(
            venue=venue_for_book, product=product_book, extraData=dict({"ean": "9876543234"})
        )
        stock_for_book = offers_factories.ThingStockFactory(offer=offer_for_book, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_book,
            dateCreated=default_booking_date,
            token="AAAAAA",
        )

        venue_for_thing = offerers_factories.VenueFactory(
            managingOfferer=offerer,
            name="Lieu pour un bien qui n'est pas un livre",
            publicName="Guitar Center",
            siret="83994784300018",
        )
        product_thing = offers_factories.ThingProductFactory(name="Harry Potter")
        offer_for_thing = offers_factories.ThingOfferFactory(venue=venue_for_thing, product=product_thing)
        stock_for_thing = offers_factories.ThingStockFactory(offer=offer_for_thing, price=0)
        bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock_for_thing,
            dateCreated=(default_booking_date - timedelta(days=1)),
            token="ABCDEF",
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=pro, booking_period=(one_year_before_booking, one_year_after_booking)
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 3
        data_dicts = [dict(zip(headers, line)) for line in data]
        venues = [data_dict["Lieu"] for data_dict in data_dicts]
        assert sorted(venues) == [venue_for_thing.publicName, venue_for_book.publicName, venue_for_event.publicName]

    def test_should_return_only_booking_for_requested_venue(self, app: fixture):
        # Given
        pro_user = users_factories.ProFactory()
        user_offerer = offerers_factories.UserOffererFactory(user=pro_user)

        bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=user_offerer.offerer)
        booking_two = bookings_factories.BookingFactory(stock__offer__venue__managingOfferer=user_offerer.offerer)

        # When
        bookings_csv = booking_repository.get_export(
            user=pro_user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            venue_id=booking_two.venue.id,
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Nom de l’offre"] == booking_two.stock.offer.name

    def test_should_return_only_booking_for_requested_event_date(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        event_date = datetime(2020, 12, 24, 10, 30)
        expected_booking = bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=event_date, offer__venue__managingOfferer=user_offerer.offerer
            )
        )
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(offer__venue__managingOfferer=user_offerer.offerer)
        )
        bookings_factories.BookingFactory(
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer)
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            event_date=event_date.date(),
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Contremarque"] == expected_booking.token

    def should_consider_venue_locale_datetime_when_filtering_by_event_date(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        event_datetime = datetime(2020, 4, 21, 20, 00)

        offer_in_cayenne = offers_factories.OfferFactory(
            venue__postalCode="97300", venue__managingOfferer=user_offerer.offerer
        )
        cayenne_event_datetime = datetime(2020, 4, 22, 2, 0)
        stock_in_cayenne = offers_factories.EventStockFactory(
            offer=offer_in_cayenne, beginningDatetime=cayenne_event_datetime
        )
        cayenne_booking = bookings_factories.BookingFactory(stock=stock_in_cayenne)

        offer_in_mayotte = offers_factories.OfferFactory(
            venue__postalCode="97600", venue__managingOfferer=user_offerer.offerer
        )
        mayotte_event_datetime = datetime(2020, 4, 20, 22, 0)
        stock_in_mayotte = offers_factories.EventStockFactory(
            offer=offer_in_mayotte, beginningDatetime=mayotte_event_datetime
        )
        mayotte_booking = bookings_factories.BookingFactory(stock=stock_in_mayotte)

        # When
        bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
            event_date=event_datetime.date(),
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2
        data_dicts = [dict(zip(headers, line)) for line in data]
        tokens = [booking["Contremarque"] for booking in data_dicts]
        assert sorted(tokens) == sorted([cayenne_booking.token, mayotte_booking.token])

    def test_should_return_only_bookings_for_requested_booking_period(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        booking_beginning_period = datetime(2020, 12, 24, 10, 30).date()
        booking_ending_period = datetime(2020, 12, 26, 15, 00).date()
        expected_booking = bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 26, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )
        bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 29, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )
        bookings_factories.BookingFactory(
            dateCreated=datetime(2020, 12, 22, 15, 30),
            stock=offers_factories.ThingStockFactory(offer__venue__managingOfferer=user_offerer.offerer),
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(booking_beginning_period, booking_ending_period),
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 1
        data_dict = dict(zip(headers, data[0]))
        assert data_dict["Date et heure de réservation"] == str(
            expected_booking.dateCreated.astimezone(tz.gettz("Europe/Paris"))
        )

    def test_should_consider_venue_locale_datetime_when_filtering_by_booking_period(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        requested_booking_period_beginning = datetime(2020, 4, 21, 20, 00).date()
        requested_booking_period_ending = datetime(2020, 4, 22, 20, 00).date()

        offer_in_cayenne = offers_factories.OfferFactory(
            venue__postalCode="97300", venue__managingOfferer=user_offerer.offerer
        )
        cayenne_booking_datetime = datetime(2020, 4, 22, 2, 0)
        stock_in_cayenne = offers_factories.EventStockFactory(
            offer=offer_in_cayenne,
        )
        cayenne_booking = bookings_factories.BookingFactory(
            stock=stock_in_cayenne, dateCreated=cayenne_booking_datetime
        )

        offer_in_mayotte = offers_factories.OfferFactory(
            venue__postalCode="97600", venue__managingOfferer=user_offerer.offerer
        )
        mayotte_booking_datetime = datetime(2020, 4, 20, 23, 0)
        stock_in_mayotte = offers_factories.EventStockFactory(
            offer=offer_in_mayotte,
        )
        mayotte_booking = bookings_factories.BookingFactory(
            stock=stock_in_mayotte, dateCreated=mayotte_booking_datetime
        )

        # When
        bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(requested_booking_period_beginning, requested_booking_period_ending),
        )

        # Then
        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2
        data_dicts = [dict(zip(headers, line)) for line in data]
        tokens = [booking["Contremarque"] for booking in data_dicts]
        assert sorted(tokens) == sorted([cayenne_booking.token, mayotte_booking.token])

    def test_should_output_the_correct_offer_type_depending_wether_offer_educational_or_not(self, app: fixture):
        # Given
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        booking_1 = bookings_factories.BookingFactory(
            stock__offer__venue=venue,
        )

        booking_2 = bookings_factories.UsedBookingFactory(
            stock__offer__venue=venue,
        )

        # When
        beginning_period = datetime.fromisoformat("2021-10-15")
        ending_period = datetime.fromisoformat("2032-02-15")
        bookings_csv = booking_repository.get_export(
            user=pro,
            booking_period=(beginning_period, ending_period),
        )

        # Then
        expected_type = {
            booking_1.stock.offer.name: "offre grand public",
            booking_2.stock.offer.name: "offre grand public",
        }

        headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
        assert len(data) == 2
        data_dicts = [dict(zip(headers, line)) for line in data]
        for data_dict in data_dicts:
            offer_name = data_dict["Nom de l’offre"]
            assert data_dict["Type d'offre"] == expected_type[offer_name]

    def test_should_return_only_bookings_for_requested_offer_type(self, app: fixture):
        # Given
        user_offerer = offerers_factories.UserOffererFactory()
        bookings_factories.BookingFactory(
            dateCreated=default_booking_date,
            stock__offer__venue__managingOfferer=user_offerer.offerer,
        )

        # When
        individual_bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
        )
        all_bookings_csv = booking_repository.get_export(
            user=user_offerer.user,
            booking_period=(one_year_before_booking, one_year_after_booking),
        )

        # Then
        headers, *individual_bookings_data = csv.reader(StringIO(individual_bookings_csv), delimiter=";")
        assert len(individual_bookings_data) == 1
        individual_bookings_data_dict = dict(zip(headers, individual_bookings_data[0]))
        assert individual_bookings_data_dict["Type d'offre"] == "offre grand public"
        headers, *all_bookings_data = csv.reader(StringIO(all_bookings_csv), delimiter=";")
        assert len(all_bookings_data) == 1

    class BookingStatusInCsvReportTest:
        def test_should_output_the_correct_status_for_individual_bookings_before_cancellationLimitDate(self, app):
            # Given
            date_created = datetime.utcnow() - timedelta(hours=6)

            pro = users_factories.ProFactory()
            offerer = offerers_factories.OffererFactory()
            offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
            venue = offerers_factories.VenueFactory(managingOfferer=offerer)

            bookings_factories.CancelledBookingFactory(
                stock__offer__venue=venue,
                dateCreated=date_created,
            )
            bookings_factories.BookingFactory(
                stock__offer__venue=venue,
                dateCreated=date_created,
            )

            # When
            beginning_period = datetime.utcnow().date() - timedelta(days=1)
            ending_period = datetime.utcnow().date() + timedelta(days=360)
            bookings_csv = booking_repository.get_export(
                user=pro,
                booking_period=(beginning_period, ending_period),
            )

            # Then
            headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
            assert len(data) == 2
            pos_cm = headers.index("Statut de la contremarque")
            assert sorted([line[pos_cm] for line in data]) == ["annulé", "réservé"]

        def test_should_output_the_correct_status_for_individual_bookings_things_after_cancellationLimitDate(self, app):
            # Given
            date_created = datetime.utcnow() - timedelta(days=10)

            pro = users_factories.ProFactory()
            offerer = offerers_factories.OffererFactory()
            offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
            venue = offerers_factories.VenueFactory(managingOfferer=offerer)

            bookings_factories.CancelledBookingFactory(
                stock__offer__venue=venue,
                dateCreated=date_created,
            )
            bookings_factories.BookingFactory(
                stock__offer__venue=venue,
                dateCreated=date_created,
            )
            bookings_factories.UsedBookingFactory(
                stock__offer__venue=venue,
                dateCreated=date_created,
            )
            bookings_factories.UsedBookingFactory(
                status=BookingStatus.REIMBURSED,
                stock__offer__venue=venue,
                dateCreated=date_created,
            )

            # When
            beginning_period = datetime.utcnow().date() - timedelta(days=11)
            ending_period = datetime.utcnow().date() + timedelta(days=360)
            bookings_csv = booking_repository.get_export(
                user=pro,
                booking_period=(beginning_period, ending_period),
            )

            # Then
            headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
            assert len(data) == 4
            pos_cm = headers.index("Statut de la contremarque")
            assert sorted([line[pos_cm] for line in data]) == ["annulé", "remboursé", "réservé", "validé"]

        def test_should_output_the_correct_status_for_individual_bookings_events_after_cancellationLimitDate(self, app):
            # Given
            date_created = datetime.utcnow() - timedelta(days=10)
            event_date = datetime.utcnow() + timedelta(days=2)

            pro = users_factories.ProFactory()
            offerer = offerers_factories.OffererFactory()
            offerers_factories.UserOffererFactory(user=pro, offerer=offerer)
            venue = offerers_factories.VenueFactory(managingOfferer=offerer)
            event_stock = offers_factories.EventStockFactory(beginningDatetime=event_date, offer__venue=venue)
            bookings_factories.CancelledBookingFactory(
                stock=event_stock,
                dateCreated=date_created,
            )
            bookings_factories.BookingFactory(
                stock=event_stock,
                dateCreated=date_created,
            )
            bookings_factories.UsedBookingFactory(
                stock=event_stock,
                dateCreated=date_created,
            )
            bookings_factories.UsedBookingFactory(
                status=BookingStatus.REIMBURSED,
                stock=event_stock,
                dateCreated=date_created,
            )

            # When
            beginning_period = datetime.utcnow().date() - timedelta(days=11)
            ending_period = datetime.utcnow().date() + timedelta(days=360)
            bookings_csv = booking_repository.get_export(
                user=pro,
                booking_period=(beginning_period, ending_period),
            )

            # Then
            headers, *data = csv.reader(StringIO(bookings_csv), delimiter=";")
            assert len(data) == 4
            pos_cm = headers.index("Statut de la contremarque")
            assert sorted([line[pos_cm] for line in data]) == ["annulé", "confirmé", "remboursé", "validé"]


class GetExcelReportTest:
    def test_should_return_excel_export_according_to_booking_attributes(self):
        # Given
        beneficiary = users_factories.BeneficiaryGrant18Factory(
            email="beneficiary@example.com", firstName="Ron", lastName="Weasley"
        )
        pro = users_factories.ProFactory()
        offerer = offerers_factories.OffererFactory()
        offerers_factories.UserOffererFactory(user=pro, offerer=offerer)

        venue = offerers_factories.VenueFactory(managingOfferer=offerer)

        offer = offers_factories.ThingOfferFactory(venue=venue)
        stock = offers_factories.ThingStockFactory(offer=offer, price=0)
        booking_date = datetime(2020, 1, 1, 10, 0, 0) - timedelta(days=1)
        booking = bookings_factories.UsedBookingFactory(
            user=beneficiary,
            stock=stock,
            dateCreated=booking_date,
            token="ABCDEF",
            amount=12,
        )

        # When
        bookings_excel = booking_repository.get_export(
            user=pro,
            booking_period=(booking_date - timedelta(days=365), booking_date + timedelta(days=365)),
            export_type=BookingExportType.EXCEL,
        )
        wb = openpyxl.load_workbook(BytesIO(bookings_excel))
        sheet = wb.active

        # Then
        # Headers
        for i in range(1, len(HEADERS)):
            assert sheet.cell(row=1, column=i).value == HEADERS[i - 1]
        # Lieu
        assert sheet.cell(row=2, column=1).value == venue.name
        # Nom de l’offre
        assert sheet.cell(row=2, column=2).value == offer.name
        # Date de l'évènement
        assert sheet.cell(row=2, column=3).value == "None"
        # EAN
        assert sheet.cell(row=2, column=4).value == ((offer.extraData or {}).get("ean") or None)
        # Nom et prénom du bénéficiaire
        assert sheet.cell(row=2, column=5).value == " ".join((beneficiary.lastName, beneficiary.firstName))
        # Email du bénéficiaire
        assert sheet.cell(row=2, column=6).value == beneficiary.email
        # Téléphone du bénéficiaire
        assert sheet.cell(row=2, column=7).value == (beneficiary.phoneNumber or None)
        # Date et heure de réservation
        assert sheet.cell(row=2, column=8).value == str(booking.dateCreated.astimezone(tz.gettz("Europe/Paris")))
        # Date et heure de validation
        assert sheet.cell(row=2, column=9).value == str(booking.dateUsed.astimezone(tz.gettz("Europe/Paris")))
        # Contremarque
        assert sheet.cell(row=2, column=10).value == booking.token
        # Intitulé du tarif
        assert sheet.cell(row=2, column=11).value is None
        # Prix de la réservation
        assert sheet.cell(row=2, column=12).value == booking.amount
        # Statut de la contremarque
        assert sheet.cell(row=2, column=13).value == booking_repository.BOOKING_STATUS_LABELS[booking.status]
        # Date et heure de remboursement
        assert sheet.cell(row=2, column=14).value == "None"
        # Type d'offre
        assert sheet.cell(row=2, column=15).value == "offre grand public"
        # Code postal du bénéficiaire
        assert sheet.cell(row=2, column=16).value == beneficiary.postalCode
        # Duo
        assert sheet.cell(row=2, column=17).value == "Non"


class FindSoonToBeExpiredBookingsTest:
    def test_should_return_only_soon_to_be_expired_individual_bookings(self, app: fixture):
        # Given
        expired_creation_date = date.today() - timedelta(days=23)
        expired_creation_date = datetime.combine(expired_creation_date, time(12, 34, 17))
        non_expired_creation_date = date.today() - timedelta(days=24)
        non_expired_creation_date = datetime.combine(non_expired_creation_date, time(12, 34, 17))
        too_old_expired_creation_date = date.today() - timedelta(days=22)
        too_old_expired_creation_date = datetime.combine(too_old_expired_creation_date, time(12, 34, 17))

        expected_booking = bookings_factories.BookingFactory(
            dateCreated=expired_creation_date,
            stock__offer__subcategoryId=subcategories.SUPPORT_PHYSIQUE_FILM.id,
        )
        # offer type not expirable
        bookings_factories.BookingFactory(
            dateCreated=expired_creation_date,
            stock__offer__subcategoryId=subcategories.TELECHARGEMENT_LIVRE_AUDIO.id,
        )
        bookings_factories.BookingFactory(
            dateCreated=non_expired_creation_date,
            stock__offer__subcategoryId=subcategories.SUPPORT_PHYSIQUE_FILM.id,
        )
        bookings_factories.BookingFactory(
            dateCreated=too_old_expired_creation_date,
            stock__offer__subcategoryId=subcategories.SUPPORT_PHYSIQUE_FILM.id,
        )

        # When
        expired_bookings = booking_repository.find_soon_to_be_expiring_individual_bookings_ordered_by_user().all()

        # Then
        assert expired_bookings == [expected_booking]

    def test_should_return_only_soon_to_be_expired_bookings_books_case(self):
        soon_expired_creation_date = datetime.combine(date.today() - timedelta(days=5), time(12, 34, 17))
        too_old_creation_date = datetime.combine(date.today() - timedelta(days=6), time(12, 34, 17))
        non_expired_creation_date = datetime.combine(date.today() - timedelta(days=4), time(12, 34, 17))

        soon_expired_books_booking = bookings_factories.BookingFactory(
            dateCreated=soon_expired_creation_date,
            stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        )
        bookings_factories.BookingFactory(
            dateCreated=soon_expired_creation_date,
            stock__offer__subcategoryId=subcategories.SUPPORT_PHYSIQUE_FILM.id,
        )
        bookings_factories.BookingFactory(
            dateCreated=too_old_creation_date,
            stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        )
        bookings_factories.BookingFactory(
            dateCreated=non_expired_creation_date,
            stock__offer__subcategoryId=subcategories.LIVRE_PAPIER.id,
        )

        assert booking_repository.find_soon_to_be_expiring_individual_bookings_ordered_by_user().all() == [
            soon_expired_books_booking
        ]


class GetLegacyActiveBookingsQuantityForVenueTest:
    def test_return_bookings_quantity_for_venue(self):
        # Given
        booking = bookings_factories.BookingFactory(quantity=2)
        venue = booking.venue
        bookings_factories.BookingFactory(stock__offer__venue=venue)

        # When
        active_bookings_quantity = booking_repository.get_active_bookings_quantity_for_venue(venue.id)

        # Then
        assert active_bookings_quantity == 3

    def test_return_0_when_no_bookings_exist(self):
        # Given
        venue = offerers_factories.VenueFactory()

        # When
        active_bookings_quantity = booking_repository.get_active_bookings_quantity_for_venue(venue.id)

        # Then
        assert active_bookings_quantity == 0

    def test_excludes_confirmed_used_or_cancelled_bookings(self):
        # Given
        booking = bookings_factories.BookingFactory()
        venue = booking.venue
        bookings_factories.UsedBookingFactory(stock__offer__venue=venue)
        bookings_factories.CancelledBookingFactory(stock__offer__venue=venue)
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.BookingFactory(cancellation_limit_date=yesterday, quantity=2, stock__offer__venue=venue)

        # When
        active_bookings_quantity = booking_repository.get_active_bookings_quantity_for_venue(venue.id)

        # Then
        assert active_bookings_quantity == 1

    def test_excludes_other_venues_bookings(self):
        # Given
        booking = bookings_factories.BookingFactory()
        venue = booking.venue
        another_venue = offerers_factories.VenueFactory(managingOfferer=venue.managingOfferer)
        bookings_factories.BookingFactory(stock__offer__venue=another_venue)

        # When
        active_bookings_quantity = booking_repository.get_active_bookings_quantity_for_venue(venue.id)

        # Then
        assert active_bookings_quantity == 1


class GetLegacyValidatedBookingsQuantityForVenueTest:
    def test_return_used_bookings_quantity_for_venue(self):
        # Given
        booking = bookings_factories.UsedBookingFactory(quantity=2)
        venue = booking.venue
        bookings_factories.UsedBookingFactory(stock__offer__venue=venue)

        # When
        validated_bookings_quantity = booking_repository.get_validated_bookings_quantity_for_venue(venue.id)

        # Then
        assert validated_bookings_quantity == 3

    def test_return_confirmed_bookings_quantity_for_venue(self):
        # Given
        yesterday = datetime.utcnow() - timedelta(days=1)
        booking = bookings_factories.BookingFactory(cancellation_limit_date=yesterday, quantity=2)
        venue = booking.venue

        # When
        validated_bookings_quantity = booking_repository.get_validated_bookings_quantity_for_venue(venue.id)

        # Then
        assert validated_bookings_quantity == 2

    def test_return_0_when_no_bookings_exist(self):
        # Given
        venue = offerers_factories.VenueFactory()

        # When
        validated_bookings_quantity = booking_repository.get_validated_bookings_quantity_for_venue(venue.id)

        # Then
        assert validated_bookings_quantity == 0

    def test_excludes_unused_or_cancelled_bookings(self):
        # Given
        booking = bookings_factories.UsedBookingFactory()
        venue = booking.venue
        bookings_factories.BookingFactory(stock__offer__venue=venue)
        bookings_factories.CancelledBookingFactory(stock__offer__venue=venue)

        # When
        validated_bookings_quantity = booking_repository.get_validated_bookings_quantity_for_venue(venue.id)

        # Then
        assert validated_bookings_quantity == 1

    def test_excludes_other_venues_bookings(self):
        # Given
        booking = bookings_factories.UsedBookingFactory()
        venue = booking.venue
        another_venue = offerers_factories.VenueFactory(managingOfferer=venue.managingOfferer)
        bookings_factories.UsedBookingFactory(stock__offer__venue=another_venue)

        # When
        validated_bookings_quantity = booking_repository.get_validated_bookings_quantity_for_venue(venue.id)

        # Then
        assert validated_bookings_quantity == 1


class FindExpiringBookingsTest:
    def test_find_expired_bookings_before_and_after_enabling_feature_flag(self):
        now = datetime.utcnow()
        book_offer = offers_factories.OfferFactory(subcategoryId=subcategories.LIVRE_PAPIER.id)
        movie_offer = offers_factories.OfferFactory(subcategoryId=subcategories.SUPPORT_PHYSIQUE_FILM.id)

        _movie_booking_too_recent = bookings_factories.BookingFactory(
            stock__offer=movie_offer, dateCreated=now - timedelta(days=30)
        )
        movie_booking_old_enough = bookings_factories.BookingFactory(
            stock__offer=movie_offer, dateCreated=now - timedelta(days=31)
        )

        _book_booking_too_recent = bookings_factories.BookingFactory(
            stock__offer=book_offer, dateCreated=now - timedelta(days=10)
        )
        book_booking_old_enough = bookings_factories.BookingFactory(
            stock__offer=book_offer, dateCreated=now - timedelta(days=11)
        )

        bookings = set(booking_repository.find_expiring_individual_bookings_query())
        assert bookings == {book_booking_old_enough, movie_booking_old_enough}


def test_get_bookings_from_deposit():
    deposit1 = users_factories.DepositGrantFactory()
    deposit2 = users_factories.DepositGrantFactory()
    booking1 = bookings_factories.BookingFactory(deposit=deposit1)
    booking2 = bookings_factories.BookingFactory(deposit=deposit2)
    _cancelled_booking = bookings_factories.CancelledBookingFactory(deposit=deposit2)

    assert get_bookings_from_deposit(deposit1.id) == [booking1]
    assert get_bookings_from_deposit(deposit2.id) == [booking2]


class SoonExpiringBookingsTest:
    def test_get_soon_expiring_bookings(self):
        stock = offers_factories.ThingStockFactory()

        bookings_factories.UsedBookingFactory(stock=stock)
        bookings_factories.CancelledBookingFactory(stock=stock)
        booking = bookings_factories.BookingFactory(stock=stock)

        creation_date = datetime.utcnow() - timedelta(days=1)
        bookings_factories.BookingFactory(stock=stock, dateCreated=creation_date)

        remaining_days = (booking.expirationDate.date() - date.today()).days

        bookings = set(booking_repository.get_soon_expiring_bookings(remaining_days))
        assert {booking.id for booking in bookings} == {booking.id}

    def test_no_unexpected_queries(self):
        stocks = offers_factories.ThingStockFactory.create_batch(5)
        bookings = [bookings_factories.BookingFactory(stock=stock) for stock in stocks]

        remaining_days = (bookings[0].expirationDate.date() - date.today()).days
        with assert_no_duplicated_queries():
            list(booking_repository.get_soon_expiring_bookings(remaining_days))


class GetTomorrowEventOfferTest:
    def test_find_tomorrow_event_offer(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=tomorrow,
            )
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 1

    def should_not_select_given_before_tomorrow_booking_event(self):
        yesterday = datetime.utcnow() - timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=yesterday,
            )
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 0

    def should_not_select_given_after_tomorrow_booking_event(self):
        after_tomorrow = datetime.utcnow() + timedelta(days=2)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=after_tomorrow,
            )
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 0

    def should_not_select_given_not_booking_event(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(stock__beginningDatetime=tomorrow)

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 0

    def should_do_only_one_query(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=tomorrow,
            )
        )

        with assert_no_duplicated_queries():
            bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 1

    def should_not_select_digital_event(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=tomorrow,
                offer__url="http://digitaloffer.pass",
            )
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 0

    @pytest.mark.parametrize(
        "offer_url",
        [
            None,
            "",
        ],
    )
    def should_select_not_digital_event(self, offer_url):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=tomorrow,
                offer__url=offer_url,
            )
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 1

    def should_not_select_cancelled_booking(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        bookings_factories.BookingFactory(
            stock=offers_factories.EventStockFactory(
                beginningDatetime=tomorrow,
            ),
            status=BookingStatus.CANCELLED,
        )

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 0

    def should_select_several_bookings_given_one_stock_with_several_bookings(self):
        tomorrow = datetime.utcnow() + timedelta(days=1)
        stock = offers_factories.EventStockFactory(
            beginningDatetime=tomorrow,
        )
        bookings_factories.BookingFactory(stock=stock)
        bookings_factories.BookingFactory(stock=stock)

        bookings = booking_repository.find_individual_bookings_event_happening_tomorrow()

        assert len(bookings) == 2
